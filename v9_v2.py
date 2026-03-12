import psycopg2
import psycopg2.extras
import psycopg2.pool
from concurrent.futures import ThreadPoolExecutor

# ── Global connection pool & executor ────────────────────────
_db_pool: psycopg2.pool.ThreadedConnectionPool = None
_executor = ThreadPoolExecutor(max_workers=20)

"""
╔══════════════════════════════════════════════════════════════╗
║         DATA SCANNER BOT v9.0 — PHASE 1 UPGRADE            ║
║  IP/IBAN/Address Search | Bulk Search | Excel Export        ║
║  Admin Alerts | Watermark | IP Tracking | Auto-Ban          ║
║  Points System | Search History | Affiliate System         ║
╚══════════════════════════════════════════════════════════════╝
"""

import os
import re
import io
import csv
import json
import time
import signal
import shutil
import asyncio
import logging
import logging.handlers
import zipfile
import hashlib
import threading
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.error import BadRequest, Forbidden, RetryAfter, TelegramError
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    CallbackQueryHandler, ContextTypes, filters
)

# ════════════════════════════════════════════
#                   CONFIG
# ════════════════════════════════════════════
TOKEN       = os.environ.get("BOT_TOKEN")
ADMIN_IDS   = [int(x) for x in os.environ.get("ADMIN_IDS", "8589487706").split(",")]
FILES_DIR   = "data_files"
BACKUP_DIR  = "db_backups"
REFERRAL_CREDITS       = 3
MIN_KEYWORD_LEN        = 3
WHITELIST_MODE         = False
WHITELIST_IDS: set     = set()
BOT_START_TIME         = None
MAINTENANCE_MODE       = False
MAINTENANCE_MSG        = "🔧 *Bot is under maintenance.*\n\nPlease try again later."
MAX_RESULT_LINES       = 10000_000

# ── Phase 1 Config ────────────────────────────────────────
POINTS_PER_SEARCH      = 1       # نقطة لكل بحث
POINTS_PER_REFERRAL    = 10      # نقاط للإحالة
POINTS_PER_UPLOAD      = 50      # نقاط لرفع ملف
AUTO_BAN_THRESHOLD     = 20      # حظر تلقائي بعد X محاولة فاشلة
WATERMARK_TEXT         = "@DataScannerBot"
BULK_SEARCH_MAX        = 50      # أقصى عدد للبحث المتعدد
IP_TRACK_ENABLED       = True    # تتبع IP
SUPPORT_TICKET_CHANNEL = None    # channel id لتذاكر الدعم (اختياري)

# ════════════════════════════════════════════
#         PAYMENT SYSTEM CONFIG
# ════════════════════════════════════════════
PLAN_PRICES = {
    "basic":   {"usd": 5,  "egp": 250,  "label": "⭐ Basic"},
    "premium": {"usd": 10, "egp": 500,  "label": "💎 Premium"},
    "vip":     {"usd": 20, "egp": 1000, "label": "👑 VIP"},
}
PLAN_DURATIONS = {
    "1month": {"days": 30,  "label": "1 شهر / 1 Month",  "discount": 0},
    "3month": {"days": 90,  "label": "3 شهور / 3 Months", "discount": 10},
    "6month": {"days": 180, "label": "6 شهور / 6 Months", "discount": 20},
}

# Crypto wallets — غيّرهم لمحافظك
CRYPTO_WALLETS = {
    "USDT_TRC20": {
        "address": os.environ.get("WALLET_USDT_TRC20", "TYourTRC20WalletAddressHere"),
        "network": "TRC-20",
        "label": "💵 USDT (TRC-20)",
        "min_confirmations": 1,
    },
    "USDT_ERC20": {
        "address": os.environ.get("WALLET_USDT_ERC20", "0xYourERC20WalletAddressHere"),
        "network": "ERC-20",
        "label": "💵 USDT (ERC-20)",
        "min_confirmations": 6,
    },
    "BTC": {
        "address": os.environ.get("WALLET_BTC", "YourBTCWalletAddressHere"),
        "network": "Bitcoin",
        "label": "🪙 BTC (Bitcoin)",
        "min_confirmations": 1,
    },
}

# Vodafone Cash / InstaPay
MOBILE_PAYMENT = {
    "vodafone": {
        "number":    os.environ.get("VODAFONE_NUMBER", "01XXXXXXXXX"),
        "name":      os.environ.get("VODAFONE_NAME",   "Data Scanner"),
        "label":     "📱 Vodafone Cash",
        "enabled":   True,
    },
    "instapay": {
        "ipa":       os.environ.get("INSTAPAY_IPA",    "DataScanner@instapay"),
        "label":     "💳 InstaPay",
        "enabled":   True,
    },
    "fawry": {
        "code":      os.environ.get("FAWRY_CODE",      ""),
        "label":     "🏪 Fawry",
        "enabled":   False,
    },
}

PAYMENT_PENDING_TIMEOUT = 24  # ساعات قبل إلغاء الطلب تلقائياً

# ════════════════════════════════════════════
#       GAMIFICATION CONFIG
# ════════════════════════════════════════════
BADGES = {
    "newcomer":    {"label": "🆕 Newcomer",      "desc": "Joined the bot",             "threshold": 0},
    "searcher":    {"label": "🔍 Searcher",       "desc": "Made 10 searches",           "threshold": 10},
    "hunter":      {"label": "🏹 Hunter",         "desc": "Made 50 searches",           "threshold": 50},
    "elite":       {"label": "⚡ Elite Hunter",   "desc": "Made 200 searches",          "threshold": 200},
    "referrer":    {"label": "🤝 Referrer",       "desc": "Referred 3 users",           "threshold": 3},
    "vip_member":  {"label": "👑 VIP Member",     "desc": "Active VIP subscriber",      "threshold": 0},
    "data_pro":    {"label": "💎 Data Pro",       "desc": "1000 total points earned",   "threshold": 1000},
}

# ════════════════════════════════════════════
#       SCHEDULED IMPORTS CONFIG
# ════════════════════════════════════════════
SCHEDULED_IMPORT_INTERVAL = 3600   # ثانية (كل ساعة)
DEDUP_BATCH_SIZE           = 50000  # عدد السجلات في كل batch تنظيف

# ════════════════════════════════════════════
#       2FA CONFIG
# ════════════════════════════════════════════
TWO_FA_CODE_LENGTH = 6
TWO_FA_CODE_EXPIRY = 300   # ثانية (5 دقايق)
TWO_FA_ENABLED     = True  # تفعيل 2FA في البوت

# ════════════════════════════════════════════
#       URL IMPORT CONFIG
# ════════════════════════════════════════════
URL_IMPORT_ALLOWED = [".txt", ".csv", ".json", ".xlsx", ".xls", ".zip"]
URL_IMPORT_MAX_MB  = 200

# Supported languages
SUPPORTED_LANGS = {
    "en": "🇬🇧 English",
    "ar": "🇸🇦 العربية",
    "tr": "🇹🇷 Türkçe",
    "fa": "🇮🇷 فارسی",
    "fr": "🇫🇷 Français",
}

# RTL languages
RTL_LANGS = {"ar", "fa"}

# Coupon system config
COUPON_TYPES = {
    "percent":  "Percentage discount",
    "fixed":    "Fixed USD discount",
    "days":     "Free days added",
    "plan":     "Free plan upgrade",
}

# Auto-renewal config
AUTO_RENEW_DAYS_BEFORE = 3    # Alert user 3 days before expiry

# ════════════════════════════════════════════
#       DATA CATEGORIES
# ════════════════════════════════════════════
DATA_CATEGORIES = {
    "email_pass": {
        "label":   "📧 Email:Password",
        "pattern": r"[\w.+-]+@[\w-]+\.[a-zA-Z]{2,}:.+",
    },
    "url_creds": {
        "label":   "🌐 URL:User:Pass",
        "pattern": r"https?://[^\s|]+",
    },
    "phone": {
        "label":   "📱 Phone Numbers",
        "pattern": r"(\+?\d[\d\s\-]{8,}\d)",
    },
    "ip_data": {
        "label":   "🖥️ IP Addresses",
        "pattern": r"\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b",
    },
    "national_id": {
        "label":   "🪪 National IDs",
        "pattern": r"\b\d{14}\b",
    },
    "iban_data": {
        "label":   "🏦 IBAN/Bank",
        "pattern": r"\b[A-Z]{2}\d{2}[A-Z0-9]{4,30}\b",
    },
    "other": {
        "label":   "📄 Other",
        "pattern": None,
    },
}

os.makedirs(FILES_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

# ── Logging ───────────────────────────────────────────────
logging.basicConfig(format="%(asctime)s | %(levelname)s | %(message)s", level=logging.INFO)
log = logging.getLogger(__name__)
_log_dir = os.path.join(os.path.dirname(__file__), "logs")
os.makedirs(_log_dir, exist_ok=True)
_fh = logging.handlers.RotatingFileHandler(
    os.path.join(_log_dir, "bot_errors.log"),
    maxBytes=5 * 1024 * 1024,
    backupCount=3,
    encoding="utf-8"
)
_fh.setLevel(logging.WARNING)
_fh.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
log.addHandler(_fh)
logging.getLogger("telegram").addHandler(_fh)

SEARCH_TIMEOUT = 180

_last_search_time: dict = {}
_rate_limit_lock  = threading.Lock()

_cb_store: dict = {}
_user_search_locks: dict = {}
_CB_TTL = 3600

# ── User cache (reduces DB hits dramatically) ─────────────
_user_cache: dict = {}
_USER_CACHE_TTL = 30  # seconds — cache user row for 30s

# ════════════════════════════════════════════
#           DATABASE CONNECTION
# ════════════════════════════════════════════
DB_CONFIG = {
    "dbname":   os.environ.get("DB_NAME",     "scanner"),
    "user":     os.environ.get("DB_USER",     "postgres"),
    "password": os.environ.get("DB_PASSWORD", "123456"),
    "host":     os.environ.get("DB_HOST",     "127.0.0.1"),
    "port":     os.environ.get("DB_PORT",     "5432"),
}

def init_pool():
    """Initialize the global connection pool."""
    global _db_pool
    _db_pool = psycopg2.pool.ThreadedConnectionPool(
        minconn=3,
        maxconn=30,
        connect_timeout=10,
        **DB_CONFIG
    )
    log.info("✅ PostgreSQL connection pool initialized (3–30 connections)")

def get_db():
    """Get a connection from the pool."""
    return _db_pool.getconn()

def release_db(conn):
    """Return a connection back to the pool."""
    try:
        _db_pool.putconn(conn)
    except Exception:
        pass

class _PoolConn:
    """Context manager: auto-acquire and release a pool connection."""
    def __init__(self):
        self.conn = None
    def __enter__(self):
        self.conn = get_db()
        return self.conn
    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type:
            try:
                self.conn.rollback()
            except Exception:
                pass
        release_db(self.conn)
        return False

def pool_conn():
    return _PoolConn()


def get_user_cached(uid: int):
    """Get user row from cache, fall back to DB if expired."""
    now = time.monotonic()
    if uid in _user_cache:
        row, ts = _user_cache[uid]
        if now - ts < _USER_CACHE_TTL:
            return row
    row = get_user(uid)
    _user_cache[uid] = (row, now)
    return row

def invalidate_user_cache(uid: int):
    """Call this after any UPDATE to users table."""
    _user_cache.pop(uid, None)

def _get_user_lock(uid: int) -> asyncio.Lock:
    if uid not in _user_search_locks:
        _user_search_locks[uid] = asyncio.Lock()
    return _user_search_locks[uid]

def _cb_put(data: str) -> str:
    key = "cb_" + hashlib.md5(data.encode()).hexdigest()[:12]
    _cb_store[key] = (data, time.monotonic())
    if len(_cb_store) > 500:
        now_t = time.monotonic()
        expired = [k for k, (_, ts) in list(_cb_store.items()) if now_t - ts > _CB_TTL]
        for k in expired:
            _cb_store.pop(k, None)
    return key

def _cb_get(key: str) -> str:
    entry = _cb_store.get(key)
    if entry:
        return entry[0]
    return key

# ════════════════════════════════════════════
#           LANGUAGE STRINGS (i18n)
# ════════════════════════════════════════════
STRINGS = {
    "en": {
        "hello": "Hello",
        "plan": "Plan",
        "daily_left": "Daily searches left",
        "credits": "Credits",
        "db_records": "Database Records",
        "nameid_records": "Name/ID Records",
        "menu_hint": "Use the menu below to search the database.",
        "btn_search": "🔍 Search Database",
        "btn_nameid": "🪪 Name / National ID Search",
        "btn_account": "📊 My Account",
        "btn_plans": "💳 Plans",
        "btn_help": "ℹ️ Help",
        "btn_subscribe": "📋 Subscribe Request",
        "btn_language": "🌐 Language / اللغة",
        "banned": "🚫 *Your account has been banned.*\n\nContact the admin if you believe this is a mistake.",
        "no_searches": "❌ *No searches remaining.*\n\nUpgrade your plan or buy credits.",
        "choose_lang": "🌐 *Choose your language:*",
        "lang_set": "✅ Language set to *English*.",
        "sub_req_title": "📋 *Subscription Request*",
        "sub_req_prompt": "Choose the plan you want to subscribe to:",
        "sub_req_sent": "✅ *Request Sent!*\n\n📦 Plan: *{tier}*\n🔢 Request ID: `#{req_id}`\n\nWe'll notify you once approved.",
        "sub_req_exists": "⏳ You already have a pending request (`#{req_id}`).\n\nPlease wait for admin review.",
        "sub_approved_user": "✅ *Subscription Approved!*\n\nYour account has been upgraded to *{tier}*.\n\nPress /start to refresh.",
        "sub_rejected_user": "❌ *Subscription request rejected.*\n\nContact admin for more info.",
        "account_title": "📊 *My Account*",
        "plans_contact": "📩 Contact @yut3ev to upgrade.",
        "help_title": "ℹ️ *How to Use Data Scanner Bot*",
    },
    "ar": {
        "hello": "مرحباً",
        "plan": "الباقة",
        "daily_left": "البحوث المتبقية اليوم",
        "credits": "الرصيد",
        "db_records": "سجلات قاعدة البيانات",
        "nameid_records": "سجلات الاسم/الرقم القومي",
        "menu_hint": "استخدم القائمة أدناه للبحث في قاعدة البيانات.",
        "btn_search": "🔍 بحث في قاعدة البيانات",
        "btn_nameid": "🪪 بحث بالاسم / الرقم القومي",
        "btn_account": "📊 حسابي",
        "btn_plans": "💳 الباقات",
        "btn_help": "ℹ️ المساعدة",
        "btn_subscribe": "📋 طلب اشتراك",
        "btn_language": "🌐 Language / اللغة",
        "banned": "🚫 تم حظرك.",
        "no_searches": "❌ *لا توجد بحوث متبقية.*\n\nقم بترقية باقتك أو شراء رصيد.",
        "choose_lang": "🌐 *اختر لغتك:*",
        "lang_set": "✅ تم تعيين اللغة إلى *العربية*.",
        "sub_req_title": "📋 *طلب اشتراك*",
        "sub_req_prompt": "اختر الباقة التي تريد الاشتراك فيها:",
        "sub_req_sent": "✅ *تم إرسال الطلب!*\n\n📦 الباقة: *{tier}*\n🔢 رقم الطلب: `#{req_id}`\n\nسيتم إخطارك عند الموافقة.",
        "sub_req_exists": "⏳ لديك طلب قيد المراجعة بالفعل (`#{req_id}`).\n\nانتظر رد الأدمن.",
        "sub_approved_user": "✅ *تمت الموافقة على الاشتراك!*\n\nتمت ترقية حسابك إلى *{tier}*.\n\nاضغط /start لتحديث لوحتك.",
        "sub_rejected_user": "❌ *تم رفض طلب الاشتراك.*\n\nتواصل مع الأدمن لمزيد من المعلومات.",
        "account_title": "📊 *حسابي*",
        "plans_contact": "📩 تواصل مع @yut3ev للترقية.",
        "help_title": "ℹ️ *كيفية استخدام بوت DATA SCANNER*",
    },
    "tr": {
        "hello": "Merhaba",
        "plan": "Plan",
        "daily_left": "Bugün kalan aramalar",
        "credits": "Kredi",
        "db_records": "Veritabanı kayıtları",
        "nameid_records": "Ad/Kimlik kayıtları",
        "menu_hint": "Veritabanında arama yapmak için aşağıdaki menüyü kullanın.",
        "btn_search": "🔍 Veritabanında Ara",
        "btn_nameid": "🪪 Ad / Kimlik No ile Ara",
        "btn_account": "📊 Hesabım",
        "btn_plans": "💳 Planlar",
        "btn_help": "ℹ️ Yardım",
        "btn_subscribe": "📋 Abonelik Talebi",
        "btn_language": "🌐 Dil / Language",
        "banned": "🚫 Hesabınız engellendi.",
        "no_searches": "❌ *Arama hakkınız kalmadı.*\n\nPlanınızı yükseltin veya kredi satın alın.",
        "choose_lang": "🌐 *Dilinizi seçin:*",
        "lang_set": "✅ Dil *Türkçe* olarak ayarlandı.",
        "sub_req_title": "📋 *Abonelik Talebi*",
        "sub_req_prompt": "Abone olmak istediğiniz planı seçin:",
        "sub_req_sent": "✅ *Talep gönderildi!*\n\n📦 Plan: *{tier}*\n🔢 Talep No: `#{req_id}`\n\nOnaylandığında bildirim alacaksınız.",
        "sub_req_exists": "⏳ Zaten bekleyen bir talebiniz var (`#{req_id}`).\n\nAdmin yanıtını bekleyin.",
        "sub_approved_user": "✅ *Abonelik onaylandı!*\n\nHesabınız *{tier}* planına yükseltildi.\n\nGüncellemek için /start'a basın.",
        "sub_rejected_user": "❌ *Abonelik talebi reddedildi.*\n\nDaha fazla bilgi için adminle iletişime geçin.",
        "account_title": "📊 *Hesabım*",
        "plans_contact": "📩 Yükseltme için @yut3ev ile iletişime geçin.",
        "help_title": "ℹ️ *DATA SCANNER Bot Kullanım Kılavuzu*",
    },
    "fa": {
        "hello": "سلام",
        "plan": "پلن",
        "daily_left": "جستجوهای باقی‌مانده امروز",
        "credits": "اعتبار",
        "db_records": "رکوردهای پایگاه داده",
        "nameid_records": "رکوردهای نام/کد ملی",
        "menu_hint": "از منوی زیر برای جستجو در پایگاه داده استفاده کنید.",
        "btn_search": "🔍 جستجو در پایگاه داده",
        "btn_nameid": "🪪 جستجو با نام / کد ملی",
        "btn_account": "📊 حساب من",
        "btn_plans": "💳 پلن‌ها",
        "btn_help": "ℹ️ راهنما",
        "btn_subscribe": "📋 درخواست اشتراک",
        "btn_language": "🌐 زبان / Language",
        "banned": "🚫 حساب شما مسدود شده است.",
        "no_searches": "❌ *جستجوی باقی‌مانده‌ای ندارید.*\n\nپلن خود را ارتقا دهید یا اعتبار بخرید.",
        "choose_lang": "🌐 *زبان خود را انتخاب کنید:*",
        "lang_set": "✅ زبان به *فارسی* تنظیم شد.",
        "sub_req_title": "📋 *درخواست اشتراک*",
        "sub_req_prompt": "پلن مورد نظر خود را انتخاب کنید:",
        "sub_req_sent": "✅ *درخواست ارسال شد!*\n\n📦 پلن: *{tier}*\n🔢 شماره درخواست: `#{req_id}`\n\nپس از تأیید اطلاع‌رسانی می‌شوید.",
        "sub_req_exists": "⏳ یک درخواست در انتظار دارید (`#{req_id}`).\n\nمنتظر پاسخ ادمین باشید.",
        "sub_approved_user": "✅ *اشتراک تأیید شد!*\n\nحساب شما به *{tier}* ارتقا یافت.\n\nبرای بروزرسانی /start را بزنید.",
        "sub_rejected_user": "❌ *درخواست اشتراک رد شد.*\n\nبرای اطلاعات بیشتر با ادمین تماس بگیرید.",
        "account_title": "📊 *حساب من*",
        "plans_contact": "📩 برای ارتقا با @yut3ev تماس بگیرید.",
        "help_title": "ℹ️ *راهنمای استفاده از ربات DATA SCANNER*",
    },
    "fr": {
        "hello": "Bonjour",
        "plan": "Plan",
        "daily_left": "Recherches restantes aujourd'hui",
        "credits": "Crédits",
        "db_records": "Enregistrements DB",
        "nameid_records": "Enregistrements Nom/ID",
        "menu_hint": "Utilisez le menu ci-dessous pour rechercher dans la base de données.",
        "btn_search": "🔍 Rechercher dans la DB",
        "btn_nameid": "🪪 Chercher par Nom / ID",
        "btn_account": "📊 Mon Compte",
        "btn_plans": "💳 Plans",
        "btn_help": "ℹ️ Aide",
        "btn_subscribe": "📋 Demande d'abonnement",
        "btn_language": "🌐 Langue / Language",
        "banned": "🚫 Votre compte a été banni.",
        "no_searches": "❌ *Plus de recherches disponibles.*\n\nAméliorez votre plan ou achetez des crédits.",
        "choose_lang": "🌐 *Choisissez votre langue:*",
        "lang_set": "✅ Langue définie sur *Français*.",
        "sub_req_title": "📋 *Demande d'abonnement*",
        "sub_req_prompt": "Choisissez le plan souhaité:",
        "sub_req_sent": "✅ *Demande envoyée!*\n\n📦 Plan: *{tier}*\n🔢 N° Demande: `#{req_id}`\n\nVous serez notifié à l'approbation.",
        "sub_req_exists": "⏳ Vous avez déjà une demande en attente (`#{req_id}`).\n\nAttendez la réponse de l'admin.",
        "sub_approved_user": "✅ *Abonnement approuvé!*\n\nVotre compte a été mis à niveau vers *{tier}*.\n\nAppuyez sur /start pour rafraîchir.",
        "sub_rejected_user": "❌ *Demande d'abonnement rejetée.*\n\nContactez l'admin pour plus d'informations.",
        "account_title": "📊 *Mon Compte*",
        "plans_contact": "📩 Contactez @yut3ev pour une mise à niveau.",
        "help_title": "ℹ️ *Guide d'utilisation du bot DATA SCANNER*",
    },
}

def get_lang(uid: int) -> str:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT lang FROM users WHERE user_id=%s", (uid,))
                row = cur.fetchone()
        return row[0] if row and row[0] else "en"
    except Exception:
        return "en"

def s(uid: int, key: str) -> str:
    lang = get_lang(uid)
    return STRINGS.get(lang, STRINGS["en"]).get(key, STRINGS["en"].get(key, key))

TIERS = {
    "free":    {"label": "🆓 Free",    "daily": 0,      "max_results": 0,       "full_scan": False},
    "basic":   {"label": "⭐ Basic",   "daily": 10,     "max_results": 200,     "full_scan": False},
    "premium": {"label": "💎 Premium", "daily": 15,     "max_results": 1000,    "full_scan": True},
    "vip":     {"label": "👑 VIP",     "daily": 100000, "max_results": 1000000, "full_scan": True},
}

NAMEID_TIERS = {
    "free":    {"daily_nameid": 0,  "max_nameid": 0},
    "basic":   {"daily_nameid": 2,  "max_nameid": 100},
    "premium": {"daily_nameid": 5,  "max_nameid": 150},
    "vip":     {"daily_nameid": 100, "max_nameid": 2000},
}

# ════════════════════════════════════════════
#                  DATABASE INIT
# ════════════════════════════════════════════
def init_db():
    conn = get_db()
    cur  = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id BIGINT PRIMARY KEY,
            username TEXT,
            full_name TEXT,
            tier TEXT DEFAULT 'free',
            daily_limit INTEGER DEFAULT 5,
            credits INTEGER DEFAULT 0,
            is_banned INTEGER DEFAULT 0,
            expires_at TEXT,
            joined_at TEXT,
            lang TEXT DEFAULT 'en',
            referred_by BIGINT DEFAULT NULL,
            referral_count INTEGER DEFAULT 0,
            last_search_at TEXT DEFAULT NULL,
            daily_nameid_limit INTEGER DEFAULT 0,
            frozen_until TEXT DEFAULT NULL,
            updated_at TEXT DEFAULT NULL,
            last_search_type TEXT DEFAULT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS data_index (
            id BIGSERIAL PRIMARY KEY,
            line TEXT NOT NULL,
            source TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS name_id_index (
            id BIGSERIAL PRIMARY KEY,
            full_name TEXT NOT NULL,
            national_id TEXT NOT NULL,
            source TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS uploaded_files (
            id SERIAL PRIMARY KEY,
            saved_name TEXT,
            original_name TEXT,
            size_bytes BIGINT,
            records INTEGER DEFAULT 0,
            uploaded_by BIGINT,
            uploaded_at TEXT,
            file_md5 TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS search_logs (
            user_id BIGINT,
            keyword TEXT,
            category TEXT,
            results INTEGER,
            timestamp TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS sub_history (
            user_id BIGINT,
            tier TEXT,
            amount INTEGER,
            admin_id BIGINT,
            timestamp TEXT,
            referral_source BIGINT DEFAULT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS sub_requests (
            id SERIAL PRIMARY KEY,
            user_id BIGINT,
            username TEXT,
            full_name TEXT,
            requested_tier TEXT,
            status TEXT DEFAULT 'pending',
            timestamp TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS admin_op_logs (
            id SERIAL PRIMARY KEY,
            admin_id BIGINT,
            action TEXT,
            target TEXT,
            details TEXT,
            timestamp TEXT
        )
    """)

    # ── Phase 1 Tables ────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS ip_logs (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT,
            ip_address TEXT,
            action TEXT,
            timestamp TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS failed_attempts (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT,
            reason TEXT,
            timestamp TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_points (
            user_id BIGINT PRIMARY KEY,
            points INTEGER DEFAULT 0,
            total_earned INTEGER DEFAULT 0,
            last_updated TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_favorites (
            id SERIAL PRIMARY KEY,
            user_id BIGINT,
            keyword TEXT,
            stype TEXT,
            label TEXT,
            created_at TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS support_tickets (
            id SERIAL PRIMARY KEY,
            user_id BIGINT,
            username TEXT,
            full_name TEXT,
            subject TEXT,
            message TEXT,
            status TEXT DEFAULT 'open',
            admin_reply TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS bulk_search_jobs (
            id SERIAL PRIMARY KEY,
            user_id BIGINT,
            keywords TEXT,
            stype TEXT,
            status TEXT DEFAULT 'pending',
            results TEXT,
            created_at TEXT,
            completed_at TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS affiliate_links (
            id SERIAL PRIMARY KEY,
            user_id BIGINT,
            code TEXT UNIQUE,
            clicks INTEGER DEFAULT 0,
            conversions INTEGER DEFAULT 0,
            earnings INTEGER DEFAULT 0,
            created_at TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS notifications (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT,
            message TEXT,
            is_read INTEGER DEFAULT 0,
            created_at TEXT
        )
    """)

    # ── Phase 1 Indexes ───────────────────────────────────
    cur.execute("CREATE INDEX IF NOT EXISTS idx_ip_logs_uid      ON ip_logs(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_ip_logs_ip       ON ip_logs(ip_address)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_failed_uid       ON failed_attempts(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_favorites_uid    ON user_favorites(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_tickets_uid      ON support_tickets(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_tickets_status   ON support_tickets(status)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_affiliate_code   ON affiliate_links(code)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_notif_uid        ON notifications(user_id)")

    # ── Payment System Tables ─────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS payment_orders (
            id SERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL,
            plan TEXT NOT NULL,
            duration TEXT NOT NULL,
            method TEXT NOT NULL,
            amount_usd REAL NOT NULL,
            amount_egp REAL,
            currency TEXT DEFAULT 'USD',
            status TEXT DEFAULT 'pending',
            tx_id TEXT,
            screenshot_file_id TEXT,
            admin_id BIGINT,
            notes TEXT,
            created_at TEXT,
            expires_at TEXT,
            confirmed_at TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS invoices (
            id SERIAL PRIMARY KEY,
            order_id INTEGER REFERENCES payment_orders(id),
            user_id BIGINT,
            plan TEXT,
            duration TEXT,
            amount REAL,
            currency TEXT,
            method TEXT,
            status TEXT,
            issued_at TEXT,
            paid_at TEXT
        )
    """)

    cur.execute("CREATE INDEX IF NOT EXISTS idx_orders_uid     ON payment_orders(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_orders_status  ON payment_orders(status)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_orders_tx      ON payment_orders(tx_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_invoices_uid   ON invoices(user_id)")

    # ── Gamification & Extras Tables ────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_badges (
            id SERIAL PRIMARY KEY,
            user_id BIGINT,
            badge_key TEXT,
            awarded_at TEXT,
            UNIQUE(user_id, badge_key)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS scheduled_imports (
            id SERIAL PRIMARY KEY,
            url TEXT NOT NULL,
            label TEXT,
            stype TEXT DEFAULT 'auto',
            frequency_hours INTEGER DEFAULT 24,
            last_run TEXT,
            next_run TEXT,
            status TEXT DEFAULT 'active',
            records_added INTEGER DEFAULT 0,
            created_by BIGINT,
            created_at TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS daily_stats (
            id SERIAL PRIMARY KEY,
            stat_date TEXT UNIQUE,
            new_users INTEGER DEFAULT 0,
            total_searches INTEGER DEFAULT 0,
            revenue_usd REAL DEFAULT 0,
            active_users INTEGER DEFAULT 0,
            new_subs INTEGER DEFAULT 0
        )
    """)

    cur.execute("CREATE INDEX IF NOT EXISTS idx_badges_uid    ON user_badges(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_daily_stats   ON daily_stats(stat_date)")

    # ── Coupon System ─────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS coupons (
            id SERIAL PRIMARY KEY,
            code TEXT UNIQUE NOT NULL,
            type TEXT NOT NULL,
            value REAL NOT NULL,
            max_uses INTEGER DEFAULT 1,
            used_count INTEGER DEFAULT 0,
            valid_from TEXT,
            valid_until TEXT,
            plan_restriction TEXT,
            created_by BIGINT,
            created_at TEXT,
            is_active INTEGER DEFAULT 1
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS coupon_uses (
            id SERIAL PRIMARY KEY,
            coupon_id INTEGER REFERENCES coupons(id),
            user_id BIGINT,
            order_id INTEGER,
            used_at TEXT
        )
    """)

    # ── Auto-Renewal ──────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS auto_renewal (
            user_id BIGINT PRIMARY KEY,
            enabled INTEGER DEFAULT 0,
            preferred_method TEXT,
            preferred_plan TEXT,
            preferred_duration TEXT,
            last_renewed TEXT,
            created_at TEXT
        )
    """)

    cur.execute("CREATE INDEX IF NOT EXISTS idx_coupons_code ON coupons(code)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_coupon_uses_uid ON coupon_uses(user_id)")

    # User activity log
    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_activity_log (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT,
            action TEXT,
            details TEXT,
            timestamp TEXT
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_activity_uid  ON user_activity_log(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_activity_ts   ON user_activity_log(timestamp)")

    # Indexes
    cur.execute("CREATE INDEX IF NOT EXISTS idx_search_logs_uid  ON search_logs(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_search_logs_ts   ON search_logs(timestamp)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_search_logs_cat  ON search_logs(category)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_users_tier       ON users(tier)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_users_banned     ON users(is_banned)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_users_expires    ON users(expires_at)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_uploaded_ts      ON uploaded_files(uploaded_at)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_data_line        ON data_index USING gin(to_tsvector('simple', line))")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_nameid_name      ON name_id_index(full_name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_nameid_id        ON name_id_index(national_id)")

    conn.commit()
    cur.close()
    release_db(conn)

# ════════════════════════════════════════════
#                   HELPERS
# ════════════════════════════════════════════
def get_user(uid):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM users WHERE user_id=%s", (uid,))
                row = cur.fetchone()
        return row
    except Exception as e:
        log.error(f"get_user error: {e}")
        return None

def ensure_user(uid, username, full_name):
    now_iso = datetime.utcnow().isoformat()
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM users WHERE user_id=%s", (uid,))
                exists = cur.fetchone()
                if not exists:
                    cur.execute(
                        """INSERT INTO users
                           (user_id, username, full_name, tier, daily_limit, credits,
                            is_banned, expires_at, joined_at, lang, referred_by,
                            referral_count, last_search_at, daily_nameid_limit, updated_at)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (uid, username or "", full_name or "User", "free", 5, 0, 0,
                         None, now_iso, "en", None, 0, None,
                         NAMEID_TIERS["free"]["daily_nameid"], now_iso)
                    )
            conn.commit()
        invalidate_user_cache(uid)
    except Exception as e:
        log.error(f"ensure_user error: {e}")

def is_admin(uid):
    return uid in ADMIN_IDS

def is_banned(uid):
    u = get_user_cached(uid)
    if not u:
        return False
    # index 6 = is_banned, index 14 = frozen_until
    if u[6]:
        return True
    frozen_until = u[14] if len(u) > 14 else None
    if frozen_until:
        try:
            if datetime.fromisoformat(str(frozen_until)) > datetime.utcnow():
                return True
        except Exception:
            pass
    return False

def get_tier(uid):
    u = get_user_cached(uid)
    return u[3] if u else "free"

def _check_and_expire(uid: int):
    u = get_user_cached(uid)
    if not u or not u[7]:
        return
    try:
        if datetime.fromisoformat(str(u[7])) < datetime.utcnow():
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "UPDATE users SET tier='free', daily_limit=0, daily_nameid_limit=0, expires_at=NULL WHERE user_id=%s",
                        (uid,)
                    )
                conn.commit()
            log.info(f"⏰ User {uid} subscription expired — downgraded to free.")
    except Exception:
        pass

def can_search(uid):
    if is_admin(uid):
        return True
    _check_and_expire(uid)
    u = get_user_cached(uid)
    if not u:
        return False
    if u[3] in ("premium", "vip"):
        return True
    return u[4] > 0 or u[5] > 0

def deduct(uid):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT daily_limit, credits FROM users WHERE user_id=%s", (uid,))
                u = cur.fetchone()
                if u:
                    if u[0] > 0:
                        cur.execute("UPDATE users SET daily_limit=daily_limit-1 WHERE user_id=%s", (uid,))
                    elif u[1] > 0:
                        cur.execute("UPDATE users SET credits=credits-1 WHERE user_id=%s", (uid,))
            conn.commit()
        invalidate_user_cache(uid)
    except Exception as e:
        log.error(f"deduct error: {e}")

def log_admin_op(admin_id: int, action: str, target: str, details: str = ""):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO admin_op_logs (admin_id, action, target, details, timestamp) VALUES (%s,%s,%s,%s,%s)",
                    (admin_id, action, target, details, datetime.utcnow().isoformat())
                )
            conn.commit()
    except Exception as e:
        log.error(f"log_admin_op error: {e}")

# ── Spam protection ───────────────────────────────────────
SEARCH_COOLDOWN_SECS = 5

def is_search_spamming(uid: int) -> bool:
    if is_admin(uid):
        return False
    with _rate_limit_lock:
        last = _last_search_time.get(uid)
    if last is None:
        return False
    return (time.monotonic() - last) < SEARCH_COOLDOWN_SECS

def mark_search_time(uid: int):
    with _rate_limit_lock:
        _last_search_time[uid] = time.monotonic()
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE users SET last_search_at=%s WHERE user_id=%s",
                            (datetime.utcnow().isoformat(), uid))
            conn.commit()
        invalidate_user_cache(uid)
    except Exception as e:
        log.error(f"mark_search_time error: {e}")

# ── Name/ID quota helpers ─────────────────────────────────
def can_search_nameid(uid: int) -> bool:
    if is_admin(uid):
        return True
    u = get_user_cached(uid)
    if not u:
        return False
    tier = u[3]
    nt = NAMEID_TIERS.get(tier, NAMEID_TIERS["free"])
    if nt["daily_nameid"] >= 100000:
        return True
    return u[13] > 0 if len(u) > 13 else False

def deduct_nameid(uid: int):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE users SET daily_nameid_limit=GREATEST(0, daily_nameid_limit-1) WHERE user_id=%s",
                    (uid,)
                )
            conn.commit()
        invalidate_user_cache(uid)
    except Exception as e:
        log.error(f"deduct_nameid error: {e}")

def get_nameid_limit(uid: int) -> int:
    if is_admin(uid):
        return 10_000_000
    u = get_user_cached(uid)
    tier = u[3] if u else "free"
    return NAMEID_TIERS.get(tier, NAMEID_TIERS["free"])["max_nameid"]

# ── Referral helpers ──────────────────────────────────────
def process_referral(new_uid: int, ref_uid: int):
    if new_uid == ref_uid:
        return
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT referred_by FROM users WHERE user_id=%s", (new_uid,))
                already = cur.fetchone()
                if already and already[0]:
                    return
                cur.execute("UPDATE users SET referred_by=%s WHERE user_id=%s", (ref_uid, new_uid))
                cur.execute(
                    "UPDATE users SET credits=credits+%s, referral_count=referral_count+1 WHERE user_id=%s",
                    (REFERRAL_CREDITS, ref_uid)
                )
            conn.commit()
    except Exception as e:
        log.error(f"process_referral error: {e}")

def get_referral_stats(uid: int):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT referral_count FROM users WHERE user_id=%s", (uid,))
                row = cur.fetchone()
        return row[0] if row else 0
    except Exception:
        return 0

# ── DB Backup (file-based not applicable for PG, just log) ──
def backup_db():
    ts   = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(BACKUP_DIR, f"pg_backup_note_{ts}.txt")
    with open(dest, "w") as f:
        f.write(f"PostgreSQL backup marker — {ts}\n")
        f.write("Run: pg_dump -U postgres scanner > backup.sql\n")
    backups = sorted([f for f in os.listdir(BACKUP_DIR) if f.endswith(".txt")], reverse=True)
    for old in backups[7:]:
        os.remove(os.path.join(BACKUP_DIR, old))
    return dest

# ── Auto daily reset ──────────────────────────────────────
def do_daily_reset():
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                for tn, td in TIERS.items():
                    cur.execute("UPDATE users SET daily_limit=%s WHERE tier=%s", (td["daily"], tn))
                for tn, nd in NAMEID_TIERS.items():
                    cur.execute("UPDATE users SET daily_nameid_limit=%s WHERE tier=%s", (nd["daily_nameid"], tn))
            conn.commit()
    except Exception as e:
        log.error(f"do_daily_reset error: {e}")

def log_search(uid, keyword, category, count):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO search_logs VALUES (%s,%s,%s,%s,%s)",
                    (uid, keyword, category, count, datetime.utcnow().isoformat())
                )
                if category and not category.startswith("nameid_"):
                    cur.execute(
                        "UPDATE users SET last_search_type=%s WHERE user_id=%s",
                        (category, uid)
                    )
            conn.commit()
    except Exception as e:
        log.error(f"log_search error: {e}")

# ════════════════════════════════════════════
#          RESULT COUNTER
# ════════════════════════════════════════════
def escape_like(value: str) -> str:
    return value.replace("\\", "\\\\").replace("%", r"\%").replace("_", r"\_")

def count_matches_fast(keyword: str, stype: str) -> int:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                # Use GIN full-text index for keywords >= 3 chars (fast), ILIKE for short ones
                if len(keyword) >= 3:
                    cur.execute(
                        "SELECT COUNT(*) FROM data_index WHERE to_tsvector('simple', line) @@ plainto_tsquery('simple', %s)",
                        (keyword,)
                    )
                else:
                    cur.execute(
                        "SELECT COUNT(*) FROM data_index WHERE line ILIKE %s",
                        (f"%{keyword}%",)
                    )
                row = cur.fetchone()
        return row[0] if row else 0
    except Exception as e:
        log.error(f"count_matches_fast error: {e}")
        return 0

def count_nameid_matches(query: str, qtype: str) -> int:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                if qtype == "national_id":
                    cleaned = re.sub(r"\s", "", query)
                    cur.execute(
                        "SELECT COUNT(*) FROM name_id_index WHERE national_id=%s",
                        (cleaned,)
                    )
                elif qtype == "partial_id":
                    cleaned = re.sub(r"\s", "", query)
                    cur.execute(
                        "SELECT COUNT(*) FROM name_id_index WHERE national_id ILIKE %s",
                        (f"%{cleaned}%",)
                    )
                else:  # name
                    norm_q = normalize_arabic(query)
                    cur.execute(
                        "SELECT COUNT(*) FROM name_id_index WHERE full_name ILIKE %s",
                        (f"%{norm_q}%",)
                    )
                row = cur.fetchone()
        return row[0] if row else 0
    except Exception as e:
        log.error(f"count_nameid_matches error: {e}")
        return 0

# ════════════════════════════════════════════
#     NAME / NATIONAL-ID DETECTION & SEARCH
# ════════════════════════════════════════════
def normalize_arabic(text: str) -> str:
    if not text:
        return ""
    text = text.strip()
    text = re.sub(r"[ًٌٍَُِّْـ]", "", text)
    text = re.sub(r"[أإآ]", "ا", text)
    text = re.sub(r"ة", "ه", text)
    text = re.sub(r"ى", "ي", text)
    text = re.sub(r"[^\w\s]", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.lower()

def is_national_id(value: str) -> bool:
    cleaned = re.sub(r"\s", "", value)
    return bool(re.fullmatch(r"\d{14}", cleaned))

def is_partial_national_id(value: str) -> bool:
    cleaned = re.sub(r"\s", "", value)
    return bool(re.fullmatch(r"\d{4,13}", cleaned))

def search_by_name(query: str, limit: int = 50) -> list:
    norm_q = normalize_arabic(query)
    query_words = [w for w in norm_q.split() if len(w) >= 2]
    if not query_words:
        return []
    results = []
    seen = set()
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                # Use ILIKE for each word
                conditions = " AND ".join(["full_name ILIKE %s"] * len(query_words))
                params = [f"%{w}%" for w in query_words] + [limit * 5]
                cur.execute(
                    f"SELECT full_name, national_id FROM name_id_index WHERE {conditions} LIMIT %s",
                    params
                )
                rows = cur.fetchall()
        for name, nat_id in rows:
            key = (name.strip(), nat_id.strip())
            if key not in seen:
                seen.add(key)
                results.append({"name": name.strip(), "national_id": nat_id.strip()})
            if len(results) >= limit:
                break
    except Exception as e:
        log.error(f"search_by_name error: {e}")
    return results

def search_by_national_id(query: str, limit: int = 50) -> list:
    cleaned = re.sub(r"\s", "", query)
    results = []
    seen = set()
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                if is_national_id(query):
                    cur.execute(
                        "SELECT full_name, national_id FROM name_id_index WHERE national_id=%s LIMIT %s",
                        (cleaned, limit)
                    )
                else:
                    cur.execute(
                        "SELECT full_name, national_id FROM name_id_index WHERE national_id ILIKE %s LIMIT %s",
                        (f"%{cleaned}%", limit)
                    )
                rows = cur.fetchall()
        for name, nat_id in rows:
            key = (name.strip(), nat_id.strip())
            if key not in seen:
                seen.add(key)
                results.append({"name": name.strip(), "national_id": nat_id.strip()})
    except Exception as e:
        log.error(f"search_by_national_id error: {e}")
    return results

def detect_nameid_query_type(query: str) -> str:
    q = query.strip()
    if is_national_id(q):
        return "national_id"
    if is_partial_national_id(q):
        return "partial_id"
    return "name"

def parse_excel_for_name_id(path: str, original_name: str) -> list:
    try:
        df = pd.read_excel(path, dtype=str)
    except Exception:
        return []

    NAME_ALIASES = {"اسم", "الاسم", "اسم كامل", "الاسم الكامل",
                    "name", "full name", "fullname", "full_name", "customer name"}
    ID_ALIASES   = {"رقم قومي", "الرقم القومي", "رقم هوية",
                    "national id", "nationalid", "national_id", "id number",
                    "id_number", "nid", "ssn", "identity"}

    name_col = None
    id_col   = None

    for col in df.columns:
        col_norm = str(col).strip().lower().replace("_", " ")
        if col_norm in NAME_ALIASES or any(a in col_norm for a in ("اسم", "name")):
            name_col = col
            break

    for col in df.columns:
        col_norm = str(col).strip().lower().replace("_", " ")
        if col_norm in ID_ALIASES or any(a in col_norm for a in ("قومي", "national", "id", "هوية")):
            sample  = df[col].dropna().astype(str).head(20)
            matches = sample.apply(lambda x: bool(re.fullmatch(r"\d{6,14}", re.sub(r"\D", "", x))))
            if matches.mean() > 0.4:
                id_col = col
                break

    if id_col is None:
        for col in df.columns:
            sample  = df[col].dropna().astype(str).head(20)
            matches = sample.apply(lambda x: bool(re.fullmatch(r"\d{14}", re.sub(r"\D", "", x))))
            if matches.mean() > 0.5:
                id_col = col
                break

    if not name_col or not id_col:
        return []

    results = []
    seen = set()
    for _, row in df.iterrows():
        name   = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
        nat_id = str(row[id_col]).strip()   if pd.notna(row[id_col])   else ""
        nat_id = re.sub(r"\D", "", nat_id)
        if name and re.fullmatch(r"\d{6,14}", nat_id):
            key = (name.lower(), nat_id)
            if key not in seen:
                seen.add(key)
                results.append((name, nat_id, original_name))
    return results

# ════════════════════════════════════════════
#         SMART LINE PARSER
# ════════════════════════════════════════════
def is_email(v):
    return bool(re.match(r"[^@\s]{1,64}@[^@\s]+\.[^@\s]{2,}", v))

def is_url(v):
    return bool(re.match(r"https?://\S+", v, re.I))

def is_domain_str(v):
    return bool(re.match(r"^(www\.)?[\w\-]+\.[a-z]{2,}(/\S*)?$", v, re.I))

def is_phone_str(v):
    cleaned = re.sub(r"[\s\-().+]", "", v)
    if re.fullmatch(r"\d{14}", cleaned):
        return False
    return bool(re.match(r"^\+?[\d\s\-().]{6,15}$", v))

def parse_line_fields(line: str) -> dict:
    line = line.strip()
    if not line:
        return {}
    if line.startswith("{") and line.endswith("}"):
        try:
            obj = json.loads(line)
            if isinstance(obj, dict):
                result = {}
                EMAIL_KEYS    = {"email", "mail", "e-mail", "e_mail"}
                PASS_KEYS     = {"password", "pass", "passwd", "pwd", "secret"}
                URL_KEYS      = {"url", "site", "domain", "host", "link", "website"}
                USERNAME_KEYS = {"username", "user", "login", "name", "uname", "nick"}
                PHONE_KEYS    = {"phone", "mobile", "tel", "cell", "number"}
                for k, v in obj.items():
                    kl = k.lower().strip()
                    sv = str(v).strip() if v else ""
                    if not sv or sv in ("null", "none", "nan"):
                        continue
                    if kl in EMAIL_KEYS and "email" not in result:
                        result["email"] = sv
                    elif kl in PASS_KEYS and "password" not in result:
                        result["password"] = sv
                    elif kl in URL_KEYS and "url" not in result:
                        result["url"] = sv
                    elif kl in USERNAME_KEYS and "username" not in result:
                        result["username"] = sv
                    elif kl in PHONE_KEYS and "phone" not in result:
                        result["phone"] = sv
                if result:
                    return result
        except (json.JSONDecodeError, Exception):
            pass

    if "|" in line or ";" in line or "\t" in line:
        parts = [p.strip() for p in re.split(r"[|;\t]", line) if p.strip()]
    else:
        url_match = re.match(r"(https?://[^:|\s]+)(.*)", line, re.I)
        if url_match:
            url_part = url_match.group(1)
            rest     = url_match.group(2).lstrip(":")
            parts    = [url_part] + [p.strip() for p in rest.split(":") if p.strip()]
        else:
            parts = [p.strip() for p in line.split(":") if p.strip()]

    result = {}
    for p in parts:
        if not p:
            continue
        if is_url(p) and "url" not in result:
            result["url"] = p
        elif is_email(p) and "email" not in result:
            result["email"] = p
        elif is_domain_str(p) and "url" not in result and "domain" not in result:
            result["domain"] = p
        elif is_phone_str(p) and "phone" not in result:
            result["phone"] = p
        else:
            if "username" not in result and "email" not in result:
                result["username"] = p
            elif ("username" in result or "email" in result) and "password" not in result:
                result["password"] = p
            elif "username" not in result:
                result["username"] = p
    return result

def line_matches_keyword(line: str, keyword: str) -> bool:
    return keyword.lower() in line.lower()

def extract_for_search_type(line: str, stype: str, keyword: str):
    if not line_matches_keyword(line, keyword):
        return None
    fields = parse_line_fields(line)
    if not fields:
        return None
    kw_lower = keyword.lower()
    if stype == "domain":
        url = fields.get("url", "") or fields.get("domain", "")
        if kw_lower not in url.lower():
            return None
    elif stype == "url":
        url = fields.get("url", "")
        if kw_lower not in url.lower():
            return None
    elif stype == "email":
        email = fields.get("email", "")
        if kw_lower not in email.lower():
            return None
    elif stype == "phone":
        phone = fields.get("phone", "")
        if kw_lower not in phone.lower():
            return None
    elif stype in ("username", "login"):
        uname = fields.get("username", "") or fields.get("email", "")
        if kw_lower not in uname.lower():
            return None
    elif stype == "password":
        pwd = fields.get("password", "")
        if kw_lower not in pwd.lower():
            return None
    return fields

def smart_search(keyword: str, stype: str, limit: int) -> list:
    fetch_limit = limit * 10
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                # Use GIN full-text index for keywords >= 3 chars (dramatically faster)
                if len(keyword) >= 3:
                    cur.execute(
                        "SELECT line FROM data_index WHERE to_tsvector('simple', line) @@ plainto_tsquery('simple', %s) LIMIT %s",
                        (keyword, fetch_limit)
                    )
                else:
                    cur.execute(
                        "SELECT line FROM data_index WHERE line ILIKE %s LIMIT %s",
                        (f"%{keyword}%", fetch_limit)
                    )
                rows = cur.fetchall()
    except Exception as e:
        log.error(f"smart_search query error: {e}")
        return []

    results = []
    seen = set()
    for (line,) in rows:
        fields = extract_for_search_type(line, stype, keyword)
        if fields is None:
            continue
        fp_parts = []
        for field in ("url", "domain", "email", "username", "phone", "password", "login"):
            val = fields.get(field, "").strip().lower()
            if val:
                fp_parts.append(val)
        if not fp_parts:
            fp_parts = [line.strip().lower()]
        fp = hashlib.md5("|".join(fp_parts).encode()).hexdigest()
        if fp in seen:
            continue
        seen.add(fp)
        results.append(fields)
        if len(results) >= limit:
            break
    return results

# ════════════════════════════════════════════
#         BUILD CLEAN RESULT FILE
# ════════════════════════════════════════════
async def safe_send_document(send_fn, path: str, filename: str, caption: str, reply_markup=None):
    MAX_CAPTION = 1024
    if len(caption) > MAX_CAPTION:
        caption = caption[:MAX_CAPTION - 10] + "\n_..._"

    last_err = None
    for attempt in range(3):
        try:
            with open(path, "rb") as f:
                await send_fn(
                    document=f, filename=filename,
                    caption=caption, parse_mode="Markdown",
                    reply_markup=reply_markup
                )
            return
        except BadRequest:
            plain = re.sub(r"[*_`\[\]]", "", caption)
            try:
                with open(path, "rb") as f:
                    await send_fn(
                        document=f, filename=filename,
                        caption=plain[:MAX_CAPTION],
                        reply_markup=reply_markup
                    )
                return
            except Exception as e:
                last_err = e
                break
        except RetryAfter as e:
            await asyncio.sleep(int(e.retry_after) + 1)
            last_err = e
        except TelegramError as e:
            last_err = e
            if attempt < 2:
                await asyncio.sleep(2 ** attempt)
        except Exception as e:
            last_err = e
            break

    log.error(f"safe_send_document failed: {last_err}")

def build_result_txt(keyword: str, results: list, stype: str) -> str:
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    truncated = False
    if len(results) > MAX_RESULT_LINES:
        results = results[:MAX_RESULT_LINES]
        truncated = True

    email_pass  = []
    email_only  = []
    url_combo   = []
    login_combo = []
    phone_list  = []

    for r in results:
        url  = r.get("url", "") or r.get("domain", "")
        em   = r.get("email", "")
        user = r.get("username", "")
        pwd  = r.get("password", "")
        ph   = r.get("phone", "")
        if em and pwd:
            if url:
                url_combo.append((url, em, pwd))
            else:
                email_pass.append((em, pwd))
        elif url and (em or user) and pwd:
            url_combo.append((url, em or user, pwd))
        elif (em or user) and pwd:
            login_combo.append((em or user, pwd))
        elif em:
            email_only.append(em)
        elif ph:
            phone_list.append(ph)
        elif user:
            login_combo.append((user, pwd))

    total = len(results)
    lines = [
        "═" * 60,
        "  🔍 DATA SCANNER BOT v9.0 — SCAN RESULTS",
        "═" * 60,
        f"  📌 Target   : {keyword}",
        f"  📂 Type     : {stype.upper()}",
        f"  📊 Total    : {total:,} records" + (" (truncated to 100K)" if truncated else ""),
        f"  📧 Email:Pass : {len(email_pass) + len(url_combo):,}",
        f"  👤 Login:Pass : {len(login_combo):,}",
        f"  📱 Phone      : {len(phone_list):,}",
        f"  🕐 Generated  : {now} UTC",
        f"  🤖 Bot        : @DataScannerBot",
        "═" * 60, "",
    ]
    if truncated:
        lines.insert(3, f"  ⚠️  Results exceeded {MAX_RESULT_LINES:,} — showing first {MAX_RESULT_LINES:,} only")
    if email_pass:
        lines += [f"{'─'*55}", f"  📧 EMAIL:PASS — {len(email_pass)} results", f"{'─'*55}"]
        for em, pwd in email_pass:
            lines.append(f"{em}:{pwd}")
        lines.append("")
    if url_combo:
        lines += [f"{'─'*55}", f"  🌐 URL | USER:PASS — {len(url_combo)} results", f"{'─'*55}"]
        for url, user, pwd in url_combo:
            lines.append(f"{url}|{user}:{pwd}")
        lines.append("")
    if login_combo:
        lines += [f"{'─'*55}", f"  👤 USER:PASS — {len(login_combo)} results", f"{'─'*55}"]
        for user, pwd in login_combo:
            lines.append(f"{user}:{pwd}" if pwd else user)
        lines.append("")
    if email_only:
        lines += [f"{'─'*55}", f"  📧 EMAIL ONLY — {len(email_only)} results", f"{'─'*55}"]
        lines += email_only + [""]
    if phone_list:
        lines += [f"{'─'*55}", f"  📱 PHONE — {len(phone_list)} results", f"{'─'*55}"]
        lines += phone_list + [""]
    lines += ["═"*55, f"  ✅ Total: {total:,} clean records", "═"*55]
    content = "\n".join(lines)
    return content

def build_nameid_result_txt(keyword: str, results: list, qtype: str) -> str:
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    type_label = {"national_id": "🪪 رقم قومي", "partial_id": "🔢 رقم جزئي", "name": "👤 اسم"}.get(qtype, qtype)
    lines = [
        "═" * 55,
        "  🔍 DATA SCANNER BOT — NAME / NATIONAL ID RESULTS",
        "═" * 55,
        f"  📌 Query  : {keyword}",
        f"  📂 Type   : {type_label}",
        f"  📊 Total  : {len(results)} records",
        f"  🕐 Time   : {now}",
        "═" * 55, "",
        f"{'─'*55}",
        f"  {'الاسم':<35} {'الرقم القومي'}",
        f"{'─'*55}",
    ]
    for r in results:
        name   = r.get("name", "—")
        nat_id = r.get("national_id", "—")
        lines.append(f"  {name:<35} {nat_id}")
    lines += ["", "═"*55, f"  ✅ Total: {len(results)} records", "═"*55]
    return "\n".join(lines)

# ════════════════════════════════════════════
#               FILE PARSING
# ════════════════════════════════════════════
def _open_text_file(path: str):
    for enc in ("utf-8", "cp1256", "latin-1"):
        try:
            f = open(path, "r", encoding=enc, errors="strict")
            f.read(512)
            f.seek(0)
            return f, enc
        except (UnicodeDecodeError, Exception):
            try:
                f.close()
            except Exception:
                pass
    return open(path, "r", encoding="utf-8", errors="replace"), "utf-8(replace)"

def parse_file(path: str, original_name: str) -> list:
    results = []
    source  = original_name
    ext     = original_name.lower().rsplit(".", 1)[-1] if "." in original_name else ""
    seen    = set()

    def add_line(line: str):
        if not line:
            return
        line = line.strip()
        if not line or len(line) < 3:
            return
        if line.lower() in ("null", "none"):
            return
        if line not in seen:
            seen.add(line)
            results.append((line, source))

    if ext == "txt":
        f, enc = _open_text_file(path)
        log.info(f"Parsing TXT [{enc}]: {original_name}")
        with f:
            for line in f:
                add_line(line)
    elif ext == "csv":
        loaded = False
        for enc in ("utf-8", "cp1256", "latin-1"):
            try:
                df = pd.read_csv(path, dtype=str, on_bad_lines="skip", encoding=enc)
                loaded = True
                break
            except (UnicodeDecodeError, Exception):
                continue
        if not loaded:
            df = pd.read_csv(path, dtype=str, on_bad_lines="skip", encoding="latin-1", errors="replace")
        for _, row in df.iterrows():
            vals = [str(v).strip() for v in row if pd.notna(v) and str(v).strip()]
            if vals:
                add_line(":".join(vals))
    elif ext in ("xlsx", "xls"):
        df = pd.read_excel(path, dtype=str)
        for _, row in df.iterrows():
            vals = [str(v).strip() for v in row if pd.notna(v) and str(v).strip()]
            if vals:
                add_line(":".join(vals))
    elif ext == "json":
        f, enc = _open_text_file(path)
        with f:
            raw = f.read()
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            data = []
            for ln in raw.splitlines():
                ln = ln.strip()
                if ln:
                    try:
                        data.append(json.loads(ln))
                    except Exception:
                        pass

        def flatten(obj):
            if isinstance(obj, dict):
                yield json.dumps(obj, ensure_ascii=False)
                for v in obj.values():
                    yield from flatten(v)
            elif isinstance(obj, list):
                for item in obj:
                    yield from flatten(item)
            elif isinstance(obj, str):
                yield obj

        for v in flatten(data):
            add_line(v)

    return results

# ════════════════════════════════════════════
#                  KEYBOARDS
# ════════════════════════════════════════════
def user_main_kb(uid: int = 0):
    lang  = get_lang(uid) if uid else "en"
    is_ar = lang == "ar"
    st    = STRINGS.get(lang, STRINGS["en"])
    rows = [
        [InlineKeyboardButton(st["btn_search"],    callback_data="go_search")],
        [InlineKeyboardButton(st["btn_nameid"],    callback_data="go_nameid")],
        [InlineKeyboardButton("🔎 " + ("بحث متعدد" if is_ar else "Bulk Search"), callback_data="go_bulk_search")],
        [
            InlineKeyboardButton(st["btn_account"], callback_data="my_account"),
            InlineKeyboardButton(st["btn_plans"],   callback_data="show_plans"),
        ],
        [InlineKeyboardButton(st["btn_subscribe"], callback_data="user_subscribe")],
        [
            InlineKeyboardButton(st["btn_help"],     callback_data="show_help"),
            InlineKeyboardButton(st["btn_language"], callback_data="set_language"),
        ],
        [
            InlineKeyboardButton("⭐ " + ("المفضلة" if is_ar else "Favorites"),       callback_data="my_favorites"),
            InlineKeyboardButton("🏆 " + ("نقاطي" if is_ar else "My Points"),         callback_data="my_points"),
        ],
        [
            InlineKeyboardButton("🎮 " + ("الإنجازات" if is_ar else "Achievements"),   callback_data="my_achievements"),
            InlineKeyboardButton("📜 " + ("بحوثي الأخيرة" if is_ar else "History"),   callback_data="my_history"),
        ],
        [
            InlineKeyboardButton("🔐 " + ("الأمان / 2FA" if is_ar else "Security/2FA"), callback_data="my_security"),
            InlineKeyboardButton("🔍 " + ("بحث Regex" if is_ar else "Regex Search"),    callback_data="go_regex_search"),
        ],
        [
            InlineKeyboardButton("🔗 " + ("إحالة متقدمة" if is_ar else "Affiliate"),    callback_data="my_affiliate"),
            InlineKeyboardButton("📋 " + ("طلباتي" if is_ar else "My Orders"),           callback_data="my_orders"),
        ],
        [
            InlineKeyboardButton("🔐 " + ("الأمان" if is_ar else "Security"),            callback_data="my_security"),
            InlineKeyboardButton("🔄 " + ("تجديد تلقائي" if is_ar else "Auto-Renewal"),  callback_data="my_autorenewal"),
        ],
        [
            InlineKeyboardButton("🆔 " + ("معرفي" if is_ar else "My ID"),               callback_data="my_id"),
            InlineKeyboardButton("🏷 " + ("كوبون" if is_ar else "Coupon"),               callback_data="my_coupon"),
        ],
    ]
    if uid:
        u = get_user_cached(uid)
        if u:
            tier_v = u[3]
            exp_at = u[7]
            if tier_v == "free" or (exp_at and (datetime.fromisoformat(str(exp_at)) - datetime.utcnow()).days <= 3):
                renew_label = "🔄 تجديد الاشتراك" if is_ar else "🔄 Renew Subscription"
                rows.insert(3, [InlineKeyboardButton(renew_label, callback_data="user_subscribe")])
    return InlineKeyboardMarkup(rows)

def search_type_kb(uid: int = 0):
    last_type = None
    if uid:
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT last_search_type FROM users WHERE user_id=%s", (uid,))
                    row = cur.fetchone()
                    last_type = row[0] if row and row[0] else None
        except Exception:
            pass

    type_buttons = {
        "url":      InlineKeyboardButton("🌐 URL",      callback_data="st_url"),
        "domain":   InlineKeyboardButton("🌍 Domain",   callback_data="st_domain"),
        "login":    InlineKeyboardButton("👤 Login",    callback_data="st_login"),
        "username": InlineKeyboardButton("📝 Username", callback_data="st_username"),
        "email":    InlineKeyboardButton("📧 Email",    callback_data="st_email"),
        "phone":    InlineKeyboardButton("📱 Phone",    callback_data="st_phone"),
        "password": InlineKeyboardButton("🔑 Password", callback_data="st_password"),
    }
    rows = []
    if last_type and last_type in type_buttons:
        rows.append([InlineKeyboardButton(f"⭐ Last: {last_type.upper()}", callback_data=f"st_{last_type}")])
    rows += [
        [type_buttons["url"], type_buttons["domain"]],
        [type_buttons["login"], type_buttons["username"]],
        [type_buttons["email"], type_buttons["phone"]],
        [type_buttons["password"]],
        [InlineKeyboardButton("🔎 Full Scan 👑", callback_data="st_all")],
        [InlineKeyboardButton("🔙 Back",         callback_data="user_home")],
    ]
    return InlineKeyboardMarkup(rows)

def nameid_type_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👤 بحث بالاسم",        callback_data="ni_name")],
        [InlineKeyboardButton("🪪 بحث بالرقم القومي", callback_data="ni_national_id")],
        [InlineKeyboardButton("🔙 Back",              callback_data="user_home")],
    ])

def back_user_kb(uid: int = 0):
    label = "🔙 القائمة الرئيسية" if get_lang(uid) == "ar" else "🔙 Main Menu"
    return InlineKeyboardMarkup([[InlineKeyboardButton(label, callback_data="user_home")]])

def new_search_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔍 New Search",     callback_data="go_search")],
        [InlineKeyboardButton("🪪 Name/ID Search", callback_data="go_nameid")],
        [InlineKeyboardButton("🏠 Main Menu",      callback_data="user_home")],
    ])

def result_share_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔍 New Search",        callback_data="go_search")],
        [InlineKeyboardButton("🏠 Main Menu",          callback_data="user_home")],
    ])

def admin_main_kb():
    maint_label = "🔧 Maintenance: ON ✅" if MAINTENANCE_MODE else "🔧 Maintenance: OFF"
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("🔍 Search DB",      callback_data="go_search"),
            InlineKeyboardButton("🪪 Name/ID Search", callback_data="go_nameid"),
        ],
        [
            InlineKeyboardButton("📂 Upload Data",  callback_data="adm_upload_info"),
            InlineKeyboardButton("👥 Users",        callback_data="adm_users"),
        ],
        [
            InlineKeyboardButton("➕ Add User",     callback_data="adm_adduser_inline"),
            InlineKeyboardButton("🗑️ Del User",     callback_data="adm_deluser"),
        ],
        [
            InlineKeyboardButton("📊 Basic Stats",    callback_data="adm_stats"),
            InlineKeyboardButton("📈 Advanced Stats", callback_data="adm_advanced_stats"),
        ],
        [
            InlineKeyboardButton("💰 Add Credits",  callback_data="adm_add_credits"),
            InlineKeyboardButton("⬆️ Set Tier",     callback_data="adm_set_tier"),
        ],
        [
            InlineKeyboardButton("🔒 Ban User",     callback_data="adm_ban"),
            InlineKeyboardButton("✅ Unban User",   callback_data="adm_unban"),
        ],
        [InlineKeyboardButton("🧊 Freeze User",     callback_data="adm_freeze")],
        [
            InlineKeyboardButton("📜 Search Logs",  callback_data="adm_logs"),
            InlineKeyboardButton("🗂️ Files",         callback_data="adm_filelist"),
        ],
        [
            InlineKeyboardButton("🗑️ Delete File",   callback_data="adm_delete_file"),
            InlineKeyboardButton("🔄 Reset Daily",   callback_data="adm_reset_daily"),
        ],
        [InlineKeyboardButton("📢 Broadcast",        callback_data="adm_broadcast")],
        [InlineKeyboardButton("✉️ Message User",      callback_data="adm_msg_user")],
        [InlineKeyboardButton("📅 Set Expiry",        callback_data="adm_set_expiry")],
        [InlineKeyboardButton("🔎 Filter Logs",       callback_data="adm_filter_logs")],
        [
            InlineKeyboardButton("💾 Backup DB",         callback_data="adm_backup"),
            InlineKeyboardButton("📤 Export Users CSV",  callback_data="adm_export_csv"),
        ],
        [InlineKeyboardButton("📋 طلبات الاشتراك",   callback_data="adm_sub_requests")],
        [InlineKeyboardButton("🎫 Support Tickets",   callback_data="adm_tickets")],
        [InlineKeyboardButton("💳 Payment Orders",    callback_data="adm_payments")],
        [InlineKeyboardButton("📊 Stats Dashboard",   callback_data="adm_stats")],
        [InlineKeyboardButton("🔗 URL Import",         callback_data="adm_url_import"),
         InlineKeyboardButton("📋 User Activity",      callback_data="adm_user_log_prompt")],
        [InlineKeyboardButton("🔗 URL Import",        callback_data="adm_url_import"),
         InlineKeyboardButton("🔎 Regex Search",      callback_data="adm_regex_search")],
        [InlineKeyboardButton("🏷 Coupons",            callback_data="adm_coupons")],
        [InlineKeyboardButton("📜 سجل عمليات الأدمن", callback_data="adm_op_logs")],
        [InlineKeyboardButton("⚡ Bot Status",        callback_data="adm_bot_status")],
        [InlineKeyboardButton(maint_label,            callback_data="adm_toggle_maintenance")],
    ])

def back_admin_kb():
    return InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Admin Panel", callback_data="adm_home")]])

# ════════════════════════════════════════════
#                 /START
# ════════════════════════════════════════════
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user

    if MAINTENANCE_MODE and not is_admin(user.id):
        await update.message.reply_text(MAINTENANCE_MSG, parse_mode="Markdown")
        return

    is_new = get_user(user.id) is None

    if WHITELIST_MODE and user.id not in WHITELIST_IDS and not is_admin(user.id):
        await update.message.reply_text("🔒 This bot is in private mode. Access restricted.")
        return

    ensure_user(user.id, user.username or "", user.first_name or "")

    if is_new:
        for admin_id in ADMIN_IDS:
            try:
                await context.bot.send_message(
                    chat_id=admin_id,
                    text=(
                        f"🆕 *New User Registered!*\n"
                        f"━━━━━━━━━━━━━━━━━━━━━━\n"
                        f"🆔 ID       : `{user.id}`\n"
                        f"👤 Name     : {esc(user.first_name or 'N/A')}\n"
                        f"🔖 Username : @{esc(user.username or 'N/A')}"
                    ),
                    parse_mode="Markdown"
                )
            except Exception:
                pass

    if is_new and context.args:
        try:
            arg = context.args[0]
            if arg.startswith("ref_"):
                # Affiliate link click
                code = arg.replace("ref_", "")
                track_affiliate_click(code)
                track_affiliate_conversion(code, REFERRAL_CREDITS)
                try:
                    with pool_conn() as conn:
                        with conn.cursor() as cur:
                            cur.execute("SELECT user_id FROM affiliate_links WHERE code=%s", (code,))
                            row = cur.fetchone()
                    if row:
                        process_referral(user.id, row[0])
                except Exception:
                    pass
            else:
                ref_uid = int(arg)
                process_referral(user.id, ref_uid)
        except (ValueError, TypeError):
            pass

    if is_banned(user.id):
        await update.message.reply_text(
            "🚫 *Your account has been banned.*\n\nIf you believe this is a mistake, please contact the admin.",
            parse_mode="Markdown"
        )
        return

    if is_admin(user.id):
        await show_admin_home(update, context, send=True)
        return

    if is_new:
        name = esc(user.first_name or "there")
        await update.message.reply_text(
            f"👋 Welcome, <b>{name}</b>!\n\n"
            f"🤖 <b>DATA SCANNER YUTO BOT</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🔍 Search millions of database records\n"
            f"🪪 Name & National ID lookup\n"
            f"💎 Multiple subscription tiers\n\n"
            f"🌐 You can switch language anytime from the menu.\n\n"
            f"Press the button below to get started! 👇",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🚀 Get Started", callback_data="user_home")],
                [InlineKeyboardButton("📖 Tutorial",    callback_data="onboard_0")],
                [InlineKeyboardButton("🌐 العربية",     callback_data="lang_ar")],
            ])
        )
        award_badge(user.id, "newcomer")
        return

    await show_user_home(update, context, send=True)

def esc(text: str) -> str:
    return (str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;"))

def mesc(text: str) -> str:
    return (str(text)
            .replace("\\", "\\\\")
            .replace("*", "\\*")
            .replace("_", "\\_")
            .replace("`", "\\`")
            .replace("[", "\\["))

# ── DB Count Cache (avoids 9s COUNT(*) on 115M rows) ─────
_count_cache = {"data": 0, "nameid": 0, "users": 0, "files": 0, "searches": 0, "banned": 0, "ts": 0}

def get_cached_counts():
    """Cache COUNT queries for 5 minutes — COUNT(*) on 115M rows takes ~9s."""
    now = time.monotonic()
    if now - _count_cache["ts"] < 300:
        return _count_cache.copy()
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM data_index")
                _count_cache["data"] = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM name_id_index")
                _count_cache["nameid"] = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM users")
                _count_cache["users"] = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM uploaded_files")
                _count_cache["files"] = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM search_logs")
                _count_cache["searches"] = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM users WHERE is_banned=1")
                _count_cache["banned"] = cur.fetchone()[0]
        _count_cache["ts"] = now
    except Exception as e:
        log.error(f"get_cached_counts error: {e}")
    return _count_cache.copy()

async def show_user_home(update, context, send=False, query=None):
    uid = query.from_user.id if query else update.effective_user.id
    u   = get_user_cached(uid)
    tier_key   = u[3]  if u else "free"
    daily_left = u[4]  if u else 5
    credits    = u[5]  if u else 0
    full_name  = u[2]  if u else "User"
    expires_at = u[7]  if u else None
    t          = TIERS.get(tier_key, TIERS["free"])

    try:
        counts = get_cached_counts()
        nameid_count   = counts["nameid"]
        db_count       = counts["data"]
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM search_logs WHERE user_id=%s", (uid,))
                total_searches = cur.fetchone()[0]
    except Exception:
        nameid_count = total_searches = db_count = 0

    lang  = get_lang(uid)
    st    = STRINGS.get(lang, STRINGS["en"])
    is_ar = lang == "ar"
    nameid_left = u[13] if u and len(u) > 13 else 0

    now_utc      = datetime.utcnow()
    next_midnight = (now_utc + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    hrs_left  = int((next_midnight - now_utc).total_seconds() // 3600)
    mins_left = int(((next_midnight - now_utc).total_seconds() % 3600) // 60)
    renew_str = f"{hrs_left}h {mins_left}m" if not is_ar else f"{hrs_left}س {mins_left}د"

    exp_line = ""
    if expires_at:
        try:
            exp_dt   = datetime.fromisoformat(str(expires_at))
            days_rem = (exp_dt - now_utc).days
            exp_line = f"\n📅 {'ينتهي في' if is_ar else 'Expires'}: <code>{str(expires_at)[:10]}</code> ({days_rem}d)"
        except Exception:
            pass

    text = (
        f"🤖 <b>DATA SCANNER YUTO BOT</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👋 {st['hello']}, <b>{esc(full_name)}</b>!\n\n"
        f"📦 {st['plan']}: {esc(t['label'])}{exp_line}\n"
        f"🔍 {st['daily_left']}: <code>{daily_left}</code>\n"
        f"🪪 {'Name/ID بحوث' if is_ar else 'Name/ID searches'}: <code>{nameid_left}</code>\n"
        f"⏰ {'يتجدد خلال' if is_ar else 'Renews in'}: <code>{renew_str}</code>\n"
        f"💰 {st['credits']}: <code>{credits}</code>\n"
        f"🔢 {'بحوثي الكلية' if is_ar else 'Total searches'}: <code>{total_searches:,}</code>\n"
        f"🗄️ {st['db_records']}: <code>{db_count:,}</code>\n"
        f"🪪 {st['nameid_records']}: <code>{nameid_count:,}</code>\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"{st['menu_hint']}\n\n"
        f"<i>💡 {'اختصارات: ' if is_ar else 'Shortcuts: '}"
        f"<code>e:gmail.com</code> · <code>p:+201</code> · <code>d:site.com</code> · <code>ip:1.1.1.1</code></i>"
    )
    if send:
        await update.message.reply_text(text, parse_mode="HTML", reply_markup=user_main_kb(uid))
    else:
        await query.edit_message_text(text, parse_mode="HTML", reply_markup=user_main_kb(uid))

async def show_admin_home(update, context, send=False, query=None):
    try:
        counts = get_cached_counts()
        total_records  = counts["data"]
        total_nameid   = counts["nameid"]
        total_users    = counts["users"]
        total_files    = counts["files"]
        total_searches = counts["searches"]
        banned_count   = counts["banned"]
    except Exception as e:
        log.error(f"show_admin_home error: {e}")
        total_records = total_nameid = total_users = total_files = total_searches = banned_count = 0

    text = (
        f"⚙️ *ADMIN CONTROL PANEL*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🗄️ DB Records    : `{total_records:,}`\n"
        f"🪪 Name/ID Rows  : `{total_nameid:,}`\n"
        f"👥 Total Users   : `{total_users:,}`\n"
        f"🚫 Banned Users  : `{banned_count}`\n"
        f"📁 Indexed Files : `{total_files}`\n"
        f"🔍 Total Searches: `{total_searches:,}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📂 Send any TXT/CSV/XLSX/JSON file here to index it.\n"
        f"📋 Excel files with name + national ID columns are *also* indexed in the Name/ID table automatically."
    )
    if send:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=admin_main_kb())
    else:
        await query.edit_message_text(text, parse_mode="Markdown", reply_markup=admin_main_kb())

# ── Callback rate limiting ────────────────────────────────
_last_callback: dict = {}
_CALLBACK_COOLDOWN = 1.0

# ════════════════════════════════════════════
#            CALLBACK QUERY ROUTER
# ════════════════════════════════════════════
async def callback_router(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q    = update.callback_query
    uid  = q.from_user.id
    data = _cb_get(q.data)

    now_t = time.monotonic()
    with _rate_limit_lock:
        last_cb = _last_callback.get(uid, 0)
        if not is_admin(uid) and (now_t - last_cb) < _CALLBACK_COOLDOWN:
            await q.answer("⏳ Too fast! Please slow down.", show_alert=False)
            return
        _last_callback[uid] = now_t

    await q.answer()
    ensure_user(uid, q.from_user.username or "", q.from_user.first_name or "")

    if is_banned(uid) and not is_admin(uid):
        await q.edit_message_text(
            "🚫 *Your account has been banned.*\n\nContact the admin for support.",
            parse_mode="Markdown"
        )
        return

    if data == "user_home":
        await show_user_home(update, context, query=q)
        return

    if data == "set_language":
        await q.edit_message_text(
            "🌐 *Choose your language / اختر لغتك:*",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🇬🇧 English",  callback_data="lang_en"),
                 InlineKeyboardButton("🇸🇦 العربية",  callback_data="lang_ar")],
                [InlineKeyboardButton("🇹🇷 Türkçe",   callback_data="lang_tr"),
                 InlineKeyboardButton("🇮🇷 فارسی",    callback_data="lang_fa")],
                [InlineKeyboardButton("🇫🇷 Français", callback_data="lang_fr")],
                [InlineKeyboardButton("🔙 Back", callback_data="user_home")],
            ])
        )
        return

    if data in ("lang_en", "lang_ar", "lang_tr", "lang_fa", "lang_fr"):
        chosen = data.split("_")[1]
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("UPDATE users SET lang=%s WHERE user_id=%s", (chosen, uid))
                conn.commit()
            invalidate_user_cache(uid)
        except Exception as e:
            log.error(f"lang update error: {e}")
        confirm = STRINGS.get(chosen, STRINGS["en"])["lang_set"]
        await q.edit_message_text(
            confirm, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(
                "🔙 Menu", callback_data="user_home"
            )]])
        )
        return

    if data == "user_subscribe":
        st = STRINGS.get(get_lang(uid), STRINGS["en"])
        tiers_kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("⭐ Basic",   callback_data="sub_req_basic"),
             InlineKeyboardButton("💎 Premium", callback_data="sub_req_premium")],
            [InlineKeyboardButton("👑 VIP",     callback_data="sub_req_vip")],
            [InlineKeyboardButton("🔙 Back / رجوع", callback_data="user_home")],
        ])
        await q.edit_message_text(
            f"{st['sub_req_title']}\n\n{st['sub_req_prompt']}",
            parse_mode="Markdown", reply_markup=tiers_kb
        )
        return

    if data.startswith("sub_req_"):
        tier = data.replace("sub_req_", "")
        if tier not in TIERS or tier == "free":
            await q.answer("❌ Invalid tier", show_alert=True)
            return
        user_obj = q.from_user

        async def reply_fn(text, parse_mode=None):
            await q.edit_message_text(text, parse_mode=parse_mode, reply_markup=back_user_kb(uid))

        await _do_subscribe_request(
            uid, user_obj.username or "", user_obj.first_name or "User",
            tier, context, reply_fn
        )
        return

    if data == "adm_home":
        if not is_admin(uid):
            return
        await show_admin_home(update, context, query=q)
        return

    if data == "my_account":
        u = get_user_cached(uid)
        if not u:
            await q.edit_message_text("❌ Account not found. Send /start first.", reply_markup=back_user_kb(uid))
            return
        lang  = get_lang(uid)
        is_ar = lang == "ar"
        tier_d   = TIERS.get(u[3], TIERS["free"])
        username = esc(u[1] or "N/A")
        fullname = esc(u[2] or "User")
        daily    = u[4]
        credits  = u[5]
        exp      = esc(str(u[7]) if u[7] else ("لا يوجد انتهاء" if is_ar else "No expiry"))
        tier_name = u[3]

        max_daily = TIERS.get(tier_name, TIERS["free"])["daily"]
        used      = max(0, max_daily - daily) if max_daily > 0 else 0
        if max_daily > 0 and max_daily < 100_000:
            pct   = min(100, int(used / max_daily * 100))
            bars  = int(pct / 10)
            bar   = "█" * bars + "░" * (10 - bars)
            daily_bar = f"`{bar}` {pct}% ({used}/{max_daily})"
        else:
            daily_bar = f"`{daily}` {'متبقي' if is_ar else 'remaining'}"

        max_ni  = NAMEID_TIERS.get(tier_name, NAMEID_TIERS["free"])["daily_nameid"]
        ni_left = u[13] if len(u) > 13 else 0
        ni_used = max(0, max_ni - ni_left) if max_ni > 0 else 0
        if max_ni > 0 and max_ni < 100_000:
            ni_pct  = min(100, int(ni_used / max_ni * 100))
            ni_bars = int(ni_pct / 10)
            ni_bar  = f"`{'█'*ni_bars}{'░'*(10-ni_bars)}` {ni_pct}% ({ni_used}/{max_ni})"
        else:
            ni_bar  = f"`{ni_left}` {'متبقي' if is_ar else 'remaining'}"

        if is_ar:
            text = (
                f"📊 <b>حسابي</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━\n"
                f"👤 الاسم    : <b>{fullname}</b>\n"
                f"🔖 المعرف   : @{username}\n"
                f"🆔 المعرف الرقمي: <code>{uid}</code>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━\n"
                f"📦 الباقة   : {esc(tier_d['label'])}\n"
                f"🔍 البحث اليومي:\n  {daily_bar}\n"
                f"🪪 Name/ID اليومي:\n  {ni_bar}\n"
                f"💰 الرصيد   : <code>{credits}</code>\n"
                f"📅 الانتهاء : <code>{exp}</code>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━"
            )
        else:
            text = (
                f"📊 <b>My Account</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━\n"
                f"👤 Name    : <b>{fullname}</b>\n"
                f"🔖 Username: @{username}\n"
                f"🆔 User ID : <code>{uid}</code>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━\n"
                f"📦 Plan    : {esc(tier_d['label'])}\n"
                f"🔍 Daily Search:\n  {daily_bar}\n"
                f"🪪 Daily Name/ID:\n  {ni_bar}\n"
                f"💰 Credits : <code>{credits}</code>\n"
                f"📅 Expires : <code>{exp}</code>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━"
            )
        await q.edit_message_text(text, parse_mode="HTML", reply_markup=back_user_kb(uid))
        return

    if data == "go_nameid":
        if not can_search_nameid(uid):
            is_ar = get_lang(uid) == "ar"
            msg_no = (
                "❌ *انتهت بحوث Name/ID لليوم.*\n\nقم بترقية باقتك للحصول على المزيد."
                if is_ar else
                "❌ *Name/ID searches used up for today.*\n\nUpgrade your plan for more."
            )
            await q.edit_message_text(
                msg_no, parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("💳 View Plans", callback_data="show_plans")],
                    [InlineKeyboardButton("🔙 Back",       callback_data="user_home")],
                ])
            )
            return
        await q.edit_message_text(
            "🪪 *Name / National ID Search*\n"
            "━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "اختار نوع البحث:",
            parse_mode="Markdown",
            reply_markup=nameid_type_kb()
        )
        return

    if data in ("ni_name", "ni_national_id"):
        if not can_search(uid):
            await q.edit_message_text("❌ No searches remaining.", reply_markup=back_user_kb(uid))
            return
        context.user_data["search_type"] = data
        if data == "ni_name":
            prompt = "👤 *بحث بالاسم*\n\n✏️ ابعت الاسم أو جزء منه:\n_مثال: عبد الفتاح السيسي_"
        else:
            prompt = "🪪 *بحث بالرقم القومي*\n\n✏️ ابعت الرقم القومي (14 رقم) أو جزء منه:\n_مثال: 3060_"
        await q.edit_message_text(
            prompt, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="go_nameid")]])
        )
        return

    if data == "go_search":
        await q.edit_message_text(
            "🔍 *Choose Search Type*\n\nSelect the type of data you want to search for:",
            parse_mode="Markdown", reply_markup=search_type_kb(uid)
        )
        return

    if data.startswith("st_"):
        stype = data[3:]
        if stype == "all":
            tier = get_tier(uid)
            if tier not in ("premium", "vip") and not is_admin(uid):
                await q.edit_message_text(
                    "👑 *Full Scan* is exclusive to *Premium* and *VIP* members.\n\nUpgrade to unlock!",
                    parse_mode="Markdown",
                    reply_markup=InlineKeyboardMarkup([
                        [InlineKeyboardButton("💳 View Plans", callback_data="show_plans")],
                        [InlineKeyboardButton("🔙 Back",       callback_data="go_search")],
                    ])
                )
                return
        if not can_search(uid):
            await q.edit_message_text(
                "❌ *No searches remaining.*\n\nUpgrade your plan or buy credits.",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("💳 View Plans", callback_data="show_plans")],
                    [InlineKeyboardButton("🔙 Back",       callback_data="user_home")],
                ])
            )
            return
        context.user_data["search_type"] = stype
        icons = {"url":"🌐","domain":"🌍","login":"👤","username":"📝",
                 "email":"📧","phone":"📱","password":"🔑","all":"🔎"}
        icon = icons.get(stype, "🔍")

        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT DISTINCT keyword FROM search_logs WHERE user_id=%s "
                        "ORDER BY timestamp DESC LIMIT 5",
                        (uid,)
                    )
                    recent = cur.fetchall()
        except Exception:
            recent = []

        kb_rows = []
        if recent:
            for (kw,) in recent:
                safe_kw = kw[:35]
                kb_rows.append([InlineKeyboardButton(
                    f"🕐 {safe_kw}", callback_data=_cb_put(f"confirm_search:{stype}:{safe_kw}")
                )])
        kb_rows.append([InlineKeyboardButton("🔙 Cancel", callback_data="go_search")])

        hint = "\n\n🕐 *Or pick a recent keyword:*" if recent else ""
        await q.edit_message_text(
            f"{icon} *{stype.upper()} Search*\n\n"
            f"✏️ Send your target keyword now:\n\n"
            f"⏱️ Search runs for up to *3 minutes* to find best results.{hint}",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(kb_rows)
        )
        return

    if data == "show_plans":
        await show_payment_plans(update, context, query=q)
        return

    if data.startswith("pay_plan_"):
        plan = data.replace("pay_plan_", "")
        if plan not in PLAN_PRICES:
            await q.answer("❌ Invalid plan", show_alert=True)
            return
        await show_duration_picker(q, plan, uid)
        return

    if data.startswith("pay_dur_"):
        _, _, plan, duration = data.split("_", 3)
        if plan not in PLAN_PRICES or duration not in PLAN_DURATIONS:
            await q.answer("❌ Invalid selection", show_alert=True)
            return
        await show_method_picker(q, plan, duration, uid)
        return

    if data.startswith("pay_method_"):
        # pay_method_{plan}_{duration}_{type}_{key}
        parts = data.split("_", 5)
        if len(parts) < 6:
            await q.answer("❌ Invalid", show_alert=True)
            return
        _, _, plan, duration, method_type, method_key = parts
        await handle_payment_method(q, context, plan, duration, method_type, method_key, uid)
        return

    if data.startswith("view_order_"):
        order_id = int(data.replace("view_order_", ""))
        await show_order_view(q, order_id, uid)
        return

    if data == "my_orders":
        await show_my_orders(q, uid)
        return

    if data.startswith("cancel_order_"):
        order_id = int(data.replace("cancel_order_", ""))
        order    = get_order(order_id)
        if not order or order["user_id"] != uid:
            await q.answer("❌ Not found.", show_alert=True)
            return
        if order["status"] != "pending":
            await q.answer("❌ Order already processed.", show_alert=True)
            return
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("UPDATE payment_orders SET status='cancelled' WHERE id=%s AND user_id=%s", (order_id, uid))
                conn.commit()
            context.user_data.pop("pending_order_id", None)
            await q.edit_message_text(
                f"❌ *Order #{order_id} Cancelled.*",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Menu", callback_data="user_home")]])
            )
        except Exception as e:
            await q.answer(f"Error: {e}", show_alert=True)
        return

    # ── Admin payment callbacks ──────────────────────────
    if data == "adm_payments":
        if not is_admin(uid):
            return
        await show_admin_payments(update, context, query=q)
        return

    if data.startswith("adm_view_order_"):
        if not is_admin(uid):
            return
        order_id = int(data.replace("adm_view_order_", ""))
        await show_admin_order_detail(q, order_id, context)
        return

    if data.startswith("adm_confirm_order_"):
        if not is_admin(uid):
            return
        order_id = int(data.replace("adm_confirm_order_", ""))
        res = confirm_order(order_id, uid)
        if res["ok"]:
            plan_lbl = PLAN_PRICES.get(res["plan"], {}).get("label", res["plan"])
            await q.edit_message_text(
                f"✅ *Order #{order_id} Confirmed!*\n\n"
                f"👤 User `{res['user_id']}` upgraded to {plan_lbl}\n"
                f"⏱ {res['days']} days | Expires: {str(res['expires'])[:10]}",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Orders", callback_data="adm_payments")]])
            )
            try:
                await context.bot.send_message(
                    chat_id=res["user_id"],
                    text=(
                        f"✅ *Payment Confirmed!*\n\n"
                        f"📦 Plan: {plan_lbl}\n"
                        f"⏱ Duration: {res['days']} days\n"
                        f"🗓 Expires: {str(res['expires'])[:10]}\n\n"
                        f"Press /start to refresh your dashboard! 🚀"
                    ),
                    parse_mode="Markdown"
                )
            except Exception:
                pass
        else:
            await q.edit_message_text(f"❌ Error: {res['msg']}")
        return

    if data.startswith("adm_reject_order_"):
        if not is_admin(uid):
            return
        order_id = int(data.replace("adm_reject_order_", ""))
        context.user_data["admin_action"] = f"reject_order_{order_id}"
        await q.edit_message_text(
            f"❌ *Reject Order #{order_id}*\n\nSend rejection reason (or send `-` to skip):",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data=f"adm_view_order_{order_id}")]])
        )
        return
        is_ar = get_lang(uid) == "ar"
        if is_ar:
            text = (
                f"💳 *باقات الاشتراك*\n"
                f"━━━━━━━━━━━━━━━━━━━━━━\n"
                f"🆓 *Free* • 0 بحث DB/يوم • 0 بحث Name/ID\n\n"
                f"⭐ *Basic*\n"
                f"  🔍 Search DB: 10 بحث/يوم ← 200 نتيجة\n"
                f"  🪪 Name/ID: 5 بحوث/يوم ← 50 نتيجة\n\n"
                f"💎 *Premium*\n"
                f"  🔍 Search DB: 15 بحث/يوم ← 1,000 نتيجة + Full Scan ✅\n"
                f"  🪪 Name/ID: 10 بحوث/يوم ← 500 نتيجة\n\n"
                f"👑 *VIP*\n"
                f"  🔍 Search DB: غير محدود ← 1,000,000 نتيجة + Full Scan ✅\n"
                f"  🪪 Name/ID: غير محدود ← 100,000 نتيجة\n"
                f"━━━━━━━━━━━━━━━━━━━━━━\n"
                f"📩 تواصل مع @yut3ev للترقية."
            )
        else:
            text = (
                f"💳 *Subscription Plans*\n"
                f"━━━━━━━━━━━━━━━━━━━━━━\n"
                f"🆓 *Free* • 0 DB searches/day • 0 Name/ID searches\n\n"
                f"⭐ *Basic*\n"
                f"  🔍 Search DB: 10/day → up to 200 results\n"
                f"  🪪 Name/ID: 5/day → up to 50 results\n\n"
                f"💎 *Premium*\n"
                f"  🔍 Search DB: 15/day → up to 1,000 results + Full Scan ✅\n"
                f"  🪪 Name/ID: 10/day → up to 500 results\n\n"
                f"👑 *VIP*\n"
                f"  🔍 Search DB: Unlimited → up to 1,000,000 results + Full Scan ✅\n"
                f"  🪪 Name/ID: Unlimited → up to 100,000 results\n"
                f"━━━━━━━━━━━━━━━━━━━━━━\n"
                f"📩 Contact @yut3ev to upgrade."
            )
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("📋 طلب اشتراك / Subscribe" if is_ar else "📋 Subscribe Request", callback_data="user_subscribe")],
            [InlineKeyboardButton("🔙 القائمة / Menu", callback_data="user_home")],
        ]))
        return

    if data == "show_help":
        is_ar = get_lang(uid) == "ar"
        u     = get_user_cached(uid)
        tier_key = u[3] if u else "free"
        t  = TIERS.get(tier_key, TIERS["free"])
        nt = NAMEID_TIERS.get(tier_key, NAMEID_TIERS["free"])
        daily_left  = u[4]  if u else 0
        nameid_left = u[13] if u and len(u) > 13 else 0
        plan_info = (
            f"\n━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📦 {'باقتك الحالية' if is_ar else 'Your Current Plan'}: *{esc(t['label'])}*\n"
            f"🔍 {'البحث العادي' if is_ar else 'DB Search'}: `{daily_left}` {'متبقي' if is_ar else 'left today'} / `{t['daily']}` {'يومياً' if is_ar else '/day'}\n"
            f"🪪 Name/ID: `{nameid_left}` {'متبقي' if is_ar else 'left today'} / `{nt['daily_nameid']}` {'يومياً' if is_ar else '/day'}"
        )
        if is_ar:
            text = (
                f"ℹ️ *كيفية استخدام بوت DATA SCANNER*\n"
                f"━━━━━━━━━━━━━━━━━━━━━━\n"
                f"*🔍 البحث في قاعدة البيانات:*\n"
                f"1️⃣ اضغط *بحث في قاعدة البيانات*\n"
                f"2️⃣ اختر النوع: URL / دومين / إيميل / هاتف / إلخ\n"
                f"3️⃣ ابعت الكلمة المفتاحية ← البوت يُظهر العدد أولاً\n"
                f"4️⃣ أكّد لتحميل ملف `.txt` نظيف\n\n"
                f"*🪪 بحث بالاسم / الرقم القومي:*\n"
                f"1️⃣ اضغط *بحث بالاسم / الرقم القومي*\n"
                f"2️⃣ اختر *بحث بالاسم* أو *بحث بالرقم القومي*\n"
                f"3️⃣ ابعت الاسم أو الرقم القومي (14 رقماً)\n"
                f"4️⃣ البوت يُظهر العدد ← أكّد للتحميل\n\n"
                f"━━━━━━━━━━━━━━━━━━━━━━\n"
                f"⚠️ *ملاحظة:* الرقم القومي = 14 رقماً (مش رقم تليفون)"
                f"{plan_info}"
            )
        else:
            text = (
                f"ℹ️ *How to Use Data Scanner Bot*\n"
                f"━━━━━━━━━━━━━━━━━━━━━━\n"
                f"*🔍 Database Search:*\n"
                f"1️⃣ Tap *Search Database*\n"
                f"2️⃣ Choose type: URL / Domain / Email / Phone / etc.\n"
                f"3️⃣ Send keyword → bot shows result *count* first\n"
                f"4️⃣ Confirm to download clean `.txt` file\n\n"
                f"*🪪 Name / National ID Search:*\n"
                f"1️⃣ Tap *Name / National ID Search*\n"
                f"2️⃣ Choose *Search by Name* or *Search by National ID*\n"
                f"3️⃣ Send name or 14-digit national ID\n"
                f"4️⃣ Bot shows count → confirm to download\n\n"
                f"━━━━━━━━━━━━━━━━━━━━━━\n"
                f"⚠️ *Note:* National ID = 14 digits (not a phone number)"
                f"{plan_info}"
            )
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=back_user_kb(uid))
        return

    if data.startswith("confirm_search:"):
        parts_cb = data.split(":", 2)
        if len(parts_cb) < 3:
            await q.answer("❌ Invalid callback.", show_alert=True)
            return
        _, stype, keyword = parts_cb
        context.user_data["search_type"]  = stype
        context.user_data["confirmed_kw"] = keyword
        await q.edit_message_text(
            f"⏳ Starting scan for `{mesc(keyword)}`...",
            parse_mode="Markdown"
        )
        await do_search(update, context, keyword, stype, reply_to=q.message)
        return

    if data.startswith("confirm_nameid:"):
        parts_cb = data.split(":", 2)
        if len(parts_cb) < 3:
            await q.answer("❌ Invalid callback.", show_alert=True)
            return
        _, stype, keyword = parts_cb
        context.user_data["search_type"] = stype
        await q.edit_message_text("⏳ جاري تحضير الملف...", parse_mode="Markdown")
        await do_nameid_search(update, context, keyword, stype, reply_to=q.message)
        return

    if data.startswith("cancel_search"):
        await q.edit_message_text("❌ Search cancelled.", reply_markup=new_search_kb())
        return

    # ════ ADMIN CALLBACKS ════
    if not is_admin(uid):
        return

    if data == "adm_stats":
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT COUNT(*) FROM data_index")
                    tr = cur.fetchone()[0]
                    cur.execute("SELECT COUNT(*) FROM name_id_index")
                    tn = cur.fetchone()[0]
                    cur.execute("SELECT COUNT(*) FROM users")
                    tu = cur.fetchone()[0]
                    cur.execute("SELECT COUNT(*) FROM users WHERE is_banned=1")
                    tb = cur.fetchone()[0]
                    cur.execute("SELECT COUNT(*) FROM search_logs")
                    ts = cur.fetchone()[0]
                    cur.execute("SELECT COUNT(*) FROM uploaded_files")
                    tf = cur.fetchone()[0]
                    cur.execute("SELECT tier, COUNT(*) FROM users GROUP BY tier")
                    tc = cur.fetchall()
                    cur.execute(
                        "SELECT COALESCE(SUM(size_bytes),0), COALESCE(MAX(size_bytes),0), "
                        "COALESCE(AVG(records),0), COALESCE(MAX(records),0) FROM uploaded_files"
                    )
                    up_stats = cur.fetchone()
        except Exception as e:
            log.error(f"adm_stats error: {e}")
            await q.edit_message_text("❌ Stats error.", reply_markup=back_admin_kb())
            return

        total_sz, max_sz, avg_recs, max_recs = up_stats
        tier_lines = "\n".join([f"  {t}: `{c}`" for t, c in tc])
        await q.edit_message_text(
            f"📊 *Bot Statistics*\n━━━━━━━━━━━━━━━━━━━━━━\n"
            f"👥 Users: `{tu:,}` | 🚫 Banned: `{tb}`\n"
            f"🗄️ DB Records: `{tr:,}` | 🪪 Name/ID: `{tn:,}`\n"
            f"📁 Files: `{tf}` | 🔍 Searches: `{ts:,}`\n\n"
            f"📦 *Tier Breakdown:*\n{tier_lines}\n\n"
            f"📂 *Upload Stats:*\n"
            f"  💾 Total size: `{round((total_sz or 0)/1024/1024, 1)} MB`\n"
            f"  📄 Largest file: `{round((max_sz or 0)/1024/1024, 2)} MB`\n"
            f"  📊 Avg records/file: `{int(avg_recs or 0):,}`\n"
            f"  🏆 Max records: `{int(max_recs or 0):,}`",
            parse_mode="Markdown", reply_markup=back_admin_kb()
        )
        return

    if data == "adm_users" or data.startswith("adm_users_p") or data.startswith("adm_users_f"):
        page = 0
        tier_filter = None
        if data.startswith("adm_users_p"):
            try:
                page = int(data.split("_p")[1].split("_f")[0])
            except Exception:
                page = 0
        if "_f" in data:
            tier_filter = data.split("_f")[-1] or None
        elif data.startswith("adm_users_f"):
            tier_filter = data.replace("adm_users_f", "") or None

        per_page = 10
        offset   = page * per_page

        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    if tier_filter:
                        cur.execute("SELECT COUNT(*) FROM users WHERE tier=%s", (tier_filter,))
                    else:
                        cur.execute("SELECT COUNT(*) FROM users")
                    total_u = cur.fetchone()[0]

                    if tier_filter:
                        cur.execute(
                            "SELECT user_id, username, full_name, tier, daily_limit, credits, is_banned, last_search_at "
                            "FROM users WHERE tier=%s ORDER BY user_id DESC LIMIT %s OFFSET %s",
                            (tier_filter, per_page, offset)
                        )
                    else:
                        cur.execute(
                            "SELECT user_id, username, full_name, tier, daily_limit, credits, is_banned, last_search_at "
                            "FROM users ORDER BY user_id DESC LIMIT %s OFFSET %s",
                            (per_page, offset)
                        )
                    rows = cur.fetchall()
        except Exception as e:
            log.error(f"adm_users error: {e}")
            await q.edit_message_text("❌ Error loading users.", reply_markup=back_admin_kb())
            return

        filter_label = f" [{tier_filter}]" if tier_filter else ""
        text = f"👥 <b>Users{filter_label} — Page {page+1}</b> ({total_u} total)\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
        kb_rows = []
        for r in rows:
            uid_r, uname, fname, tier_r, dlimit, credits_r, banned, last_s = r
            icon = "🚫" if banned else "✅"
            last = (str(last_s) or "")[:10] if last_s else "never"
            text += (
                f"{icon} <code>{uid_r}</code> <b>{esc(tier_r)}</b> @{esc(uname or 'N/A')} "
                f"<i>{esc((fname or '')[:15])}</i> cr:{credits_r} last:{last}\n"
            )
            ban_lbl = "✅ Unban" if banned else "🔒 Ban"
            ban_cb  = f"quick_unban:{uid_r}" if banned else f"quick_ban:{uid_r}"
            kb_rows.append([
                InlineKeyboardButton(f"👤 {uid_r}", callback_data=f"quick_info:{uid_r}"),
                InlineKeyboardButton(ban_lbl,       callback_data=ban_cb),
            ])

        base_f = "adm_users_p0_f"
        filter_row = [
            InlineKeyboardButton("All",  callback_data="adm_users"),
            InlineKeyboardButton("⭐",   callback_data=f"{base_f}basic"),
            InlineKeyboardButton("💎",   callback_data=f"{base_f}premium"),
            InlineKeyboardButton("👑",   callback_data=f"{base_f}vip"),
        ]
        kb_rows.append(filter_row)
        nav_buttons = []
        f_suffix = f"_f{tier_filter}" if tier_filter else ""
        if page > 0:
            nav_buttons.append(InlineKeyboardButton("⬅️ Prev", callback_data=f"adm_users_p{page-1}{f_suffix}"))
        if offset + per_page < total_u:
            nav_buttons.append(InlineKeyboardButton("Next ➡️", callback_data=f"adm_users_p{page+1}{f_suffix}"))
        if nav_buttons:
            kb_rows.append(nav_buttons)
        kb_rows.append([InlineKeyboardButton("🔙 Admin Panel", callback_data="adm_home")])
        await q.edit_message_text(text[:4000], parse_mode="HTML", reply_markup=InlineKeyboardMarkup(kb_rows))
        return

    if data.startswith("quick_info:"):
        target = int(data.split(":")[1])
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT * FROM users WHERE user_id=%s", (target,))
                    row = cur.fetchone()
        except Exception:
            row = None
        if not row:
            await q.answer("User not found.", show_alert=True)
            return
        uid_r, uname, fname, tier_r, daily, credits_r, banned, expires, joined, lang_r, ref_by, ref_cnt, last_s, nameid_lim = row[:14]
        status = "🚫 Banned" if banned else "✅ Active"
        await q.answer(
            f"ID:{uid_r} | {tier_r} | cr:{credits_r}\n"
            f"daily:{daily} | nameid:{nameid_lim}\n"
            f"exp:{str(expires or 'none')[:10]} | {status}",
            show_alert=True
        )
        return

    if data.startswith("quick_ban:") or data.startswith("quick_unban:"):
        action_q, target_s = data.split(":", 1)
        target  = int(target_s)
        new_ban = 1 if action_q == "quick_ban" else 0
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("UPDATE users SET is_banned=%s WHERE user_id=%s", (new_ban, target))
                conn.commit()
        except Exception as e:
            log.error(f"quick_ban error: {e}")
        log_admin_op(uid, "quick_ban" if new_ban else "quick_unban", str(target))
        action_label = "🚫 Banned" if new_ban else "✅ Unbanned"
        await q.answer(f"{action_label} user {target}", show_alert=False)
        data = "adm_users"

    if data == "adm_logs" or data.startswith("adm_logs_p"):
        page = 0
        if data.startswith("adm_logs_p"):
            try:
                page = int(data.split("_p")[1])
            except Exception:
                page = 0
        per_page = 20
        offset   = page * per_page
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT COUNT(*) FROM search_logs")
                    total_logs = cur.fetchone()[0]
                    cur.execute(
                        "SELECT user_id, keyword, category, results, timestamp "
                        "FROM search_logs ORDER BY timestamp DESC LIMIT %s OFFSET %s",
                        (per_page, offset)
                    )
                    rows = cur.fetchall()
        except Exception as e:
            log.error(f"adm_logs error: {e}")
            rows = []
            total_logs = 0

        text = f"📜 <b>Search Logs — Page {page+1}</b> ({total_logs:,} total)\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
        for r in rows:
            cat_icon = {"email":"📧","phone":"📱","url":"🌐","nameid_name":"🪪"}.get(r[2], "🔍")
            text += f"{cat_icon} <code>{r[0]}</code> <code>{esc(r[1][:20])}</code> [{esc(r[2])}] — {r[3]}r @ {str(r[4])[:16]}\n"
        nav_buttons = []
        if page > 0:
            nav_buttons.append(InlineKeyboardButton("⬅️ Prev", callback_data=f"adm_logs_p{page-1}"))
        if offset + per_page < total_logs:
            nav_buttons.append(InlineKeyboardButton("Next ➡️", callback_data=f"adm_logs_p{page+1}"))
        kb_rows = [nav_buttons] if nav_buttons else []
        kb_rows.append([InlineKeyboardButton("🔙 Admin Panel", callback_data="adm_home")])
        await q.edit_message_text(text[:4000], parse_mode="HTML", reply_markup=InlineKeyboardMarkup(kb_rows))
        return

    if data == "adm_filelist" or data.startswith("adm_filelist_p"):
        page = 0
        if data.startswith("adm_filelist_p"):
            try:
                page = int(data.split("_p")[1])
            except Exception:
                page = 0
        per_page = 15
        offset   = page * per_page
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT COUNT(*) FROM uploaded_files")
                    total_files = cur.fetchone()[0]
                    cur.execute(
                        "SELECT id, original_name, records, size_bytes, uploaded_at "
                        "FROM uploaded_files ORDER BY uploaded_at DESC LIMIT %s OFFSET %s",
                        (per_page, offset)
                    )
                    rows = cur.fetchall()
        except Exception as e:
            log.error(f"adm_filelist error: {e}")
            rows = []
            total_files = 0

        if not rows:
            text = "📁 No files uploaded yet."
        else:
            text = f"🗂️ *Uploaded Files — Page {page+1}* ({total_files} total)\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
            for fid, fname, recs, sz, ts in rows:
                kb = round((sz or 0) / 1024, 1)
                text += f"`#{fid}` 📄 `{mesc(fname)}`\n    {(recs or 0):,} recs | {kb} KB | {str(ts)[:16]}\n\n"

        nav_buttons = []
        if page > 0:
            nav_buttons.append(InlineKeyboardButton("⬅️ Prev", callback_data=f"adm_filelist_p{page-1}"))
        if offset + per_page < total_files:
            nav_buttons.append(InlineKeyboardButton("Next ➡️", callback_data=f"adm_filelist_p{page+1}"))
        kb_rows = [nav_buttons] if nav_buttons else []
        kb_rows.append([InlineKeyboardButton("🔙 Admin Panel", callback_data="adm_home")])
        await q.edit_message_text(text[:4000], parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(kb_rows))
        return

    if data == "adm_upload_info":
        await q.edit_message_text(
            f"📂 *Upload Data to Database*\n━━━━━━━━━━━━━━━━━━━━━━\n"
            f"✅ *Supported Formats:*\n"
            f"  • `.txt` — one entry per line\n"
            f"  • `.csv` — all columns indexed\n"
            f"  • `.xlsx/.xls` — all cells indexed\n"
            f"  • `.json` — all string values indexed\n\n"
            f"🪪 *Excel with Name + National ID columns:*\n"
            f"  Automatically also indexed in Name/ID table.\n"
            f"  Columns are auto-detected by header keywords.\n\n"
            f"📌 *Small files (under 20 MB):* Send directly here.\n"
            f"📦 *Large files:* Use `uploader.py` script.",
            parse_mode="Markdown", reply_markup=back_admin_kb()
        )
        return

    if data == "adm_reset_daily":
        do_daily_reset()
        await q.edit_message_text("✅ Daily limits reset for all users (Search DB + Name/ID).", reply_markup=back_admin_kb())
        return

    if data == "adm_advanced_stats":
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT user_id, COUNT(*) as cnt FROM search_logs GROUP BY user_id ORDER BY cnt DESC LIMIT 5"
                    )
                    top_searchers = cur.fetchall()
                    today = datetime.utcnow().strftime("%Y-%m-%d")
                    cur.execute(
                        "SELECT COUNT(*) FROM search_logs WHERE timestamp LIKE %s", (f"{today}%",)
                    )
                    today_searches = cur.fetchone()[0]
                    cur.execute(
                        "SELECT keyword, COUNT(*) as cnt FROM search_logs GROUP BY keyword ORDER BY cnt DESC LIMIT 5"
                    )
                    top_keywords = cur.fetchall()
                    cur.execute(
                        "SELECT COUNT(*) FROM users WHERE joined_at LIKE %s", (f"{today}%",)
                    )
                    new_users_today = cur.fetchone()[0]
                    cur.execute(
                        "SELECT category, COUNT(*) FROM search_logs GROUP BY category ORDER BY COUNT(*) DESC LIMIT 5"
                    )
                    by_type = cur.fetchall()
                    cur.execute(
                        "SELECT SUBSTRING(timestamp, 12, 2) as hr, COUNT(*) as cnt "
                        "FROM search_logs GROUP BY hr ORDER BY cnt DESC LIMIT 3"
                    )
                    peak_hours = cur.fetchall()
                    cur.execute(
                        "SELECT user_id, referral_count FROM users WHERE referral_count > 0 "
                        "ORDER BY referral_count DESC LIMIT 5"
                    )
                    top_referrers = cur.fetchall()
                    cur.execute("SELECT COALESCE(SUM(referral_count), 0) FROM users")
                    total_referrals = cur.fetchone()[0]
                    cur.execute(
                        "SELECT SUBSTRING(timestamp, 1, 10) as day, COUNT(*) "
                        "FROM search_logs WHERE timestamp >= CURRENT_DATE - INTERVAL '7 days' "
                        "GROUP BY day ORDER BY day DESC"
                    )
                    week_trend = cur.fetchall()
        except Exception as e:
            log.error(f"adm_advanced_stats error: {e}")
            await q.edit_message_text("❌ Stats error.", reply_markup=back_admin_kb())
            return

        top_s_lines   = "\n".join([f"  `{r[0]}` — {r[1]}x" for r in top_searchers]) or "  No data"
        top_k_lines   = "\n".join([f"  `{mesc(r[0][:20])}` — {r[1]}x" for r in top_keywords]) or "  No data"
        by_type_lines = "\n".join([f"  {r[0] or 'N/A'}: `{r[1]}`" for r in by_type]) or "  No data"
        peak_lines    = " | ".join([f"`{r[0]}:00` ({r[1]})" for r in peak_hours]) or "No data"
        ref_lines     = "\n".join([f"  `{r[0]}` — {r[1]} referrals" for r in top_referrers]) or "  No referrals yet"
        trend_lines   = " | ".join([f"`{str(r[0])[5:]}` {r[1]}" for r in week_trend]) or "No data"

        await q.edit_message_text(
            f"📈 *Advanced Statistics*\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"📅 *Today ({today}):*\n"
            f"  🔍 Searches: `{today_searches}` | 👤 New users: `{new_users_today}`\n\n"
            f"📊 *7-Day Trend:*\n  {trend_lines}\n\n"
            f"⏰ *Peak Hours (UTC):* {peak_lines}\n\n"
            f"🏆 *Top Searchers:*\n{top_s_lines}\n\n"
            f"🔑 *Top Keywords:*\n{top_k_lines}\n\n"
            f"📂 *Searches by Type:*\n{by_type_lines}\n\n"
            f"🔗 *Referrals:* `{total_referrals}` total\n{ref_lines}",
            parse_mode="Markdown", reply_markup=back_admin_kb()
        )
        return

    if data == "adm_broadcast":
        context.user_data["admin_action"] = "broadcast"
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT COUNT(*) FROM users WHERE is_banned=0")
                    user_count = cur.fetchone()[0]
        except Exception:
            user_count = 0
        await q.edit_message_text(
            f"📢 *Broadcast Message*\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"👥 Will be sent to *{user_count}* active users.\n\n"
            f"✏️ Send your message now:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_home")]])
        )
        return

    if data == "adm_msg_user":
        context.user_data["admin_action"] = "msg_user"
        await q.edit_message_text(
            "✉️ *Message Specific User*\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "Send: `USER_ID Your message`\n"
            "Example: `123456789 Your account is ready!`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_home")]])
        )
        return

    if data == "my_history":
        is_ar = get_lang(uid) == "ar"
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT keyword, category, results, timestamp FROM search_logs "
                        "WHERE user_id=%s ORDER BY timestamp DESC LIMIT 15",
                        (uid,)
                    )
                    rows = cur.fetchall()
                    cur.execute("SELECT COUNT(*) FROM search_logs WHERE user_id=%s", (uid,))
                    total = cur.fetchone()[0]
        except Exception:
            rows = []
            total = 0

        if not rows:
            txt = "📜 *بحوثي الأخيرة*\n\nلا توجد بحوث بعد." if is_ar else "📜 *My Search History*\n\nNo searches yet."
        else:
            title = "📜 *بحوثي الأخيرة*" if is_ar else "📜 *My Search History*"
            txt = f"{title}\n━━━━━━━━━━━━━━━━━━━━━━\n({'إجمالي' if is_ar else 'Total'}: {total})\n\n"
            type_icons = {
                "email":"📧","phone":"📱","url":"🌐","domain":"🌍",
                "login":"👤","username":"📝","password":"🔑","all":"🔎",
                "nameid_name":"🪪","nameid_national_id":"🪪","nameid_partial_id":"🔢",
            }
            for kw, cat, res, ts in rows:
                icon = type_icons.get(cat, "🔍")
                cat_label = cat.replace("nameid_", "").replace("_", " ").upper()
                txt += f"{icon} `{mesc(kw[:22])}` `{cat_label}` — *{res}* | {str(ts)[:10]}\n"

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🗑️ مسح التاريخ" if is_ar else "🗑️ Clear History", callback_data="clear_history")],
            [InlineKeyboardButton("🔙 القائمة الرئيسية" if is_ar else "🔙 Main Menu", callback_data="user_home")],
        ])
        await q.edit_message_text(txt[:4000], parse_mode="Markdown", reply_markup=kb)
        return

    if data == "clear_history":
        is_ar = get_lang(uid) == "ar"
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("DELETE FROM search_logs WHERE user_id=%s", (uid,))
                conn.commit()
        except Exception as e:
            log.error(f"clear_history error: {e}")
        txt = "✅ *تم مسح تاريخ البحث بنجاح.*" if is_ar else "✅ *Search history cleared.*"
        await q.edit_message_text(txt, parse_mode="Markdown", reply_markup=back_user_kb(uid))
        return

    if data == "my_referral":
        is_ar     = get_lang(uid) == "ar"
        ref_count = get_referral_stats(uid)
        bot_username = (await context.bot.get_me()).username
        ref_link  = f"https://t.me/{bot_username}?start={uid}"
        if is_ar:
            txt = (
                f"🔗 *رابط الإحالة الخاص بك*\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
                f"شارك هذا الرابط مع أصدقائك:\n`{ref_link}`\n\n"
                f"👥 الأصدقاء الذين دعوتهم: `{ref_count}`\n"
                f"💰 تكسب *{REFERRAL_CREDITS}* رصيد لكل صديق جديد!\n\n"
                f"كل صديق يفتح البوت عبر رابطك تحصل على رصيد تلقائياً ✅"
            )
        else:
            txt = (
                f"🔗 *Your Referral Link*\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
                f"Share this link with friends:\n`{ref_link}`\n\n"
                f"👥 Friends referred: `{ref_count}`\n"
                f"💰 Earn *{REFERRAL_CREDITS}* credits per new user!\n\n"
                f"Every friend who opens the bot via your link earns you credits ✅"
            )
        await q.edit_message_text(txt, parse_mode="Markdown", reply_markup=back_user_kb(uid))
        return

    if data == "my_id":
        is_ar = get_lang(uid) == "ar"
        txt = (
            f"🆔 *معرفك على تيليجرام:*\n\n`{uid}`\n\n_شارك هذا مع الأدمن لإدارة حسابك_"
            if is_ar else
            f"🆔 *Your Telegram ID:*\n\n`{uid}`\n\n_Share this with the admin to manage your account_"
        )
        await q.edit_message_text(txt, parse_mode="Markdown", reply_markup=back_user_kb(uid))
        return

    # ════ PHASE 1 CALLBACKS ════
    if data == "my_points":
        await show_points(update, context)
        return

    if data == "my_favorites":
        await show_favorites(update, context, query=q)
        return

    if data == "my_support":
        await show_support_menu(update, context, query=q)
        return

    if data == "my_affiliate":
        await show_affiliate_stats(update, context, query=q)
        return

    if data == "my_orders":
        await show_my_orders(q, uid)
        return

    if data == "my_achievements":
        await show_gamification(update, context, query=q)
        return

    if data.startswith("onboard_"):
        step = data.replace("onboard_", "")
        if step == "done":
            await q.edit_message_text(
                "✅ *Tutorial complete!*\n\nUse the menu below to get started.",
                parse_mode="Markdown",
                reply_markup=user_main_kb(uid)
            )
        else:
            await send_onboarding(context.bot, uid, int(step))
            try:
                await q.message.delete()
            except Exception:
                pass
        return

    if data == "adm_stats":
        if not is_admin(uid):
            return
        await show_admin_stats(update, context, query=q)
        return

    if data == "adm_weekly_report":
        if not is_admin(uid):
            return
        await show_weekly_report(q, uid)
        return

    if data == "adm_top_users":
        if not is_admin(uid):
            return
        await show_top_users(q, uid)
        return

    if data == "adm_run_dedup":
        if not is_admin(uid):
            return
        await q.edit_message_text("⏳ *Running deduplication...*\n\nThis may take a few minutes.", parse_mode="Markdown")
        loop = asyncio.get_running_loop()
        removed_d = await loop.run_in_executor(_executor, run_deduplication)
        removed_n = await loop.run_in_executor(_executor, run_nameid_deduplication)
        await q.edit_message_text(
            f"✅ *Deduplication Complete!*\n\n"
            f"🗑️ data_index   : `{removed_d:,}` duplicates removed\n"
            f"🗑️ name_id_index: `{removed_n:,}` duplicates removed",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="adm_stats")]])
        )
        return

    if data == "adm_scheduled":
        if not is_admin(uid):
            return
        await show_scheduled_imports(q, uid)
        return

    if data == "adm_add_scheduled":
        if not is_admin(uid):
            return
        context.user_data["admin_action"] = "add_scheduled_import"
        await q.edit_message_text(
            "⏰ *Add Scheduled Import*\n\n"
            "Send in this format:\n`URL LABEL HOURS`\n\n"
            "Example:\n`https://example.com/data.txt MySource 24`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_scheduled")]])
        )
        return

    if data.startswith("del_scheduled_"):
        if not is_admin(uid):
            return
        sid = int(data.replace("del_scheduled_", ""))
        delete_scheduled_import(sid)
        await show_scheduled_imports(q, uid)
        return

    # ════ 2FA CALLBACKS ════
    if data == "my_security":
        await show_security_settings(update, context, query=q)
        return

    if data == "2fa_enable":
        set_2fa_enabled(uid, True)
        await q.answer("✅ 2FA Enabled!", show_alert=True)
        await show_security_settings(update, context, query=q)
        log_activity(uid, "2fa_enabled")
        return

    if data == "2fa_disable":
        set_2fa_enabled(uid, False)
        await q.answer("❌ 2FA Disabled.", show_alert=True)
        await show_security_settings(update, context, query=q)
        log_activity(uid, "2fa_disabled")
        return

    if data == "2fa_test":
        code = await send_2fa_code(context.bot, uid, "test")
        context.user_data["2fa_verify_purpose"] = "test"
        await q.edit_message_text(
            "🔐 *2FA Test*\n\nA code was sent to you.\nEnter it to verify:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="my_security")]])
        )
        context.user_data["waiting_2fa"] = True
        return

    # ════ REGEX SEARCH CALLBACK ════
    if data == "go_regex_search":
        is_ar = get_lang(uid) == "ar"
        u     = get_user_cached(uid)
        tier  = u[3] if u else "free"
        if tier not in ("premium", "vip") and not is_admin(uid):
            await q.edit_message_text(
                "🔍 *Regex Search — Premium Feature*\n\nRequires Premium or VIP plan.",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("💳 Upgrade", callback_data="show_plans")],
                    [InlineKeyboardButton("🔙 Back",    callback_data="user_home")],
                ])
            )
            return
        context.user_data["search_type"] = "regex"
        await q.edit_message_text(
            "🔍 *Regex Search*\n\n"
            "Send a regular expression pattern.\n\n"
            "Examples:\n"
            "`^admin.*@gmail\\.com`\n"
            "`\\+20[0-9]{10}`\n"
            "`.*:password123$`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="user_home")]])
        )
        return

    # ════ FILTER RESULTS CALLBACK ════
    if data.startswith("filter_search:"):
        parts_f = data.split(":", 3)
        if len(parts_f) < 4:
            await q.answer("❌ Invalid", show_alert=True)
            return
        _, cat, keyword, stype = parts_f
        await q.answer("⏳ Filtering...", show_alert=False)
        u     = get_user_cached(uid)
        tier  = u[3] if u else "free"
        limit = TIERS[tier]["max_results"] if not is_admin(uid) else 1_000_000
        results = await asyncio.get_running_loop().run_in_executor(
            _executor, lambda: smart_search(keyword, stype, limit)
        )
        if cat != "all":
            results = filter_results_by_category(results, cat)
        if not results:
            await q.edit_message_text(
                f"🔍 *No results after filtering by `{cat}`.*",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Menu", callback_data="user_home")]])
            )
            return
        cat_label = DATA_CATEGORIES.get(cat, {}).get("label", cat)
        content   = build_result_txt(keyword, results, stype)
        content   = add_watermark(content, uid)
        safe_kw   = re.sub(r"[^\w\-]", "_", keyword)[:20]
        filename  = f"{safe_kw}_{cat}_{len(results)}_filtered.txt"
        tmppath   = os.path.join(FILES_DIR, f"tmp_filter_{uid}.txt")
        with open(tmppath, "w", encoding="utf-8") as f:
            f.write(content)
        caption = (
            f"✅ *Filtered Results*\n"
            f"🎯 `{mesc(keyword)}` | {cat_label}\n"
            f"📊 `{len(results):,}` records"
        )
        msg2 = await q.message.reply_document(
            document=open(tmppath, "rb"),
            filename=filename,
            caption=caption,
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Menu", callback_data="user_home")]])
        )
        return

    # ════ ADMIN URL IMPORT ════
    if data == "adm_url_import":
        if not is_admin(uid):
            return
        await show_url_import_menu(q, uid, context)
        return

    # ════ ADMIN USER ACTIVITY LOG ════
    if data.startswith("adm_user_log_"):
        if not is_admin(uid):
            return
        target_uid = int(data.replace("adm_user_log_", ""))
        await show_user_activity(q, target_uid, uid)
        return

    # ── Security / 2FA callbacks ─────────────────────────
    if data == "my_security":
        await show_security_settings(update, context, query=q)
        return

    if data == "2fa_enable":
        set_2fa_enabled(uid, True)
        await q.answer("✅ 2FA Enabled!", show_alert=True)
        await show_security_settings(update, context, query=q)
        return

    if data == "2fa_disable":
        set_2fa_enabled(uid, False)
        await q.answer("❌ 2FA Disabled.", show_alert=True)
        await show_security_settings(update, context, query=q)
        return

    if data == "2fa_test":
        code = await send_2fa_code(context.bot, uid, "test")
        context.user_data["verify_2fa"] = True
        await q.edit_message_text(
            f"🔐 *2FA Test*\n\nA code was sent to you.\nEnter it to verify:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="my_security")]])
        )
        return

    # ── Admin: regex search, URL import ─────────────────
    if data == "adm_regex_search":
        if not is_admin(uid):
            return
        context.user_data["admin_action"] = "regex_search"
        await q.edit_message_text(
            "🔎 *Regex Search*\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "Send a regex pattern:\n\n"
            "Examples:\n"
            "`.*@gmail\\.com:.*`  → Gmail accounts\n"
            "`\\+20\\d{10}`        → Egyptian phones\n"
            "`\\d{14}`            → National IDs",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_home")]])
        )
        return

    if data == "adm_url_import":
        if not is_admin(uid):
            return
        await show_url_import_menu(q, uid, context)
        return

    # ── Auto-renewal callbacks ────────────────────────────
    if data == "my_autorenewal":
        await show_autorenewal_settings(update, context, query=q)
        return

    if data == "autorenewal_on":
        ar = get_auto_renewal(uid)
        set_auto_renewal(uid, True, ar.get("method"), ar.get("plan"), ar.get("duration"))
        await q.answer("✅ Auto-renewal enabled!", show_alert=True)
        await show_autorenewal_settings(update, context, query=q)
        return

    if data == "autorenewal_off":
        set_auto_renewal(uid, False)
        await q.answer("❌ Auto-renewal disabled.", show_alert=True)
        await show_autorenewal_settings(update, context, query=q)
        return

    if data == "autorenewal_setup":
        is_ar2 = get_lang(uid) == "ar"
        rows_ar = []
        for pk in PLAN_PRICES:
            for dk in PLAN_DURATIONS:
                usd = calc_price(pk, dk, "USD")
                rows_ar.append([InlineKeyboardButton(
                    f"{PLAN_PRICES[pk]['label']} — {PLAN_DURATIONS[dk]['label']} (${usd})",
                    callback_data=f"ar_set_{pk}_{dk}"
                )])
        rows_ar.append([InlineKeyboardButton("🔙 Back", callback_data="my_autorenewal")])
        await q.edit_message_text(
            f"⚙️ *{'إعداد التجديد التلقائي' if is_ar2 else 'Setup Auto-Renewal'}*\n\n"
            f"{'اختر الباقة والمدة المفضلة:' if is_ar2 else 'Choose preferred plan & duration:'}",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(rows_ar)
        )
        return

    if data.startswith("ar_set_"):
        parts_ar = data.split("_", 3)
        if len(parts_ar) >= 4:
            plan_ar, dur_ar = parts_ar[2], parts_ar[3]
            ar2 = get_auto_renewal(uid)
            set_auto_renewal(uid, True, ar2.get("method", "vodafone"), plan_ar, dur_ar)
            await q.answer(f"✅ Set: {plan_ar} / {dur_ar}", show_alert=True)
            await show_autorenewal_settings(update, context, query=q)
        return

    # ── Coupon callbacks ──────────────────────────────────
    if data == "my_coupon":
        is_ar2 = get_lang(uid) == "ar"
        context.user_data["user_action"] = "apply_coupon"
        await q.edit_message_text(
            f"🏷 *{'أدخل كود الكوبون' if is_ar2 else 'Enter Coupon Code'}*\n\n"
            f"{'أرسل الكود:' if is_ar2 else 'Send your coupon code:'}",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="user_home")]])
        )
        return

    if data == "adm_coupons":
        if not is_admin(uid):
            return
        await show_admin_coupons(update, context, query=q)
        return

    if data == "adm_create_coupon":
        if not is_admin(uid):
            return
        context.user_data["admin_action"] = "create_coupon"
        await q.edit_message_text(
            "🏷 *Create Coupon*\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "Format: `CODE TYPE VALUE MAX_USES VALID_DAYS [PLAN]`\n\n"
            "Types: `percent` | `fixed` | `days` | `plan`\n\n"
            "Examples:\n"
            "`SAVE20 percent 20 100 30`\n"
            "`EXTRA7 days 7 50 90`\n"
            "`VIP4FREE plan 0 5 14 vip`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_coupons")]])
        )
        return

    if data == "adm_delete_coupon":
        if not is_admin(uid):
            return
        context.user_data["admin_action"] = "delete_coupon"
        await q.edit_message_text(
            "🗑️ *Delete Coupon*\n\nSend the coupon code to deactivate:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_coupons")]])
        )
        return

    if data == "adm_user_log_prompt":
        if not is_admin(uid):
            return
        context.user_data["admin_action"] = "view_user_log"
        await q.edit_message_text(
            "📋 *User Activity Log*\n\nSend the User ID to view their activity:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_home")]])
        )
        return

    # ── Filter results callback ──────────────────────────
    if data.startswith("show_filter:"):
        parts_sf = data.split(":", 2)
        if len(parts_sf) < 3:
            return
        _, keyword, stype = parts_sf
        await show_filter_menu(q, keyword, stype, uid, context)
        return

    if data.startswith("filter_search:"):
        parts_f = data.split(":", 3)
        if len(parts_f) < 4:
            return
        _, cat, keyword, stype = parts_f
        u     = get_user_cached(uid)
        tier  = u[3] if u else "free"
        limit = TIERS[tier]["max_results"] if not is_admin(uid) else 1_000_000

        await q.answer("⏳ Filtering...", show_alert=False)
        msg = await q.message.reply_text(f"⏳ Filtering by `{cat}`...", parse_mode="Markdown")
        results = await asyncio.get_running_loop().run_in_executor(
            _executor, lambda: smart_search(keyword, stype, limit)
        )
        if cat != "all":
            results = filter_results_by_category(results, cat)

        if not results:
            await msg.edit_text(f"🔍 No results for category `{cat}`.", parse_mode="Markdown",
                                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="user_home")]]))
            return

        content  = build_result_txt(keyword, results, stype)
        content  = add_watermark(content, uid)
        cat_lbl  = DATA_CATEGORIES.get(cat, {}).get("label", cat)
        safe_kw  = re.sub(r"[^\w\-]", "_", keyword)[:25]
        filename = f"{safe_kw}_{cat}_{len(results)}_results.txt"
        tmppath  = os.path.join(FILES_DIR, f"tmp_filter_{uid}.txt")
        with open(tmppath, "w", encoding="utf-8") as f:
            f.write(content)
        caption = (
            f"✅ *Filtered Results*\n"
            f"🎯 Keyword: `{mesc(keyword)}`\n"
            f"🏷 Category: {cat_lbl}\n"
            f"📊 Results: `{len(results):,}`"
        )
        await msg.delete()
        await safe_send_document(
            q.message.reply_document, tmppath, filename, caption,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Menu", callback_data="user_home")]])
        )
        return

    if data == "go_bulk_search":
        await show_bulk_search_menu(update, context, query=q)
        return

    if data == "support_new_ticket":
        context.user_data["support_action"] = "new_ticket_subject"
        is_ar = get_lang(uid) == "ar"
        await q.edit_message_text(
            "🎫 *New Support Ticket*\n\n✏️ Send the subject of your ticket:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="my_support")]])
        )
        return

    if data.startswith("del_fav_"):
        fav_id = int(data.replace("del_fav_", ""))
        delete_favorite(fav_id, uid)
        await show_favorites(update, context, query=q)
        return

    if data.startswith("adm_reply_ticket_"):
        if not is_admin(uid):
            return
        ticket_id = int(data.replace("adm_reply_ticket_", ""))
        context.user_data["admin_action"] = f"reply_ticket_{ticket_id}"
        await q.edit_message_text(
            f"💬 *Reply to Ticket #{ticket_id}*\n\nSend your reply:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_home")]])
        )
        return

    if data.startswith("adm_approve_req_"):
        if not is_admin(uid):
            return
        req_id = int(data.replace("adm_approve_req_", ""))
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT user_id, requested_tier FROM sub_requests WHERE id=%s", (req_id,))
                    row = cur.fetchone()
                    if row:
                        target_uid, tier = row
                        t  = TIERS.get(tier, TIERS["free"])
                        nt = NAMEID_TIERS.get(tier, NAMEID_TIERS["free"])
                        cur.execute(
                            "UPDATE users SET tier=%s, daily_limit=%s, daily_nameid_limit=%s WHERE user_id=%s",
                            (tier, t["daily"], nt["daily_nameid"], target_uid)
                        )
                        cur.execute("UPDATE sub_requests SET status='approved' WHERE id=%s", (req_id,))
                conn.commit()
            invalidate_user_cache(target_uid)
            await context.bot.send_message(
                chat_id=target_uid,
                text=f"✅ *Subscription Approved!*\n\nYour account has been upgraded to *{tier}*.\n\nPress /start to refresh.",
                parse_mode="Markdown"
            )
            await q.edit_message_text(f"✅ Request #{req_id} approved — User upgraded to *{tier}*.", parse_mode="Markdown")
        except Exception as e:
            await q.edit_message_text(f"❌ Error: {e}")
        return

    if data.startswith("adm_reject_req_"):
        if not is_admin(uid):
            return
        req_id = int(data.replace("adm_reject_req_", ""))
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT user_id FROM sub_requests WHERE id=%s", (req_id,))
                    row = cur.fetchone()
                    cur.execute("UPDATE sub_requests SET status='rejected' WHERE id=%s", (req_id,))
                conn.commit()
            if row:
                await context.bot.send_message(
                    chat_id=row[0],
                    text="❌ *Subscription request rejected.*\n\nContact admin for more info.",
                    parse_mode="Markdown"
                )
            await q.edit_message_text(f"❌ Request #{req_id} rejected.")
        except Exception as e:
            await q.edit_message_text(f"❌ Error: {e}")
        return

    if data == "adm_tickets":
        if not is_admin(uid):
            return
        await show_admin_tickets(update, context, query=q)
        return

    if data.startswith("export_excel:"):
        parts_cb = data.split(":", 2)
        if len(parts_cb) < 3:
            await q.answer("❌ Invalid.", show_alert=True)
            return
        _, kw, stype = parts_cb
        u     = get_user_cached(uid)
        tier  = u[3] if u else "free"
        limit = TIERS[tier]["max_results"] if not is_admin(uid) else 1_000_000
        await q.answer("⏳ Generating Excel...", show_alert=False)
        msg = await q.message.reply_text("📊 *Generating Excel file...*", parse_mode="Markdown")
        results = await asyncio.get_running_loop().run_in_executor(
            _executor, lambda: smart_search(kw, stype, limit)
        )
        if not results:
            await msg.edit_text("❌ No results to export.")
            return
        xlsx_path = await asyncio.get_running_loop().run_in_executor(
            _executor, lambda: build_result_excel(kw, results, stype, uid)
        )
        safe_kw  = re.sub(r"[^\w\-]", "_", kw)[:25]
        filename = f"{safe_kw}_{len(results)}_results.xlsx"
        caption  = f"📊 *Excel Export*\n🎯 `{mesc(kw)}` | `{stype.upper()}`\n📊 `{len(results):,}` records"
        await msg.delete()
        await safe_send_document(
            q.message.reply_document, xlsx_path, filename, caption,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Menu", callback_data="user_home")]])
        )
        return

    if data.startswith("save_fav:"):
        parts_cb = data.split(":", 2)
        if len(parts_cb) < 3:
            return
        _, stype, kw = parts_cb
        saved = save_favorite(uid, kw, stype)
        await q.answer("⭐ Saved to favorites!" if saved else "❌ Max 20 favorites reached.", show_alert=True)
        return

    if data == "adm_set_expiry":
        context.user_data["admin_action"] = "set_expiry"
        await q.edit_message_text(
            "📅 *Set Subscription Expiry*\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "Send: `USER_ID DAYS`\n"
            "Example: `123456789 30` → sets expiry to 30 days from now\n\n"
            "Send `USER_ID 0` to clear expiry.",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_home")]])
        )
        return

    if data == "adm_filter_logs":
        context.user_data["admin_action"] = "filter_logs"
        await q.edit_message_text(
            "🔎 *Filter Search Logs*\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "Send a *User ID* to see their searches:\n`123456789`\n\n"
            "Or send a *keyword* to find searches containing it:\n`gmail.com`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_home")]])
        )
        return

    if data == "adm_export_csv":
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT user_id, username, full_name, tier, daily_limit, credits, is_banned, "
                        "expires_at, joined_at, lang, referral_count FROM users ORDER BY user_id DESC"
                    )
                    rows = cur.fetchall()
        except Exception as e:
            log.error(f"adm_export_csv error: {e}")
            rows = []

        import csv
        import io as _io
        output = _io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["user_id","username","full_name","tier","daily_limit","credits","is_banned","expires_at","joined_at","lang","referral_count"])
        writer.writerows(rows)
        csv_bytes = output.getvalue().encode("utf-8-sig")
        filename  = f"users_export_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.csv"
        await context.bot.send_document(
            chat_id=uid,
            document=_io.BytesIO(csv_bytes),
            filename=filename,
            caption=f"📤 *Users Export*\n{len(rows):,} users exported.",
            parse_mode="Markdown"
        )
        log_admin_op(uid, "export_csv", "users", f"{len(rows)} rows")
        await q.edit_message_text("✅ CSV sent to you.", reply_markup=back_admin_kb())
        return

    if data == "adm_backup":
        try:
            dest = backup_db()
            backups = sorted([f for f in os.listdir(BACKUP_DIR) if f.endswith(".txt")], reverse=True)
            await q.edit_message_text(
                f"💾 *Backup Note Created!*\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
                f"📄 File: `{os.path.basename(dest)}`\n"
                f"ℹ️ For PostgreSQL, run: `pg_dump -U postgres scanner > backup.sql`\n"
                f"🗂️ Total markers kept: `{len(backups)}`",
                parse_mode="Markdown", reply_markup=back_admin_kb()
            )
            log_admin_op(uid, "backup_db", dest, "pg marker")
        except Exception as e:
            await q.edit_message_text(f"❌ Backup failed: `{e}`", parse_mode="Markdown", reply_markup=back_admin_kb())
        return

    if data == "adm_sub_requests":
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id, user_id, username, full_name, requested_tier, status, timestamp "
                        "FROM sub_requests ORDER BY timestamp DESC LIMIT 20"
                    )
                    rows = cur.fetchall()
        except Exception:
            rows = []

        if not rows:
            text = "📋 *طلبات الاشتراك*\n\nلا توجد طلبات حتى الآن."
        else:
            text = "📋 *طلبات الاشتراك*\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
            for r in rows:
                status_icon = "⏳" if r[5] == "pending" else ("✅" if r[5] == "approved" else "❌")
                text += (
                    f"{status_icon} `#{r[0]}` | <code>{r[1]}</code>\n"
                    f"  👤 {esc(r[3] or 'User')} (@{esc(r[2] or 'N/A')})\n"
                    f"  📦 Tier: `{esc(r[4])}` | {str(r[6])[:10]}\n\n"
                )

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ قبول طلب",  callback_data="adm_approve_sub"),
             InlineKeyboardButton("❌ رفض طلب",   callback_data="adm_reject_sub")],
            [InlineKeyboardButton("🔙 Admin Panel", callback_data="adm_home")],
        ])
        await q.edit_message_text(text[:4000], parse_mode="HTML", reply_markup=kb)
        return

    if data == "adm_approve_sub":
        context.user_data["admin_action"] = "sub_request_approve"
        await q.edit_message_text(
            "✅ *قبول طلب اشتراك*\n\nأرسل: `REQUEST_ID` أو `REQUEST_ID TIER` لتغيير الباقة\nمثال: `5` أو `5 premium`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_sub_requests")]])
        )
        return

    if data == "adm_reject_sub":
        context.user_data["admin_action"] = "sub_request_reject"
        await q.edit_message_text(
            "❌ *رفض طلب اشتراك*\n\nأرسل رقم الطلب:\nمثال: `5`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_sub_requests")]])
        )
        return

    if data == "adm_op_logs":
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT admin_id, action, target, details, timestamp "
                        "FROM admin_op_logs ORDER BY timestamp DESC LIMIT 30"
                    )
                    rows = cur.fetchall()
        except Exception:
            rows = []

        if not rows:
            text = "📜 *سجل عمليات الأدمن*\n\nلا توجد عمليات مسجلة بعد."
        else:
            text = "📜 <b>سجل عمليات الأدمن</b>\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
            for r in rows:
                text += (
                    f"👤 <code>{r[0]}</code> → <b>{esc(r[1])}</b>\n"
                    f"  🎯 {esc(r[2])} | {esc(r[3] or '')} | {str(r[4])[:16]}\n\n"
                )
        await q.edit_message_text(text[:4000], parse_mode="HTML", reply_markup=back_admin_kb())
        return

    if data == "adm_bot_status":
        import platform, sys
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT COUNT(*) FROM data_index")
                    total_records = cur.fetchone()[0]
                    cur.execute("SELECT COUNT(*) FROM name_id_index")
                    total_nameid = cur.fetchone()[0]
                    cur.execute("SELECT COUNT(*) FROM users")
                    total_users = cur.fetchone()[0]
                    cur.execute("SELECT COUNT(*) FROM uploaded_files")
                    total_files = cur.fetchone()[0]
                    cur.execute("SELECT COUNT(*) FROM search_logs")
                    total_searches = cur.fetchone()[0]
                    cur.execute("SELECT COALESCE(SUM(size_bytes), 0) FROM uploaded_files")
                    files_size = cur.fetchone()[0]
        except Exception:
            total_records = total_nameid = total_users = total_files = total_searches = files_size = 0

        now_utc = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        await q.edit_message_text(
            f"⚡ *Bot Status*\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"🟢 *Status:* Online\n"
            f"🕐 *Time (UTC):* `{now_utc}`\n"
            f"🐍 *Python:* `{sys.version.split()[0]}`\n"
            f"🖥️ *OS:* `{platform.system()} {platform.release()}`\n"
            f"🐘 *DB:* PostgreSQL\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🗄️ DB Records: `{total_records:,}`\n"
            f"🪪 Name/ID:    `{total_nameid:,}`\n"
            f"👥 Users:      `{total_users:,}`\n"
            f"📁 Files:      `{total_files}`\n"
            f"🔍 Searches:   `{total_searches:,}`\n\n"
            f"📦 Files Size: `{round((files_size or 0)/1024/1024, 2)} MB`",
            parse_mode="Markdown", reply_markup=back_admin_kb()
        )
        return

    if data == "adm_toggle_maintenance":
        global MAINTENANCE_MODE
        MAINTENANCE_MODE = not MAINTENANCE_MODE
        status = (
            "🔧 *Maintenance mode ENABLED.*\n\nAll non-admin users will see the maintenance message."
            if MAINTENANCE_MODE else
            "✅ *Maintenance mode DISABLED.*\n\nBot is open to all users again."
        )
        await q.edit_message_text(status, parse_mode="Markdown", reply_markup=back_admin_kb())
        return

    action_map = {
        "adm_add_credits":    ("add_credits",    "💰 *Add / Deduct Credits*\n\nSend: `USER_ID AMOUNT`\nExample: `123456789 500`\n\n_Use negative to deduct: `123456789 -50`_"),
        "adm_set_tier":       ("set_tier",        f"⬆️ *Set User Tier*\n\nSend: `USER_ID TIER`\n\nTiers: `free` | `basic` | `premium` | `vip`"),
        "adm_ban":            ("ban",             "🔒 *Ban User*\n\nSend the User ID to ban:"),
        "adm_unban":          ("unban",           "✅ *Unban User*\n\nSend the User ID to unban:"),
        "adm_freeze":         ("freeze",          "🧊 *Freeze User Temporarily*\n\nSend: `USER_ID HOURS`\nExample: `123456789 24` → freeze for 24 hours"),
        "adm_adduser_inline": ("adduser_inline",  "➕ *Add User*\n\nSend: `USER_ID TIER`\nExample: `123456789 basic`\n\nTiers: `free` | `basic` | `premium` | `vip`"),
        "adm_deluser":        ("deluser",         "🗑️ *Delete User*\n\nSend the User ID to permanently delete from DB:"),
    }
    if data in action_map:
        action, prompt = action_map[data]
        context.user_data["admin_action"] = action
        await q.edit_message_text(
            prompt, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_home")]])
        )
        return

    if data == "adm_delete_file":
        context.user_data["admin_action"] = "delete_file"
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id, original_name, records FROM uploaded_files ORDER BY uploaded_at DESC LIMIT 20"
                    )
                    files = cur.fetchall()
        except Exception:
            files = []

        if not files:
            await q.edit_message_text("No files to delete.", reply_markup=back_admin_kb())
            return
        text = "🗑️ *Delete File*\n\nSend the file ID:\n\n"
        for fid, fname, recs in files:
            text += f"`#{fid}` — `{fname}` ({(recs or 0):,} records)\n"
        await q.edit_message_text(
            text, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_home")]])
        )
        return

# ════════════════════════════════════════════
#          TEXT HANDLER + SEARCH LOGIC
# ════════════════════════════════════════════
async def text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    text = update.message.text.strip()
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(_executor, ensure_user, uid, update.effective_user.username or "", update.effective_user.first_name or "")

    if is_banned(uid) and not is_admin(uid):
        log_failed_attempt(uid, "banned_access")
        await update.message.reply_text(
            "🚫 *Your account has been banned.*\n\nPlease contact the admin if you have any questions.",
            parse_mode="Markdown"
        )
        return

    if is_admin(uid) and "admin_action" in context.user_data:
        action = context.user_data["admin_action"]
        # Handle ticket reply
        if action.startswith("reply_ticket_"):
            ticket_id = int(action.replace("reply_ticket_", ""))
            context.user_data.pop("admin_action")
            try:
                with pool_conn() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT user_id FROM support_tickets WHERE id=%s", (ticket_id,))
                        row = cur.fetchone()
                if row:
                    reply_ticket(ticket_id, text)
                    await context.bot.send_message(
                        chat_id=row[0],
                        text=f"💬 *Reply to your ticket #{ticket_id}:*\n\n{mesc(text)}",
                        parse_mode="Markdown"
                    )
                    await update.message.reply_text(f"✅ Reply sent to ticket #{ticket_id}.", reply_markup=back_admin_kb())
            except Exception as e:
                await update.message.reply_text(f"❌ Error: {e}", reply_markup=back_admin_kb())
            return
        await handle_admin_text(update, context, context.user_data.pop("admin_action"), text)
        return

    # ── 2FA verification ──────────────────────────────────
    if context.user_data.get("verify_2fa"):
        if verify_2fa_code(uid, text.strip()):
            context.user_data.pop("verify_2fa")
            await update.message.reply_text(
                "✅ *2FA Verified Successfully!*",
                parse_mode="Markdown",
                reply_markup=user_main_kb(uid)
            )
        else:
            await update.message.reply_text(
                "❌ *Invalid or expired code.* Try again or tap Cancel.",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="my_security")]])
            )
        return

    # ── User coupon input ──────────────────────────────────
    if context.user_data.get("user_action") == "apply_coupon":
        context.user_data.pop("user_action")
        is_ar2 = get_lang(uid) == "ar"
        res = validate_coupon(text.strip(), uid)
        if res["ok"]:
            cp = res["coupon"]
            await update.message.reply_text(
                f"✅ *{'كوبون صالح!' if is_ar2 else 'Coupon Valid!'}*\n\n"
                f"{res['msg']}\n\n"
                f"{'اذهب للاشتراك واستخدم الكود عند الدفع.' if is_ar2 else 'Go to subscribe and use the code at checkout.'}",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("💳 " + ("اشترك الآن" if is_ar2 else "Subscribe Now"), callback_data="user_subscribe")],
                    [InlineKeyboardButton("🔙 Back", callback_data="user_home")],
                ])
            )
            # Store for use at checkout
            context.user_data["active_coupon"] = cp["code"]
        else:
            await update.message.reply_text(
                res["msg"],
                parse_mode="Markdown",
                reply_markup=user_main_kb(uid)
            )
        return

    # ── Quick search shortcuts ─────────────────────────────
    shortcut_type, shortcut_kw = detect_shortcut(text)
    if shortcut_type and shortcut_kw:
        if not can_search(uid):
            await update.message.reply_text(
                s(uid, "no_searches"),
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("💳 Plans", callback_data="show_plans")]])
            )
            return
        if shortcut_type == "ip":
            if await handle_advanced_search(update, context, shortcut_kw):
                return
        await show_search_counter(update, context, shortcut_kw, shortcut_type)
        return

    # ── 2FA Verification ──────────────────────────────────
    if context.user_data.get("waiting_2fa"):
        purpose = context.user_data.pop("2fa_verify_purpose", "action")
        context.user_data.pop("waiting_2fa")
        if verify_2fa_code(uid, text.strip()):
            await update.message.reply_text(
                f"✅ *2FA Verified!*\nCode accepted for: _{purpose}_",
                parse_mode="Markdown",
                reply_markup=user_main_kb(uid)
            )
            log_activity(uid, "2fa_verified", purpose)
        else:
            await update.message.reply_text(
                "❌ *Invalid or expired code.*\n\nRequest a new one from Security Settings.",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔐 Security", callback_data="my_security")]])
            )
            log_failed_attempt(uid, "2fa_failed")
        return

    # ── Payment: TX ID submission ──────────────────────────
    if context.user_data.get("pending_order_id"):
        order_id = context.user_data["pending_order_id"]
        order    = get_order(order_id)
        # If looks like a TX hash (crypto) — hex string ≥ 20 chars
        if order and order.get("status") == "pending" and re.fullmatch(r"[0-9a-fA-FxX\-]{20,100}", text.strip()):
            update_order_txid(order_id, text.strip())
            await update.message.reply_text(
                f"🔗 *TX ID Saved!*\n\n"
                f"📋 Order `#{order_id}`\n"
                f"🔗 TX: `{mesc(text.strip())}`\n\n"
                f"✅ Admin will verify and confirm shortly.",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📋 View Order", callback_data=f"view_order_{order_id}")]])
            )
            # Forward to admin
            order2 = get_order(order_id)
            u2     = get_user_cached(uid)
            notif  = (
                f"🔗 *TX ID Submitted*\n"
                f"Order `#{order_id}` | @{u2[1] if u2 else uid}\n"
                f"TX: `{mesc(text.strip())}`\n"
                f"Plan: {PLAN_PRICES.get(order2.get('plan',''),{}).get('label','?')}"
            )
            adm_kb = InlineKeyboardMarkup([[
                InlineKeyboardButton(f"✅ Confirm", callback_data=f"adm_confirm_order_{order_id}"),
                InlineKeyboardButton(f"❌ Reject",  callback_data=f"adm_reject_order_{order_id}"),
            ]])
            for admin_id in ADMIN_IDS:
                try:
                    await context.bot.send_message(chat_id=admin_id, text=notif, parse_mode="Markdown", reply_markup=adm_kb)
                except Exception:
                    pass
            return

    # ── Phase 1: Support ticket flow ──────────────────────
    if "support_action" in context.user_data:
        action = context.user_data["support_action"]
        if action == "new_ticket_subject":
            context.user_data["support_action"] = "new_ticket_message"
            context.user_data["ticket_subject"]  = text
            await update.message.reply_text(
                "📝 Now send your message / problem details:",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="my_support")]])
            )
            return
        elif action == "new_ticket_message":
            subject = context.user_data.pop("ticket_subject", "No subject")
            context.user_data.pop("support_action", None)
            u = get_user_cached(uid)
            ticket_id = create_ticket(
                uid,
                update.effective_user.username or "",
                u[2] if u else "User",
                subject, text
            )
            await update.message.reply_text(
                f"✅ *Ticket #{ticket_id} Created!*\n\n📝 Subject: {subject}\n\nWe'll reply soon.",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Menu", callback_data="user_home")]])
            )
            await notify_admins_new_ticket(context.application, ticket_id, uid, update.effective_user.username or str(uid), subject)
            return

    # ── Phase 1: Bulk search flow ──────────────────────────
    if context.user_data.get("bulk_search_waiting"):
        context.user_data.pop("bulk_search_waiting")
        keywords = [k.strip() for k in text.splitlines() if k.strip()]
        if not keywords:
            await update.message.reply_text("❌ No keywords found.", reply_markup=user_main_kb(uid))
            return
        stype = context.user_data.get("search_type", "email")
        context.user_data.pop("search_type", None)
        u     = get_user_cached(uid)
        tier  = u[3] if u else "free"
        limit = min(TIERS[tier]["max_results"], 50) if not is_admin(uid) else 200

        msg = await update.message.reply_text(
            f"⏳ *Bulk Search...*\n\n🔢 Keywords: {len(keywords)}\nPlease wait...",
            parse_mode="Markdown"
        )
        results = await asyncio.get_running_loop().run_in_executor(
            _executor, lambda: do_bulk_search(keywords, stype, limit)
        )
        content  = build_bulk_result_txt(results, stype, uid)
        filename = f"bulk_{stype}_{len(keywords)}kw_{int(time.time())}.txt"
        tmppath  = os.path.join(FILES_DIR, f"tmp_bulk_{uid}.txt")
        with open(tmppath, "w", encoding="utf-8") as f:
            f.write(content)

        total = sum(len(v) for v in results.values())
        caption = (
            f"✅ *Bulk Search Complete*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🔢 Keywords: `{len(results)}`\n"
            f"📊 Total: `{total:,}` records"
        )
        await msg.delete()
        await safe_send_document(
            update.message.reply_document, tmppath, filename, caption,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Menu", callback_data="user_home")]])
        )
        if not is_admin(uid):
            deduct(uid)
            add_points(uid, POINTS_PER_SEARCH * len(keywords), "bulk_search")
        return

    if "search_type" in context.user_data:
        stype = context.user_data.pop("search_type")
        if stype in ("ni_name", "ni_national_id"):
            await show_nameid_counter(update, context, text, stype)
            return
        if stype == "regex":
            await handle_regex_search(update, context, text)
            return
        # Check for advanced search first
        if await handle_advanced_search(update, context, text):
            return
        await show_search_counter(update, context, text, stype)
        return

    # Check if it's an advanced search query (IP/IBAN/Address)
    if detect_advanced_search_type(text):
        if await handle_advanced_search(update, context, text):
            return

    if is_admin(uid):
        await update.message.reply_text("⚙️ Admin Panel:", reply_markup=admin_main_kb())
    else:
        await update.message.reply_text("Use the menu:", reply_markup=user_main_kb(uid))

# ════════════════════════════════════════════
#    COUNTER PREVIEW BEFORE DOWNLOAD
# ════════════════════════════════════════════
async def show_search_counter(update: Update, context: ContextTypes.DEFAULT_TYPE, keyword: str, stype: str):
    uid = update.effective_user.id

    if len(keyword.strip()) < MIN_KEYWORD_LEN:
        await update.message.reply_text(
            f"❌ *Keyword too short!*\n\nMinimum *{MIN_KEYWORD_LEN} characters* required.\nTry a more specific keyword.",
            parse_mode="Markdown", reply_markup=back_user_kb(uid)
        )
        return

    if is_search_spamming(uid):
        await update.message.reply_text("⏳ Please wait a few seconds before searching again.", reply_markup=back_user_kb(uid))
        return
    if not can_search(uid):
        await update.message.reply_text(
            "❌ No searches remaining. Upgrade your plan.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("💳 View Plans", callback_data="show_plans")]])
        )
        return

    u     = get_user_cached(uid)
    tier  = u[3] if u else "free"
    limit = TIERS[tier]["max_results"] if not is_admin(uid) else 1_000_000
    mark_search_time(uid)

    msg = await update.message.reply_text(
        f"🔎 *Checking database...*\n\n`{mesc(keyword)}` [{stype.upper()}]",
        parse_mode="Markdown"
    )

    raw_count = await asyncio.get_running_loop().run_in_executor(
        _executor, lambda: count_matches_fast(keyword, stype)
    )
    capped = min(raw_count, limit)

    icons = {"url":"🌐","domain":"🌍","login":"👤","username":"📝",
             "email":"📧","phone":"📱","password":"🔑","all":"🔎"}
    icon = icons.get(stype, "🔍")

    if raw_count == 0:
        alt_types = [t for t in ["email","url","domain","login","username","phone","password","all"] if t != stype]
        suggestions = " | ".join(f"`{t}`" for t in alt_types[:4])
        await msg.edit_text(
            f"🔍 *No results found*\n\n"
            f"Target: `{mesc(keyword)}` | Type: `{stype.upper()}`\n\n"
            f"💡 *Suggestions:*\n"
            f"• Try a shorter keyword\n"
            f"• Try a different search type: {suggestions}\n"
            f"• Use `Full Scan` to search all types at once",
            parse_mode="Markdown", reply_markup=new_search_kb()
        )
        return

    safe_kw    = keyword[:40]
    cb_confirm = _cb_put(f"confirm_search:{stype}:{safe_kw}")
    cb_cancel  = "cancel_search"

    preview_results = await asyncio.get_running_loop().run_in_executor(
        _executor, lambda: smart_search(keyword, stype, 3)
    )
    preview_lines = ""
    if preview_results:
        preview_lines = "\n\n👁️ *Preview (first 3):*\n"
        for r in preview_results[:3]:
            em  = r.get("email") or r.get("username", "")
            pwd = r.get("password", "")
            url = r.get("url") or r.get("domain", "")
            ph  = r.get("phone", "")
            if em and pwd:
                preview_lines += f"• `{mesc(em[:30])}:{mesc(pwd[:20])}`\n"
            elif url and em:
                preview_lines += f"• `{mesc(url[:25])}` | `{mesc(em[:20])}`\n"
            elif ph:
                preview_lines += f"• `{mesc(ph)}`\n"
            elif em:
                preview_lines += f"• `{mesc(em[:40])}`\n"

    await msg.edit_text(
        f"{icon} *Search Preview*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🎯 Target    : `{mesc(keyword)}`\n"
        f"📂 Type      : `{stype.upper()}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📊 Available : `{raw_count:,}` raw matches\n"
        f"✅ Your limit: `{capped:,}` records\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"Tap *Download* to start the scan and get your `.txt` file.\n"
        f"⏱️ May take up to 3 minutes.{preview_lines}",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(f"📥 Download ({capped:,} records)", callback_data=cb_confirm)],
            [InlineKeyboardButton("❌ Cancel", callback_data=cb_cancel)],
        ])
    )

async def show_nameid_counter(update: Update, context: ContextTypes.DEFAULT_TYPE, query: str, stype: str):
    uid = update.effective_user.id

    if stype == "ni_national_id":
        cleaned = re.sub(r"\s", "", query)
        if not re.fullmatch(r"\d{4,14}", cleaned):
            await update.message.reply_text(
                "❌ *رقم غير صالح!*\n\nأدخل رقماً يتكون من 4 إلى 14 رقماً فقط.\n_مثال: `30604` أو `30604150100123`_",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 رجوع", callback_data="go_nameid")]])
            )
            return
    elif stype == "ni_name":
        if len(query.strip()) < 2:
            await update.message.reply_text(
                "❌ *الاسم قصير جداً!*\n\nأدخل اسماً من حرفين على الأقل.",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 رجوع", callback_data="go_nameid")]])
            )
            return
        if re.fullmatch(r"[\d\s]+", query.strip()):
            await update.message.reply_text(
                "❌ *الاسم يجب أن يحتوي على حروف!*\n\nلبحث بالرقم اختر 🪪 بحث بالرقم القومي.",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 رجوع", callback_data="go_nameid")]])
            )
            return

    if is_search_spamming(uid):
        await update.message.reply_text("⏳ Please wait a few seconds before searching again.", reply_markup=back_user_kb(uid))
        return
    if not can_search_nameid(uid):
        is_ar = get_lang(uid) == "ar"
        msg_out = (
            "❌ *انتهت بحوث Name/ID لليوم.*\n\nقم بترقية باقتك للحصول على المزيد."
            if is_ar else
            "❌ *Name/ID searches used up for today.*\n\nUpgrade your plan for more."
        )
        await update.message.reply_text(
            msg_out, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("💳 Plans", callback_data="show_plans")],
                [InlineKeyboardButton("🔙 Back",  callback_data="user_home")],
            ])
        )
        return

    limit = get_nameid_limit(uid)
    mark_search_time(uid)

    if stype == "ni_national_id":
        qtype = "national_id" if is_national_id(query) else "partial_id"
    else:
        qtype = "name"

    type_labels = {
        "national_id": "🪪 رقم قومي",
        "partial_id":  "🔢 رقم جزئي",
        "name":        "👤 اسم",
    }

    msg = await update.message.reply_text(
        f"🔎 *جاري الفحص...*\n\n`{mesc(query)}`", parse_mode="Markdown"
    )

    raw_count = await asyncio.get_running_loop().run_in_executor(
        _executor, lambda: count_nameid_matches(query, qtype)
    )
    capped = min(raw_count, limit)

    if raw_count == 0:
        await msg.edit_text(
            f"🔍 *لا توجد نتائج*\n\nQuery: `{mesc(query)}`\n\n"
            f"💡 *اقتراحات:*\n• جرب اسماً أقصر أو جزءاً من الرقم\n• تأكد من الإملاء الصحيح",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🪪 بحث جديد", callback_data="go_nameid")],
                [InlineKeyboardButton("🏠 Main Menu", callback_data="user_home")],
            ])
        )
        return

    preview_results = await asyncio.get_running_loop().run_in_executor(
        _executor, lambda: (search_by_name(query, 3) if qtype == "name" else search_by_national_id(query, 3))
    )
    preview_lines = ""
    if preview_results:
        preview_lines = "\n\n👁️ *معاينة (أول 3):*\n"
        for r in preview_results[:3]:
            preview_lines += f"• 👤 `{mesc(r.get('name','')[:25])}` | 🪪 `{r.get('national_id','')}`\n"

    safe_q     = query[:40]
    cb_confirm = _cb_put(f"confirm_nameid:{stype}:{safe_q}")
    cb_cancel  = "cancel_search"

    await msg.edit_text(
        f"🪪 *نتائج البحث — معاينة*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🎯 Query  : `{mesc(query)}`\n"
        f"📂 Type   : {type_labels[qtype]}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📊 متاح   : `{raw_count:,}` سجل\n"
        f"✅ حدك    : `{capped:,}` سجل"
        f"{preview_lines}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"اضغط *تحميل* للحصول على الملف.",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(f"📥 تحميل ({capped:,} سجل)", callback_data=cb_confirm)],
            [InlineKeyboardButton("❌ إلغاء", callback_data=cb_cancel)],
        ])
    )

# ════════════════════════════════════════════
#       NAME / NATIONAL ID SEARCH HANDLER
# ════════════════════════════════════════════
async def do_nameid_search(update: Update, context: ContextTypes.DEFAULT_TYPE,
                           query: str, stype: str, reply_to=None):
    uid = update.effective_user.id
    if context.user_data.get("search_running"):
        send_fn = reply_to.reply_text if reply_to else update.message.reply_text
        await send_fn("⏳ A search is already in progress. Please wait.")
        return
    context.user_data["search_running"] = True
    try:
        if not can_search_nameid(uid):
            send_fn = reply_to.reply_text if reply_to else update.message.reply_text
            await send_fn(
                "❌ *Name/ID searches used up for today.*\n\nUpgrade your plan for more.",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("💳 Plans", callback_data="show_plans")]])
            )
            return

        limit = get_nameid_limit(uid)

        if stype == "ni_national_id":
            qtype = "national_id" if is_national_id(query) else "partial_id"
        else:
            qtype = "name"

        type_labels = {
            "national_id": "🪪 رقم قومي (14 رقم)",
            "partial_id":  "🔢 رقم جزئي",
            "name":        "👤 اسم",
        }

        send_fn = reply_to.reply_text if reply_to else update.message.reply_text
        msg = await send_fn(
            f"⏳ *جاري البحث...*\n\n"
            f"🎯 Query: `{mesc(query)}`\n"
            f"📂 Type: {type_labels[qtype]}",
            parse_mode="Markdown"
        )

        if qtype == "name":
            results = await asyncio.get_running_loop().run_in_executor(
                _executor, lambda: search_by_name(query, limit)
            )
        else:
            results = await asyncio.get_running_loop().run_in_executor(
                _executor, lambda: search_by_national_id(query, limit)
            )

        if not is_admin(uid):
            deduct_nameid(uid)
        log_search(uid, query, f"nameid_{qtype}", len(results))

        if not results:
            await msg.edit_text(
                f"🔍 *لا توجد نتائج*\n\nQuery: `{mesc(query)}`",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🪪 بحث جديد", callback_data="go_nameid")],
                    [InlineKeyboardButton("🏠 Main Menu", callback_data="user_home")],
                ])
            )
            return

        if len(results) <= 3:
            lines = [f"✅ *نتائج البحث* — `{len(results)}` نتيجة\n"]
            for r in results:
                lines.append(f"👤 *{mesc(r['name'])}*\n🪪 `{r['national_id']}`\n")
            await msg.edit_text(
                "\n".join(lines), parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🪪 بحث جديد", callback_data="go_nameid")],
                    [InlineKeyboardButton("🏠 Main Menu", callback_data="user_home")],
                ])
            )
            return

        content  = build_nameid_result_txt(query, results, qtype)
        safe_kw  = re.sub(r"[^\w\-]", "_", query)[:30]
        filename = f"nameid_{safe_kw}_{len(results)}_results.txt"
        tmppath  = os.path.join(FILES_DIR, f"tmp_nameid_{uid}.txt")
        with open(tmppath, "w", encoding="utf-8") as f:
            f.write(content)

        caption = (
            f"✅ *نتائج Name/ID Search*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🎯 Query  : `{mesc(query)}`\n"
            f"📂 Type   : {type_labels[qtype]}\n"
            f"📊 Total  : `{len(results):,}` سجل\n"
            f"📄 File   : `{mesc(filename)}`"
        )

        await msg.delete()
        send_doc = reply_to.reply_document if reply_to else update.message.reply_document

        file_size_kb = os.path.getsize(tmppath) / 1024
        if file_size_kb > 200:
            zippath = tmppath + ".zip"
            with zipfile.ZipFile(zippath, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.write(tmppath, filename)
            send_path, send_filename = zippath, filename + ".zip"
            caption += f"\n📦 Compressed"
        else:
            send_path, send_filename = tmppath, filename

        await safe_send_document(
            send_doc, send_path, send_filename, caption,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🪪 بحث جديد", callback_data="go_nameid")],
                [InlineKeyboardButton("🏠 Main Menu", callback_data="user_home")],
            ])
        )
    finally:
        context.user_data["search_running"] = False

# ════════════════════════════════════════════
#         REGULAR SEARCH + TIMER
# ════════════════════════════════════════════
async def do_search(update: Update, context: ContextTypes.DEFAULT_TYPE,
                    keyword: str, stype: str, reply_to=None):
    uid = update.effective_user.id
    if context.user_data.get("search_running"):
        await update.message.reply_text("⏳ A search is already in progress. Please wait.")
        return
    context.user_data["search_running"] = True
    try:
        if not can_search(uid):
            await update.message.reply_text(
                "❌ No searches remaining. Upgrade your plan.",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("💳 View Plans", callback_data="show_plans")]])
            )
            return

        u     = get_user_cached(uid)
        tier  = u[3] if u else "free"
        limit = TIERS[tier]["max_results"] if not is_admin(uid) else 1_000_000

        send_fn = reply_to.reply_text if reply_to else update.message.reply_text
        msg = await send_fn(
            f"⏳ *Scanning database...*\n\n"
            f"🎯 Target: `{mesc(keyword)}`\n"
            f"📂 Type: `{stype.upper()}`\n\n"
            f"⏱️ Time remaining: *3:00*\n"
            f"`░░░░░░░░░░░░░░░░░░░░` 0%",
            parse_mode="Markdown"
        )

        results = await run_search_with_timer(msg, keyword, stype, limit)

        if not is_admin(uid):
            deduct(uid)
            add_points(uid, POINTS_PER_SEARCH, "search")
            # Check daily challenge
            ch = get_daily_challenge_progress(uid)
            if ch["done"] and ch["count"] == 5:  # Exactly hit 5 — award bonus once
                add_points(uid, ch["reward"], "daily_challenge")
            # Check badges in background
            asyncio.create_task(
                asyncio.get_event_loop().run_in_executor(_executor, check_and_award_badges, uid)
            )
        log_search(uid, keyword, stype, len(results))

        if not results:
            alt_types = [t for t in ["email","url","domain","login","username","phone","password","all"] if t != stype]
            suggestions = " | ".join(f"`{t}`" for t in alt_types[:4])
            await msg.edit_text(
                f"🔍 *No results found*\n\n"
                f"Target: `{mesc(keyword)}` | Type: `{stype.upper()}`\n\n"
                f"💡 *Suggestions:*\n"
                f"• Try a shorter keyword\n"
                f"• Try a different type: {suggestions}\n"
                f"• Use `Full Scan` to search all types",
                parse_mode="Markdown", reply_markup=new_search_kb()
            )
            return

        content  = build_result_txt(keyword, results, stype)
        content  = add_watermark(content, uid)
        safe_kw  = re.sub(r"[^\w\-]", "_", keyword)[:30]
        filename = f"{safe_kw}_{len(results)}_results.txt"
        tmppath  = os.path.join(FILES_DIR, f"tmp_{uid}.txt")
        with open(tmppath, "w", encoding="utf-8") as f:
            f.write(content)

        email_count = sum(1 for r in results if r.get("email"))
        other_count = len(results) - email_count

        caption = (
            f"✅ *Scan Complete*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🎯 Target    : `{mesc(keyword)}`\n"
            f"📂 Type      : `{stype.upper()}`\n"
            f"📊 Total     : `{len(results):,}` records\n"
            f"📧 With Email: `{email_count:,}` | 👤 Username: `{other_count:,}`\n"
            f"📄 File      : `{filename}`"
        )

        await msg.delete()
        send_doc = reply_to.reply_document if reply_to else update.message.reply_document

        file_size_kb = os.path.getsize(tmppath) / 1024
        if file_size_kb > 200:
            zippath = tmppath + ".zip"
            with zipfile.ZipFile(zippath, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.write(tmppath, filename)
            send_path, send_filename = zippath, filename + ".zip"
            caption += f"\n📦 Compressed"
        else:
            send_path, send_filename = tmppath, filename

        # Offer Excel export button
        result_kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔍 New Search",      callback_data="go_search"),
             InlineKeyboardButton("📊 Export Excel",    callback_data=_cb_put(f"export_excel:{keyword}:{stype}"))],
            [InlineKeyboardButton("🔽 Filter by Category", callback_data=_cb_put(f"show_filter:{keyword}:{stype}"))],
            [InlineKeyboardButton("⭐ Save to Favorites",  callback_data=_cb_put(f"save_fav:{stype}:{keyword}"))],
            [InlineKeyboardButton("🏠 Main Menu",          callback_data="user_home")],
        ])

        await safe_send_document(send_doc, send_path, send_filename, caption, reply_markup=result_kb)
    finally:
        context.user_data["search_running"] = False

async def run_search_with_timer(msg, keyword: str, stype: str, limit: int) -> list:
    total_secs = SEARCH_TIMEOUT
    bar_chars  = 20

    async def update_timer():
        intervals = [15, 30, 45, 60, 75, 90, 105, 120, 135, 150, 165, 175]
        for elapsed in intervals:
            await asyncio.sleep(15)
            remaining = total_secs - elapsed
            if remaining <= 0:
                break
            pct  = int((elapsed / total_secs) * 100)
            fill = int(bar_chars * pct / 100)
            bar  = "█" * fill + "░" * (bar_chars - fill)
            mins = remaining // 60
            secs = remaining % 60
            try:
                await msg.edit_text(
                    f"⏳ *Scanning database...*\n\n"
                    f"🎯 Target: `{mesc(keyword)}`\n"
                    f"📂 Type: `{stype.upper()}`\n\n"
                    f"⏱️ Time remaining: *{mins}:{secs:02d}*\n"
                    f"`{bar}` {pct}%",
                    parse_mode="Markdown"
                )
            except Exception:
                pass

    timer_task = asyncio.create_task(update_timer())
    try:
        results = await asyncio.wait_for(
            asyncio.get_running_loop().run_in_executor(
                _executor, lambda: smart_search(keyword, stype, limit)
            ),
            timeout=SEARCH_TIMEOUT + 10
        )
    except asyncio.TimeoutError:
        results = []
    finally:
        timer_task.cancel()
    return results

# ════════════════════════════════════════════
#         ADMIN TEXT ACTION HANDLER
# ════════════════════════════════════════════
async def handle_admin_text(update, context, action, text):
    uid   = update.effective_user.id
    parts = text.strip().split()

    if action == "add_credits":
        if len(parts) != 2 or not parts[1].lstrip("-").isdigit():
            await update.message.reply_text(
                "❌ Format: `USER_ID AMOUNT`\n\n_Use negative to deduct, e.g. `123456 -50`_",
                parse_mode="Markdown", reply_markup=back_admin_kb()
            )
            return
        target, amount = int(parts[0]), int(parts[1])
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    if amount < 0:
                        cur.execute(
                            "UPDATE users SET credits=GREATEST(0, credits+%s) WHERE user_id=%s",
                            (amount, target)
                        )
                        cur.execute("SELECT credits FROM users WHERE user_id=%s", (target,))
                        row = cur.fetchone()
                        new_balance = row[0] if row else 0
                        result_msg = f"✅ Deducted `{abs(amount)}` credits from user `{target}`. New balance: `{new_balance}`."
                    else:
                        cur.execute("UPDATE users SET credits=credits+%s WHERE user_id=%s", (amount, target))
                        result_msg = f"✅ Added `{amount}` credits to user `{target}`."
                    cur.execute("INSERT INTO sub_history VALUES (%s,%s,%s,%s,%s)",
                                (target, "credits", amount, uid, datetime.utcnow().isoformat()))
                conn.commit()
        except Exception as e:
            log.error(f"add_credits error: {e}")
            result_msg = f"❌ Error: {e}"
        await update.message.reply_text(result_msg, parse_mode="Markdown", reply_markup=back_admin_kb())
        log_admin_op(uid, "add_credits", str(target), f"{amount:+} credits")

    elif action == "set_tier":
        if len(parts) != 2 or parts[1] not in TIERS:
            await update.message.reply_text(
                f"❌ Format: `USER_ID TIER`\nTiers: {', '.join(TIERS)}",
                parse_mode="Markdown", reply_markup=back_admin_kb()
            )
            return
        target, new_tier = int(parts[0]), parts[1]
        t  = TIERS[new_tier]
        nt = NAMEID_TIERS[new_tier]
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT 1 FROM users WHERE user_id=%s", (target,))
                    if cur.fetchone():
                        cur.execute(
                            "UPDATE users SET tier=%s, daily_limit=%s, daily_nameid_limit=%s WHERE user_id=%s",
                            (new_tier, t["daily"], nt["daily_nameid"], target)
                        )
                    else:
                        cur.execute(
                            "INSERT INTO users (user_id, username, full_name, tier, daily_limit, credits, "
                            "is_banned, expires_at, joined_at, lang, daily_nameid_limit) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                            (target, "unknown", "User", new_tier, t["daily"], 0, 0, None,
                             datetime.utcnow().isoformat(), "en", nt["daily_nameid"])
                        )
                    cur.execute("INSERT INTO sub_history VALUES (%s,%s,%s,%s,%s)",
                                (target, new_tier, 0, uid, datetime.utcnow().isoformat()))
                conn.commit()
        except Exception as e:
            log.error(f"set_tier error: {e}")
        await update.message.reply_text(f"✅ User `{target}` → *{new_tier}*.", parse_mode="Markdown", reply_markup=back_admin_kb())
        log_admin_op(uid, "set_tier", str(target), f"tier={new_tier}")

    elif action == "ban":
        if not parts or not parts[0].lstrip("-").isdigit():
            await update.message.reply_text("❌ Send a valid User ID.", reply_markup=back_admin_kb())
            return
        target = int(parts[0])
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("UPDATE users SET is_banned=1 WHERE user_id=%s", (target,))
                conn.commit()
        except Exception as e:
            log.error(f"ban error: {e}")
        await update.message.reply_text(f"🔒 User `{target}` banned.", parse_mode="Markdown", reply_markup=back_admin_kb())
        log_admin_op(uid, "ban", str(target))

    elif action == "unban":
        if not parts or not parts[0].lstrip("-").isdigit():
            await update.message.reply_text("❌ Send a valid User ID.", reply_markup=back_admin_kb())
            return
        target = int(parts[0])
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("UPDATE users SET is_banned=0 WHERE user_id=%s", (target,))
                conn.commit()
        except Exception as e:
            log.error(f"unban error: {e}")
        await update.message.reply_text(f"✅ User `{target}` unbanned.", parse_mode="Markdown", reply_markup=back_admin_kb())
        log_admin_op(uid, "unban", str(target))

    elif action == "delete_file":
        if not parts or not parts[0].lstrip("#").isdigit():
            await update.message.reply_text("❌ Send a valid file ID.", reply_markup=back_admin_kb())
            return
        fid = int(parts[0].lstrip("#"))
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT saved_name, original_name, records FROM uploaded_files WHERE id=%s", (fid,))
                    row = cur.fetchone()
                    if not row:
                        await update.message.reply_text(f"❌ File `#{fid}` not found.", parse_mode="Markdown", reply_markup=back_admin_kb())
                        return
                    saved_name, orig_name, record_count = row
                    cur.execute("DELETE FROM data_index WHERE source=%s", (orig_name,))
                    cur.execute("DELETE FROM name_id_index WHERE source=%s", (orig_name,))
                    cur.execute("DELETE FROM uploaded_files WHERE id=%s", (fid,))
                conn.commit()
        except Exception as e:
            log.error(f"delete_file error: {e}")
            await update.message.reply_text(f"❌ Error: {e}", reply_markup=back_admin_kb())
            return
        fpath = os.path.join(FILES_DIR, saved_name)
        if os.path.exists(fpath):
            os.remove(fpath)
        await update.message.reply_text(
            f"🗑️ `{orig_name}` deleted. `{(record_count or 0):,}` records removed.",
            parse_mode="Markdown", reply_markup=back_admin_kb()
        )
        log_admin_op(uid, "delete_file", orig_name, f"{record_count:,} records removed")

    elif action == "broadcast":
        msg_text = text.strip()
        if not msg_text:
            await update.message.reply_text("❌ Message is empty.", reply_markup=back_admin_kb())
            return
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT user_id FROM users WHERE is_banned=0")
                    user_ids = [r[0] for r in cur.fetchall()]
        except Exception:
            user_ids = []

        progress = await update.message.reply_text(
            f"📢 Broadcasting to *{len(user_ids)}* users...", parse_mode="Markdown"
        )
        sent = failed = flood_waits = 0
        for i, target_uid in enumerate(user_ids):
            try:
                await context.bot.send_message(
                    chat_id=target_uid,
                    text=f"📢 *Message from Admin:*\n\n{mesc(msg_text)}",
                    parse_mode="Markdown"
                )
                sent += 1
            except RetryAfter as e:
                flood_waits += 1
                wait_secs = int(e.retry_after) + 1
                await asyncio.sleep(wait_secs)
                try:
                    await context.bot.send_message(chat_id=target_uid, text=f"📢 Message from Admin:\n\n{msg_text}")
                    sent += 1
                except Exception:
                    failed += 1
            except (Forbidden, BadRequest):
                failed += 1
            except Exception:
                failed += 1
            if (i + 1) % 50 == 0:
                try:
                    await progress.edit_text(
                        f"📢 *Broadcasting...*\n\n"
                        f"📤 Sent: `{sent}` / `{len(user_ids)}`\n"
                        f"❌ Failed: `{failed}`",
                        parse_mode="Markdown"
                    )
                except Exception:
                    pass
            await asyncio.sleep(0.05)

        await progress.edit_text(
            f"✅ *Broadcast Complete*\n\n"
            f"📤 Sent: `{sent}`\n"
            f"❌ Failed: `{failed}`\n"
            f"⏸️ FloodWaits: `{flood_waits}`\n"
            f"👥 Total: `{len(user_ids)}`",
            parse_mode="Markdown", reply_markup=back_admin_kb()
        )
        log_admin_op(uid, "broadcast", "all_users", f"sent={sent}, failed={failed}")

    elif action == "freeze":
        if len(parts) != 2 or not parts[0].lstrip("-").isdigit() or not parts[1].isdigit():
            await update.message.reply_text("❌ Format: `USER_ID HOURS`", parse_mode="Markdown", reply_markup=back_admin_kb())
            return
        target, hours = int(parts[0]), int(parts[1])
        if hours <= 0:
            try:
                with pool_conn() as conn:
                    with conn.cursor() as cur:
                        cur.execute("UPDATE users SET frozen_until=NULL WHERE user_id=%s", (target,))
                    conn.commit()
            except Exception as e:
                log.error(f"unfreeze error: {e}")
            await update.message.reply_text(f"✅ User `{target}` unfrozen.", parse_mode="Markdown", reply_markup=back_admin_kb())
            log_admin_op(uid, "unfreeze", str(target))
        else:
            until = (datetime.utcnow() + timedelta(hours=hours)).isoformat()
            try:
                with pool_conn() as conn:
                    with conn.cursor() as cur:
                        cur.execute("UPDATE users SET frozen_until=%s WHERE user_id=%s", (until, target))
                    conn.commit()
            except Exception as e:
                log.error(f"freeze error: {e}")
            await update.message.reply_text(
                f"🧊 User `{target}` frozen for `{hours}` hours.",
                parse_mode="Markdown", reply_markup=back_admin_kb()
            )
            log_admin_op(uid, "freeze", str(target), f"{hours}h")
            try:
                await context.bot.send_message(
                    chat_id=target,
                    text=f"🧊 Your account has been temporarily frozen for *{hours}* hour(s).",
                    parse_mode="Markdown"
                )
            except Exception:
                pass

    elif action == "msg_user":
        await handle_msg_user(update, context, text)

    elif action == "adduser_inline":
        if len(parts) != 2 or not parts[0].lstrip("-").isdigit() or parts[1] not in TIERS:
            await update.message.reply_text(
                f"❌ Format: `USER_ID TIER`\nTiers: {', '.join(TIERS)}",
                parse_mode="Markdown", reply_markup=back_admin_kb()
            )
            return
        target, tier_val = int(parts[0]), parts[1]
        t  = TIERS[tier_val]
        nt = NAMEID_TIERS[tier_val]
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT 1 FROM users WHERE user_id=%s", (target,))
                    if cur.fetchone():
                        cur.execute(
                            "UPDATE users SET tier=%s, daily_limit=%s, daily_nameid_limit=%s WHERE user_id=%s",
                            (tier_val, t["daily"], nt["daily_nameid"], target)
                        )
                        result_msg = f"✅ Updated user `{target}` → tier `{tier_val}`."
                    else:
                        cur.execute(
                            "INSERT INTO users (user_id, username, full_name, tier, daily_limit, credits, "
                            "is_banned, expires_at, joined_at, lang, daily_nameid_limit) "
                            "VALUES (%s,%s,%s,%s,%s,0,0,NULL,%s,%s,%s)",
                            (target, "", "", tier_val, t["daily"], datetime.utcnow().isoformat(), "en", nt["daily_nameid"])
                        )
                        result_msg = f"✅ User `{target}` added with tier `{tier_val}`."
                conn.commit()
        except Exception as e:
            log.error(f"adduser_inline error: {e}")
            result_msg = f"❌ Error: {e}"
        await update.message.reply_text(result_msg, parse_mode="Markdown", reply_markup=back_admin_kb())
        log_admin_op(uid, "adduser_inline", str(target), tier_val)

    elif action == "deluser":
        if len(parts) != 1 or not parts[0].lstrip("-").isdigit():
            await update.message.reply_text("❌ Send a valid User ID.", parse_mode="Markdown", reply_markup=back_admin_kb())
            return
        target = int(parts[0])
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT full_name, username FROM users WHERE user_id=%s", (target,))
                    exists = cur.fetchone()
                    if not exists:
                        await update.message.reply_text(f"❌ User `{target}` not found.", parse_mode="Markdown", reply_markup=back_admin_kb())
                        return
                    cur.execute("DELETE FROM users WHERE user_id=%s", (target,))
                    cur.execute("DELETE FROM search_logs WHERE user_id=%s", (target,))
                    cur.execute("DELETE FROM sub_history WHERE user_id=%s", (target,))
                    cur.execute("DELETE FROM sub_requests WHERE user_id=%s", (target,))
                conn.commit()
        except Exception as e:
            log.error(f"deluser error: {e}")
            await update.message.reply_text(f"❌ Error: {e}", reply_markup=back_admin_kb())
            return
        fname_del = esc(exists[0] or "N/A")
        await update.message.reply_text(
            f"✅ User `{target}` ({fname_del}) permanently deleted.",
            parse_mode="Markdown", reply_markup=back_admin_kb()
        )
        log_admin_op(uid, "deluser", str(target), f"name={exists[0]}")

    elif action == "set_expiry":
        if len(parts) != 2 or not parts[0].lstrip("-").isdigit() or not parts[1].lstrip("-").isdigit():
            await update.message.reply_text("❌ Format: `USER_ID DAYS`", parse_mode="Markdown", reply_markup=back_admin_kb())
            return
        target, days = int(parts[0]), int(parts[1])
        if days <= 0:
            exp_val  = None
            msg_out  = f"✅ Expiry cleared for user `{target}`."
        else:
            exp_val  = (datetime.utcnow() + timedelta(days=days)).isoformat()
            msg_out  = f"✅ User `{target}` expires in `{days}` days (`{exp_val[:10]}`)."
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("UPDATE users SET expires_at=%s WHERE user_id=%s", (exp_val, target))
                conn.commit()
        except Exception as e:
            log.error(f"set_expiry error: {e}")
        await update.message.reply_text(msg_out, parse_mode="Markdown", reply_markup=back_admin_kb())
        log_admin_op(uid, "set_expiry", str(target), f"days={days}")

    elif action == "filter_logs":
        query_val = text.strip()
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    if query_val.lstrip("-").isdigit():
                        cur.execute(
                            "SELECT user_id, keyword, category, results, timestamp FROM search_logs "
                            "WHERE user_id=%s ORDER BY timestamp DESC LIMIT 30",
                            (int(query_val),)
                        )
                        title = f"🔎 Logs for user `{mesc(query_val)}`"
                    else:
                        cur.execute(
                            "SELECT user_id, keyword, category, results, timestamp FROM search_logs "
                            "WHERE keyword ILIKE %s ORDER BY timestamp DESC LIMIT 30",
                            (f"%{query_val}%",)
                        )
                        title = f"🔎 Logs matching `{esc(query_val)}`"
                    rows = cur.fetchall()
        except Exception as e:
            log.error(f"filter_logs error: {e}")
            rows = []
            title = "❌ Error"

        if not rows:
            await update.message.reply_text(f"📭 No logs found for `{esc(query_val)}`.", parse_mode="Markdown", reply_markup=back_admin_kb())
            return
        text_out = f"{title}\n━━━━━━━━━━━━━━━━━━━━━━\n\n"
        for r in rows:
            text_out += f"`{r[0]}` → `{esc(r[1][:20])}` [{esc(r[2])}] {r[3]} res @ {str(r[4])[:10]}\n"
        await update.message.reply_text(text_out[:4000], parse_mode="Markdown", reply_markup=back_admin_kb())

    elif action == "sub_request_approve":
        if not parts or not parts[0].isdigit():
            await update.message.reply_text("❌ Send a valid request ID.", reply_markup=back_admin_kb())
            return
        req_id        = int(parts[0])
        override_tier = parts[1] if len(parts) > 1 and parts[1] in TIERS else None
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT user_id, username, full_name, requested_tier FROM sub_requests WHERE id=%s", (req_id,))
                    req = cur.fetchone()
                    if not req:
                        await update.message.reply_text(f"❌ Request `#{req_id}` not found.", parse_mode="Markdown", reply_markup=back_admin_kb())
                        return
                    target_uid, uname, fname, req_tier = req
                    final_tier = override_tier or req_tier
                    t  = TIERS[final_tier]
                    nt = NAMEID_TIERS[final_tier]
                    cur.execute("SELECT 1 FROM users WHERE user_id=%s", (target_uid,))
                    if cur.fetchone():
                        cur.execute(
                            "UPDATE users SET tier=%s, daily_limit=%s, daily_nameid_limit=%s WHERE user_id=%s",
                            (final_tier, t["daily"], nt["daily_nameid"], target_uid)
                        )
                    else:
                        cur.execute(
                            "INSERT INTO users (user_id, username, full_name, tier, daily_limit, credits, "
                            "is_banned, expires_at, joined_at, lang, daily_nameid_limit) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                            (target_uid, uname, fname, final_tier, t["daily"], 0, 0, None,
                             datetime.utcnow().isoformat(), "en", nt["daily_nameid"])
                        )
                    cur.execute("UPDATE sub_requests SET status='approved' WHERE id=%s", (req_id,))
                    cur.execute("INSERT INTO sub_history VALUES (%s,%s,%s,%s,%s)",
                                (target_uid, final_tier, 0, uid, datetime.utcnow().isoformat()))
                conn.commit()
        except Exception as e:
            log.error(f"sub_request_approve error: {e}")
            await update.message.reply_text(f"❌ Error: {e}", reply_markup=back_admin_kb())
            return

        try:
            user_lang = get_lang(target_uid)
            user_st   = STRINGS.get(user_lang, STRINGS["en"])
            await context.bot.send_message(
                chat_id=target_uid,
                text=user_st["sub_approved_user"].format(tier=final_tier),
                parse_mode="Markdown"
            )
        except Exception:
            pass
        await update.message.reply_text(
            f"✅ Request `#{req_id}` approved → `{final_tier}` for user `{target_uid}`.",
            parse_mode="Markdown", reply_markup=back_admin_kb()
        )
        log_admin_op(uid, "approve_sub", str(target_uid), f"req={req_id}, tier={final_tier}")

    elif action == "sub_request_reject":
        if not parts or not parts[0].isdigit():
            await update.message.reply_text("❌ Send a valid request ID.", reply_markup=back_admin_kb())
            return
        req_id = int(parts[0])
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT user_id FROM sub_requests WHERE id=%s", (req_id,))
                    req = cur.fetchone()
                    if not req:
                        await update.message.reply_text(f"❌ Request `#{req_id}` not found.", parse_mode="Markdown", reply_markup=back_admin_kb())
                        return
                    target_uid = req[0]
                    cur.execute("UPDATE sub_requests SET status='rejected' WHERE id=%s", (req_id,))
                conn.commit()
        except Exception as e:
            log.error(f"sub_request_reject error: {e}")
            await update.message.reply_text(f"❌ Error: {e}", reply_markup=back_admin_kb())
            return

        try:
            user_lang = get_lang(target_uid)
            user_st   = STRINGS.get(user_lang, STRINGS["en"])
            await context.bot.send_message(
                chat_id=target_uid,
                text=user_st["sub_rejected_user"],
                parse_mode="Markdown"
            )
        except Exception:
            pass
        await update.message.reply_text(
            f"❌ Request `#{req_id}` rejected.",
            parse_mode="Markdown", reply_markup=back_admin_kb()
        )
        log_admin_op(uid, "reject_sub", str(target_uid), f"req={req_id}")

    elif action.startswith("reject_order_"):
        order_id = int(action.replace("reject_order_", ""))
        reason   = text.strip() if text.strip() != "-" else ""
        res = reject_order(order_id, uid, reason)
        if res["ok"]:
            await update.message.reply_text(
                f"❌ *Order #{order_id} Rejected.*",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Orders", callback_data="adm_payments")]])
            )
            try:
                await context.bot.send_message(
                    chat_id=res["user_id"],
                    text=(
                        f"❌ *Payment Rejected.*\n\n"
                        f"Order `#{order_id}` was not confirmed.\n"
                        f"{('Reason: ' + reason) if reason else ''}\n\n"
                        f"Contact support if you have questions."
                    ),
                    parse_mode="Markdown"
                )
            except Exception:
                pass
        else:
            await update.message.reply_text(f"❌ Error: {res['msg']}", reply_markup=back_admin_kb())

    elif action == "add_scheduled_import":
        parts2 = text.strip().split(None, 2)
        if len(parts2) < 1 or not parts2[0].startswith("http"):
            await update.message.reply_text(
                "❌ Format: `URL LABEL HOURS`\nExample: `https://site.com/data.txt MyLabel 24`",
                parse_mode="Markdown", reply_markup=back_admin_kb()
            )
            return
        url    = parts2[0]
        label  = parts2[1] if len(parts2) > 1 else url[:20]
        hours  = int(parts2[2]) if len(parts2) > 2 and parts2[2].isdigit() else 24
        sid = create_scheduled_import(url, label, "auto", hours, uid)
        await update.message.reply_text(
            f"✅ *Scheduled Import Created!*\n\n"
            f"🔢 ID: `#{sid}`\n🔗 URL: `{url}`\n⏰ Every `{hours}` hours",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📋 View All", callback_data="adm_scheduled")]])
        )

    elif action == "url_import":
        parts3 = text.strip().split(None, 1)
        url    = parts3[0]
        label  = parts3[1] if len(parts3) > 1 else "url_import"
        if not url.startswith("http"):
            await update.message.reply_text("❌ Invalid URL.", reply_markup=back_admin_kb())
            return
        await update.message.reply_text(f"⏳ Importing from URL...\n`{mesc(url)}`", parse_mode="Markdown")
        result = await import_from_url(url, label, uid, context.bot)
        await update.message.reply_text(
            f"{'✅' if result['ok'] else '❌'} *URL Import*\n\n{result['msg']}",
            parse_mode="Markdown", reply_markup=back_admin_kb()
        )

    elif action == "view_user_log":
        if not parts or not parts[0].lstrip("-").isdigit():
            await update.message.reply_text("❌ Send a valid User ID.", reply_markup=back_admin_kb())
            return
        target_uid2 = int(parts[0])
        rows = get_user_activity(target_uid2, 25)
        if not rows:
            text2 = f"📋 No activity for user `{target_uid2}`."
        else:
            text2 = f"📋 *Activity Log — `{target_uid2}`* (last 25)\n\n"
            for action2, details2, ts2 in rows:
                text2 += f"`{str(ts2)[:16]}` {action2}"
                if details2:
                    text2 += f": _{details2[:40]}_"
                text2 += "\n"
        await update.message.reply_text(text2, parse_mode="Markdown", reply_markup=back_admin_kb())

    elif action == "regex_search":
        pattern = text.strip()
        await update.message.reply_text(f"⏳ Running regex: `{mesc(pattern)}`...", parse_mode="Markdown")
        try:
            results = await asyncio.get_running_loop().run_in_executor(
                _executor, lambda: search_by_regex(pattern, 500)
            )
        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}", reply_markup=back_admin_kb())
            return

        if isinstance(results, dict) and "error" in results:
            await update.message.reply_text(f"❌ Invalid regex: `{results['error']}`", parse_mode="Markdown", reply_markup=back_admin_kb())
            return

        if not results:
            await update.message.reply_text("🔍 No results.", reply_markup=back_admin_kb())
            return

        content  = build_result_txt(pattern, results, "regex")
        content  = add_watermark(content, uid)
        filename = f"regex_{len(results)}_results.txt"
        tmppath  = os.path.join(FILES_DIR, f"tmp_regex_{uid}.txt")
        with open(tmppath, "w", encoding="utf-8") as f:
            f.write(content)
        caption  = f"🔎 *Regex Results*\n`{mesc(pattern)}`\n📊 `{len(results):,}` records"
        await safe_send_document(
            update.message.reply_document, tmppath, filename, caption,
            reply_markup=back_admin_kb()
        )

    elif action == "url_import":
        parts3 = text.strip().split(None, 1)
        if not parts3 or not parts3[0].startswith("http"):
            await update.message.reply_text("❌ Send a valid URL.", reply_markup=back_admin_kb())
            return
        url   = parts3[0]
        label = parts3[1] if len(parts3) > 1 else url.split("/")[-1][:30]
        result = await import_from_url(url, label, uid, context.bot)
        await update.message.reply_text(
            f"{'✅' if result['ok'] else '❌'} *URL Import*\n\n{result['msg']}",
            parse_mode="Markdown",
            reply_markup=back_admin_kb()
        )

    elif action == "create_coupon":
        parts_c = text.strip().split()
        if len(parts_c) < 5:
            await update.message.reply_text(
                "❌ Format: `CODE TYPE VALUE MAX_USES VALID_DAYS [PLAN]`",
                parse_mode="Markdown", reply_markup=back_admin_kb()
            )
            return
        code_c = parts_c[0]
        ctype  = parts_c[1]
        if ctype not in COUPON_TYPES:
            await update.message.reply_text(
                f"❌ Invalid type. Use: {', '.join(COUPON_TYPES.keys())}",
                reply_markup=back_admin_kb()
            )
            return
        try:
            value     = float(parts_c[2])
            max_uses  = int(parts_c[3])
            valid_days= int(parts_c[4])
            plan_r    = parts_c[5] if len(parts_c) > 5 else None
        except (ValueError, IndexError):
            await update.message.reply_text("❌ Invalid numbers.", reply_markup=back_admin_kb())
            return
        cid = create_coupon(code_c, ctype, value, max_uses, valid_days, plan_r, uid)
        await update.message.reply_text(
            f"✅ *Coupon Created!*\n\n"
            f"🏷 Code: `{code_c.upper()}`\n"
            f"📌 Type: `{ctype}` | Value: `{value}`\n"
            f"🔢 Max uses: `{max_uses}` | Valid: `{valid_days}` days",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📋 All Coupons", callback_data="adm_coupons")]])
        )
        log_admin_op(uid, "create_coupon", code_c, f"type={ctype} val={value}")

    elif action == "delete_coupon":
        code_d = text.strip().upper()
        deleted = delete_coupon(code_d)
        await update.message.reply_text(
            f"{'✅ Coupon `' + code_d + '` deactivated.' if deleted else '❌ Coupon not found.'}",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📋 Coupons", callback_data="adm_coupons")]])
        )

# ════════════════════════════════════════════
#     FILE UPLOAD HANDLER
# ════════════════════════════════════════════
MAX_UPLOAD_MB  = 100
_last_upload: dict = {}
_UPLOAD_COOLDOWN   = 1

async def file_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not is_admin(uid):
        await update.message.reply_text(
            "📁 Only the admin can upload files.\n\nUse /start to search.",
            reply_markup=user_main_kb()
        )
        return

    now_t = time.monotonic()
    with _rate_limit_lock:
        last_up = _last_upload.get(uid, 0)
        if (now_t - last_up) < _UPLOAD_COOLDOWN:
            remaining = int(_UPLOAD_COOLDOWN - (now_t - last_up))
            await update.message.reply_text(
                f"⏳ Please wait *{remaining}s* before uploading another file.",
                parse_mode="Markdown"
            )
            return
        _last_upload[uid] = now_t

    doc   = update.message.document
    fname = doc.file_name or "upload.bin"
    ext   = fname.lower().rsplit(".", 1)[-1] if "." in fname else ""

    if ext not in ("txt", "csv", "xlsx", "xls", "json"):
        await update.message.reply_text(
            f"❌ Unsupported: `{mesc(fname)}`\nAllowed: TXT, CSV, XLSX, XLS, JSON",
            parse_mode="Markdown"
        )
        return

    file_size_bytes = doc.file_size or 0
    if file_size_bytes > MAX_UPLOAD_MB * 1024 * 1024:
        await update.message.reply_text(
            f"❌ *File too large!*\n\n"
            f"Max allowed: `{MAX_UPLOAD_MB} MB`\n"
            f"Your file: `{round(file_size_bytes/1024/1024, 1)} MB`",
            parse_mode="Markdown"
        )
        return

    msg = await update.message.reply_text(f"📥 *Downloading* `{fname}`...", parse_mode="Markdown")

    ts_prefix = datetime.utcnow().strftime("%Y%m%d_%H%M%S_")
    save_name = ts_prefix + fname
    save_path = os.path.join(FILES_DIR, save_name)

    try:
        file_obj = await doc.get_file()
        await file_obj.download_to_drive(save_path)
        file_size = os.path.getsize(save_path)
    except Exception as e:
        await msg.edit_text(f"❌ Download failed: `{e}`", parse_mode="Markdown")
        return

    await msg.edit_text(f"⚙️ *Parsing* `{fname}`...", parse_mode="Markdown")

    file_md5 = hashlib.md5(open(save_path, "rb").read()).hexdigest()
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT original_name FROM uploaded_files WHERE file_md5=%s", (file_md5,))
                dup = cur.fetchone()
                if dup:
                    await msg.edit_text(
                        f"⚠️ *Duplicate File Detected!*\n\n"
                        f"This file was already uploaded as `{mesc(dup[0])}`.\n"
                        f"Upload aborted to prevent duplicate records.",
                        parse_mode="Markdown", reply_markup=back_admin_kb()
                    )
                    os.remove(save_path)
                    return
    except Exception as e:
        log.error(f"dup check error: {e}")

    if ext in ("xlsx", "xls"):
        nameid_rows = parse_excel_for_name_id(save_path, fname)
        if nameid_rows:
            try:
                with pool_conn() as conn:
                    with conn.cursor() as cur:
                        for i in range(0, len(nameid_rows), 2000):
                            batch = nameid_rows[i:i+2000]
                            cur.executemany(
                                "INSERT INTO name_id_index (full_name, national_id, source) VALUES (%s,%s,%s) "
                                "ON CONFLICT DO NOTHING",
                                batch
                            )
                        cur.execute(
                            "INSERT INTO uploaded_files (saved_name, original_name, size_bytes, records, uploaded_by, uploaded_at, file_md5) "
                            "VALUES (%s,%s,%s,%s,%s,%s,%s)",
                            (save_name, fname, file_size, len(nameid_rows), uid, datetime.utcnow().isoformat(), file_md5)
                        )
                    conn.commit()
            except Exception as e:
                log.error(f"nameid insert error: {e}")
                await msg.edit_text(f"❌ DB error: `{mesc(str(e))}`", parse_mode="Markdown")
                return
            await msg.edit_text(
                f"✅ *Excel Indexed (Name/ID)*\n\n"
                f"📄 File    : `{mesc(fname)}`\n"
                f"🪪 Records : `{len(nameid_rows):,}` Name/ID rows\n"
                f"💾 Size    : `{round(file_size/1024,1)} KB`",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🗂️ View Files",  callback_data="adm_filelist")],
                    [InlineKeyboardButton("🔙 Admin Panel", callback_data="adm_home")],
                ])
            )
            return
        await msg.edit_text("⚙️ *No Name/ID columns found. Indexing all data...*", parse_mode="Markdown")

    try:
        rows = parse_file(save_path, fname)
    except Exception as e:
        await msg.edit_text(f"❌ Parse failed: `{mesc(str(e))}`", parse_mode="Markdown")
        return

    if not rows:
        await msg.edit_text(f"⚠️ No valid data found in `{mesc(fname)}`.", parse_mode="Markdown")
        return

    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                for i in range(0, len(rows), 2000):
                    cur.executemany(
                        "INSERT INTO data_index (line, source) VALUES (%s,%s)",
                        rows[i:i+2000]
                    )
                cur.execute(
                    "INSERT INTO uploaded_files (saved_name, original_name, size_bytes, records, uploaded_by, uploaded_at, file_md5) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (save_name, fname, file_size, len(rows), uid, datetime.utcnow().isoformat(), file_md5)
                )
            conn.commit()
    except Exception as e:
        log.error(f"data_index insert error: {e}")
        await msg.edit_text(f"❌ DB error: `{mesc(str(e))}`", parse_mode="Markdown")
        return

    await msg.edit_text(
        f"✅ *File Indexed Successfully!*\n\n"
        f"📄 File    : `{mesc(fname)}`\n"
        f"📊 Records : `{len(rows):,}` lines\n"
        f"💾 Size    : `{round(file_size/1024,1)} KB`",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🗂️ View Files",  callback_data="adm_filelist")],
            [InlineKeyboardButton("🔙 Admin Panel", callback_data="adm_home")],
        ])
    )
    log_admin_op(uid, "upload_file", fname, f"{len(rows):,} records | {round(file_size/1024,1)} KB")

# ════════════════════════════════════════════
#               COMMANDS
# ════════════════════════════════════════════
async def cmd_hello(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    uid  = user.id
    ensure_user(uid, user.username or "", user.first_name or "")
    if is_banned(uid):
        await update.message.reply_text("🚫 *Your account has been banned.*", parse_mode="Markdown")
        return
    if is_admin(uid):
        await show_admin_home(update, context, send=True)
    else:
        await show_user_home(update, context, send=True)

async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    ensure_user(uid, update.effective_user.username or "", update.effective_user.first_name or "")
    if is_admin(uid):
        await update.message.reply_text("⚙️ Admin Panel:", reply_markup=admin_main_kb())
    else:
        await update.message.reply_text("ℹ️ Use /start to open the main menu.", reply_markup=user_main_kb())

async def _do_subscribe_request(uid: int, username: str, full_name: str, tier: str, context, reply_fn):
    st = STRINGS.get(get_lang(uid), STRINGS["en"])
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id FROM sub_requests WHERE user_id=%s AND status='pending'", (uid,)
                )
                existing = cur.fetchone()
                if existing:
                    msg = st["sub_req_exists"].format(req_id=existing[0])
                    await reply_fn(msg, parse_mode="Markdown")
                    return
                cur.execute(
                    "INSERT INTO sub_requests (user_id, username, full_name, requested_tier, timestamp) VALUES (%s,%s,%s,%s,%s)",
                    (uid, username or "", full_name or "User", tier, datetime.utcnow().isoformat())
                )
                cur.execute("SELECT lastval()")
                req_id = cur.fetchone()[0]
            conn.commit()
    except Exception as e:
        log.error(f"_do_subscribe_request error: {e}")
        await reply_fn("❌ Error submitting request.", parse_mode="Markdown")
        return

    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=(
                    f"📋 *New Subscription Request!*\n\n"
                    f"🆔 Request ID: `#{req_id}`\n"
                    f"👤 User: `{uid}` — {esc(full_name or 'User')} (@{esc(username or 'N/A')})\n"
                    f"📦 Tier: *{tier}*\n\n"
                    f"Go to 📋 طلبات الاشتراك in the admin panel."
                ),
                parse_mode="Markdown"
            )
        except Exception:
            pass
    msg = st["sub_req_sent"].format(tier=tier, req_id=req_id)
    await reply_fn(msg, parse_mode="Markdown")

async def cmd_subscribe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    uid  = user.id
    ensure_user(uid, user.username or "", user.first_name or "")
    if is_banned(uid):
        await update.message.reply_text("🚫 *Your account has been banned.*", parse_mode="Markdown")
        return

    u = get_user_cached(uid)
    current_tier = u[3] if u else "free"
    t_current    = TIERS.get(current_tier, TIERS["free"])

    args = context.args
    tier = args[0].lower() if args and args[0].lower() in TIERS else None
    if not tier or tier == "free":
        tiers_list = " | ".join(t for t in TIERS if t != "free")
        await update.message.reply_text(
            f"📋 *طلب اشتراك / Subscription Request*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📦 *Current Plan:* `{current_tier}` — {esc(t_current['label'])}\n\n"
            f"استخدم: `/subscribe TIER`\n"
            f"الباقات المتاحة: `{tiers_list}`\n\n"
            f"مثال: `/subscribe premium`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⭐ Basic",   callback_data="sub_req_basic"),
                 InlineKeyboardButton("💎 Premium", callback_data="sub_req_premium")],
                [InlineKeyboardButton("👑 VIP",     callback_data="sub_req_vip")],
            ])
        )
        return

    await _do_subscribe_request(
        uid, user.username or "", user.first_name or "User",
        tier, context, update.message.reply_text
    )

async def cmd_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    await update.message.reply_text(
        f"🆔 *Your Telegram ID:*\n\n`{user.id}`\n\n"
        f"_(Share this with the admin to manage your account)_",
        parse_mode="Markdown"
    )

async def cmd_finduser(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    if not context.args:
        await update.message.reply_text("Usage: `/finduser USER_ID` or `/finduser @username`", parse_mode="Markdown")
        return
    q_val = context.args[0].lstrip("@")
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                if q_val.lstrip("-").isdigit():
                    cur.execute("SELECT * FROM users WHERE user_id=%s", (int(q_val),))
                else:
                    cur.execute("SELECT * FROM users WHERE username ILIKE %s", (f"%{q_val}%",))
                row = cur.fetchone()
    except Exception as e:
        log.error(f"cmd_finduser error: {e}")
        row = None

    if not row:
        await update.message.reply_text(f"❌ User `{mesc(q_val)}` not found.", parse_mode="Markdown")
        return
    uid_r, uname, fname, tier, daily, credits, banned, expires, joined, lang_r, ref_by, ref_cnt, last_s, nameid_lim = row[:14]
    status = "🚫 Banned" if banned else "✅ Active"
    await update.message.reply_text(
        f"👤 *User Info*\n━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 ID       : `{uid_r}`\n"
        f"👤 Name     : {esc(fname or 'N/A')}\n"
        f"🔖 Username : @{esc(uname or 'N/A')}\n"
        f"📦 Tier     : `{tier}`\n"
        f"🔍 Daily    : `{daily}` left\n"
        f"🪪 Name/ID  : `{nameid_lim}` left\n"
        f"💰 Credits  : `{credits}`\n"
        f"📅 Expires  : `{str(expires or 'None')}`\n"
        f"📆 Joined   : `{str(joined or '')[:10]}`\n"
        f"🌐 Lang     : `{lang_r}`\n"
        f"🔗 Referred by: `{ref_by or 'None'}`\n"
        f"👥 Referrals: `{ref_cnt}`\n"
        f"Status     : {status}",
        parse_mode="Markdown", reply_markup=back_admin_kb()
    )

async def cmd_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.pop("search_type", None)
    context.user_data.pop("confirmed_kw", None)
    context.user_data.pop("admin_action", None)
    await update.message.reply_text(
        "❌ Cancelled. Use the menu to start a new search.",
        reply_markup=user_main_kb(update.effective_user.id)
    )

async def cmd_ping(update: Update, context: ContextTypes.DEFAULT_TYPE):
    t0  = time.monotonic()
    msg = await update.message.reply_text("🏓 Pong!")
    latency_ms = round((time.monotonic() - t0) * 1000)
    await msg.edit_text(
        f"🏓 *Pong!*\n⚡ Latency: `{latency_ms}ms`\n🟢 Bot is online.",
        parse_mode="Markdown"
    )

async def cmd_version(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uptime_str = "N/A"
    if BOT_START_TIME:
        delta = datetime.utcnow() - BOT_START_TIME
        h, rem = divmod(int(delta.total_seconds()), 3600)
        m, s   = divmod(rem, 60)
        uptime_str = f"{h}h {m}m {s}s"
    await update.message.reply_text(
        f"🤖 *DATA SCANNER BOT*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 Version : `v9.0-PG`\n"
        f"⏱️ Uptime  : `{uptime_str}`\n"
        f"🐍 Python  : `{__import__('sys').version.split()[0]}`\n"
        f"🐘 DB      : PostgreSQL",
        parse_mode="Markdown"
    )

async def cmd_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM data_index");    tr = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM name_id_index"); tn = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM users");         tu = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM users WHERE is_banned=1"); tb = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM search_logs");   ts = cur.fetchone()[0]
                today = datetime.utcnow().strftime("%Y-%m-%d")
                cur.execute("SELECT COUNT(*) FROM search_logs WHERE timestamp LIKE %s", (f"{today}%",))
                ts_today = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM users WHERE joined_at LIKE %s", (f"{today}%",))
                new_today = cur.fetchone()[0]
    except Exception as e:
        log.error(f"cmd_stats error: {e}")
        await update.message.reply_text("❌ Stats error.")
        return

    await update.message.reply_text(
        f"📊 *Quick Stats*\n━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🗄️ DB Records : `{tr:,}`\n"
        f"🪪 Name/ID    : `{tn:,}`\n"
        f"👥 Users      : `{tu:,}` (🚫 {tb} banned)\n"
        f"🔍 Searches   : `{ts:,}` total | `{ts_today}` today\n"
        f"🆕 New today  : `{new_today}`",
        parse_mode="Markdown", reply_markup=back_admin_kb()
    )

async def cmd_adduser(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    args = context.args
    if len(args) < 2:
        await update.message.reply_text("Usage: /adduser USER_ID TIER [CREDITS]")
        return
    target_id = int(args[0])
    tier      = args[1] if args[1] in TIERS else "free"
    credits   = int(args[2]) if len(args) > 2 else 0
    t  = TIERS[tier]
    nt = NAMEID_TIERS[tier]
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM users WHERE user_id=%s", (target_id,))
                if cur.fetchone():
                    cur.execute(
                        "UPDATE users SET tier=%s, daily_limit=%s, credits=credits+%s WHERE user_id=%s",
                        (tier, t["daily"], credits, target_id)
                    )
                else:
                    cur.execute(
                        "INSERT INTO users (user_id, username, full_name, tier, daily_limit, credits, "
                        "is_banned, expires_at, joined_at, lang, referral_count, daily_nameid_limit) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (target_id, "unknown", "User", tier, t["daily"], credits, 0, None,
                         datetime.utcnow().isoformat(), "en", 0, nt["daily_nameid"])
                    )
            conn.commit()
    except Exception as e:
        log.error(f"cmd_adduser error: {e}")
    await update.message.reply_text(f"✅ User `{target_id}` → `{tier}` + `{credits}` credits.", parse_mode="Markdown")

# ════════════════════════════════════════════
#         GLOBAL ERROR HANDLER
# ════════════════════════════════════════════
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    err_type = type(context.error).__name__
    err_msg  = str(context.error)
    log.error(f"Unhandled exception [{err_type}]: {err_msg}", exc_info=context.error)

    if isinstance(update, Update) and update.effective_message:
        uid_ctx = update.effective_user.id if update.effective_user else "?"
        try:
            if isinstance(context.error, (BadRequest, TelegramError)):
                user_msg = "⚠️ *Telegram error* — please try again."
            elif isinstance(context.error, asyncio.TimeoutError):
                user_msg = "⏱️ *Request timed out.* Try a more specific keyword."
            else:
                user_msg = "⚠️ *Unexpected error.* Please press /start and try again."
            await update.effective_message.reply_text(user_msg, parse_mode="Markdown")
        except Exception:
            pass
    else:
        uid_ctx = "N/A"

    import traceback
    tb_str   = "".join(traceback.format_exception(type(context.error), context.error, context.error.__traceback__))
    err_text = (
        f"⚠️ *Bot Error*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 User: `{uid_ctx}`\n"
        f"🔴 Error: `{mesc(err_type)}`\n"
        f"📝 Message: `{mesc(err_msg[:200])}`\n\n"
        f"```\n{mesc(tb_str[-500:])}\n```"
    )
    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(chat_id=admin_id, text=err_text[:4000], parse_mode="Markdown")
        except Exception:
            pass

# ════════════════════════════════════════════
#     EXPIRY CHECKER
# ════════════════════════════════════════════
async def check_expiry_notifications(app):
    now = datetime.utcnow()
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT user_id, tier, expires_at, lang FROM users WHERE expires_at IS NOT NULL AND is_banned=0"
                )
                rows = cur.fetchall()
    except Exception:
        return

    for uid, tier, exp_str, lang in rows:
        try:
            exp       = datetime.fromisoformat(str(exp_str))
            days_left = (exp - now).days
            if 0 <= days_left <= 3:
                is_ar = (lang or "en") == "ar"
                if is_ar:
                    msg = (
                        f"⚠️ *تنبيه انتهاء الاشتراك!*\n\n"
                        f"📦 باقتك *{tier}* ستنتهي خلال *{days_left}* يوم.\n"
                        f"تواصل مع الأدمن للتجديد."
                    )
                else:
                    msg = (
                        f"⚠️ *Subscription Expiry Notice!*\n\n"
                        f"📦 Your *{tier}* plan expires in *{days_left}* day(s).\n"
                        f"Contact admin to renew."
                    )
                await app.bot.send_message(chat_id=uid, text=msg, parse_mode="Markdown")
        except Exception:
            pass

# ════════════════════════════════════════════
#         ADMIN: MESSAGE SPECIFIC USER
# ════════════════════════════════════════════
async def handle_msg_user(update, context, text):
    uid   = update.effective_user.id
    parts = text.strip().split(" ", 1)
    if len(parts) < 2 or not parts[0].lstrip("-").isdigit():
        await update.message.reply_text(
            "❌ Format: `USER_ID Your message here`\nExample: `123456789 Hello!`",
            parse_mode="Markdown", reply_markup=back_admin_kb()
        )
        return
    target_uid = int(parts[0])
    msg_text   = parts[1].strip()
    try:
        await context.bot.send_message(
            chat_id=target_uid,
            text=f"📩 *Message from Admin:*\n\n{mesc(msg_text)}",
            parse_mode="Markdown"
        )
        await update.message.reply_text(
            f"✅ Message sent to `{target_uid}`.", parse_mode="Markdown", reply_markup=back_admin_kb()
        )
        log_admin_op(uid, "msg_user", str(target_uid), msg_text[:50])
    except Exception as e:
        await update.message.reply_text(f"❌ Failed: `{e}`", parse_mode="Markdown", reply_markup=back_admin_kb())

# ════════════════════════════════════════════
#            AUTO EXPIRE & CLEANUP
# ════════════════════════════════════════════
def auto_expire_subscriptions():
    now = datetime.utcnow().isoformat()
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT user_id, tier FROM users WHERE expires_at IS NOT NULL AND expires_at <= %s AND tier != 'free'",
                    (now,)
                )
                expired = cur.fetchall()
                if expired:
                    cur.execute(
                        "UPDATE users SET tier='free', daily_limit=0, daily_nameid_limit=0, expires_at=NULL "
                        "WHERE expires_at IS NOT NULL AND expires_at <= %s AND tier != 'free'",
                        (now,)
                    )
            conn.commit()
        return expired
    except Exception as e:
        log.error(f"auto_expire_subscriptions error: {e}")
        return []

def cleanup_temp_files():
    removed = 0
    try:
        for fname in os.listdir(FILES_DIR):
            if fname.startswith("tmp_"):
                fpath = os.path.join(FILES_DIR, fname)
                try:
                    age = time.time() - os.path.getmtime(fpath)
                    if age > 3600:
                        os.remove(fpath)
                        removed += 1
                except Exception:
                    pass
    except Exception:
        pass
    return removed

# ════════════════════════════════════════════
#                    MAIN
# ════════════════════════════════════════════
def main():
    global BOT_START_TIME
    BOT_START_TIME = datetime.utcnow()
    init_pool()
    init_db()

    if not TOKEN or not re.match(r"^\d+:[A-Za-z0-9_-]{35,}$", TOKEN):
        log.error("❌ BOT_TOKEN is missing or invalid! Set BOT_TOKEN environment variable.")
        return

    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start",     cmd_start))
    app.add_handler(CommandHandler("hello",     cmd_hello))
    app.add_handler(CommandHandler("help",      cmd_help))
    app.add_handler(CommandHandler("id",        cmd_id))
    app.add_handler(CommandHandler("cancel",    cmd_cancel))
    app.add_handler(CommandHandler("version",   cmd_version))
    app.add_handler(CommandHandler("ping",      cmd_ping))
    app.add_handler(CommandHandler("finduser",  cmd_finduser))
    app.add_handler(CommandHandler("stats",     cmd_stats))
    app.add_handler(CommandHandler("adduser",   cmd_adduser))
    app.add_handler(CommandHandler("subscribe", cmd_subscribe))
    app.add_handler(CallbackQueryHandler(callback_router))
    app.add_handler(MessageHandler(filters.Document.ALL, file_handler))
    app.add_handler(MessageHandler(filters.PHOTO, handle_payment_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_handler))
    app.add_error_handler(error_handler)

    def _shutdown(signum, frame):
        log.info(f"🛑 Received signal {signum} — shutting down gracefully...")
        app.stop_running()

    signal.signal(signal.SIGINT,  _shutdown)
    signal.signal(signal.SIGTERM, _shutdown)

    async def post_init(application):
        await check_expiry_notifications(application)

        async def daily_job():
            while True:
                now           = datetime.utcnow()
                next_midnight = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
                wait_secs     = (next_midnight - now).total_seconds()
                await asyncio.sleep(wait_secs)

                do_daily_reset()
                log.info("✅ Daily limits reset.")

                expired_users = auto_expire_subscriptions()
                for uid_exp, tier_exp in expired_users:
                    try:
                        lang  = get_lang(uid_exp)
                        is_ar = lang == "ar"
                        msg   = (
                            f"⏰ *انتهى اشتراكك!*\n\nباقتك *{tier_exp}* انتهت.\nتواصل مع الأدمن للتجديد."
                            if is_ar else
                            f"⏰ *Subscription Expired!*\n\nYour *{tier_exp}* plan has ended.\nContact admin to renew."
                        )
                        await application.bot.send_message(chat_id=uid_exp, text=msg, parse_mode="Markdown")
                    except Exception:
                        pass
                if expired_users:
                    log.info(f"⏰ Auto-expired {len(expired_users)} subscriptions.")

                try:
                    backup_db()
                    log.info("✅ Auto backup marker done.")
                except Exception as e:
                    log.warning(f"Auto backup failed: {e}")

                await check_expiry_notifications(application)

                removed = cleanup_temp_files()
                if removed:
                    log.info(f"🗑️ Cleaned up {removed} temp files.")

                # ── Daily stats snapshot ─────────────────────
                try:
                    record_daily_stats()
                    log.info("📊 Daily stats recorded.")
                except Exception as e:
                    log.warning(f"Daily stats error: {e}")

                # ── Daily report to admin ─────────────────────
                try:
                    await send_daily_report(application)
                    log.info("📊 Daily report sent to admins.")
                except Exception as e:
                    log.warning(f"Daily report error: {e}")

                # ── Auto-renewal reminders ────────────────────
                try:
                    await check_auto_renewals(application)
                    log.info("🔄 Auto-renewal check done.")
                except Exception as e:
                    log.warning(f"Auto-renewal check error: {e}")

                # ── Run scheduled imports ─────────────────────
                try:
                    await run_scheduled_imports(application)
                except Exception as e:
                    log.warning(f"Scheduled imports error: {e}")

                # ── Auto-expire pending payment orders ────────
                try:
                    cutoff = (datetime.utcnow() - timedelta(hours=PAYMENT_PENDING_TIMEOUT)).isoformat()
                    with pool_conn() as conn:
                        with conn.cursor() as cur:
                            cur.execute(
                                "UPDATE payment_orders SET status='expired' WHERE status='pending' AND created_at < %s",
                                (cutoff,)
                            )
                            expired_orders = cur.rowcount
                        conn.commit()
                    if expired_orders:
                        log.info(f"⏳ Auto-expired {expired_orders} payment orders.")
                except Exception as e:
                    log.warning(f"Order expiry error: {e}")

        asyncio.create_task(daily_job())
        asyncio.create_task(scheduled_import_loop(application))

    app.post_init = post_init
    log.info("🚀 Data Scanner Bot v9.0-PG running with PostgreSQL...")

    webhook_url = os.environ.get("WEBHOOK_URL", "").strip()
    if webhook_url:
        port = int(os.environ.get("PORT", 8443))
        log.info(f"🌐 Webhook mode: {webhook_url} on port {port}")
        app.run_webhook(listen="0.0.0.0", port=port, webhook_url=webhook_url)
    else:
        log.info("🔄 Polling mode")
        app.run_polling()


# ════════════════════════════════════════════
#         PHASE 1 — IP TRACKING
# ════════════════════════════════════════════
def log_ip(user_id: int, ip_address: str, action: str = "request"):
    if not IP_TRACK_ENABLED or not ip_address:
        return
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO ip_logs (user_id, ip_address, action, timestamp) VALUES (%s,%s,%s,%s)",
                    (user_id, ip_address, action, datetime.utcnow().isoformat())
                )
            conn.commit()
    except Exception as e:
        log.error(f"log_ip error: {e}")

def get_user_ips(user_id: int) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT DISTINCT ip_address, MAX(timestamp) as last_seen, COUNT(*) as hits "
                    "FROM ip_logs WHERE user_id=%s GROUP BY ip_address ORDER BY last_seen DESC LIMIT 10",
                    (user_id,)
                )
                return cur.fetchall()
    except Exception:
        return []

def get_ip_users(ip_address: str) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT DISTINCT user_id, MAX(timestamp) as last_seen FROM ip_logs "
                    "WHERE ip_address=%s GROUP BY user_id ORDER BY last_seen DESC LIMIT 20",
                    (ip_address,)
                )
                return cur.fetchall()
    except Exception:
        return []

# ════════════════════════════════════════════
#         PHASE 1 — AUTO BAN SYSTEM
# ════════════════════════════════════════════
def log_failed_attempt(user_id: int, reason: str):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO failed_attempts (user_id, reason, timestamp) VALUES (%s,%s,%s)",
                    (user_id, reason, datetime.utcnow().isoformat())
                )
                # Check if should auto-ban
                cur.execute(
                    "SELECT COUNT(*) FROM failed_attempts WHERE user_id=%s AND timestamp > %s",
                    (user_id, (datetime.utcnow() - timedelta(hours=1)).isoformat())
                )
                count = cur.fetchone()[0]
                if count >= AUTO_BAN_THRESHOLD:
                    cur.execute(
                        "UPDATE users SET is_banned=1 WHERE user_id=%s",
                        (user_id,)
                    )
                    invalidate_user_cache(user_id)
                    log.warning(f"🚫 Auto-banned user {user_id} after {count} failed attempts")
            conn.commit()
    except Exception as e:
        log.error(f"log_failed_attempt error: {e}")

# ════════════════════════════════════════════
#         PHASE 1 — POINTS SYSTEM
# ════════════════════════════════════════════
def add_points(user_id: int, points: int, reason: str = ""):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """INSERT INTO user_points (user_id, points, total_earned, last_updated)
                       VALUES (%s, %s, %s, %s)
                       ON CONFLICT (user_id) DO UPDATE SET
                       points = user_points.points + %s,
                       total_earned = user_points.total_earned + %s,
                       last_updated = %s""",
                    (user_id, points, points, datetime.utcnow().isoformat(),
                     points, points, datetime.utcnow().isoformat())
                )
            conn.commit()
        log.info(f"Points +{points} to user {user_id} — {reason}")
    except Exception as e:
        log.error(f"add_points error: {e}")

def get_points(user_id: int) -> int:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT points FROM user_points WHERE user_id=%s", (user_id,))
                row = cur.fetchone()
        return row[0] if row else 0
    except Exception:
        return 0

def get_points_leaderboard(limit: int = 10) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """SELECT up.user_id, u.full_name, up.points, up.total_earned
                       FROM user_points up
                       LEFT JOIN users u ON u.user_id = up.user_id
                       ORDER BY up.points DESC LIMIT %s""",
                    (limit,)
                )
                return cur.fetchall()
    except Exception:
        return []

# ════════════════════════════════════════════
#         PHASE 1 — WATERMARK
# ════════════════════════════════════════════
def add_watermark(content: str, user_id: int) -> str:
    """Add watermark to result files."""
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    watermark = (
        f"\n{'═'*55}\n"
        f"  🤖 {WATERMARK_TEXT}\n"
        f"  👤 User ID : {user_id}\n"
        f"  🕐 Generated: {now}\n"
        f"  ⚠️  For authorized use only\n"
        f"{'═'*55}\n"
    )
    return content + watermark

# ════════════════════════════════════════════
#         PHASE 1 — EXCEL EXPORT
# ════════════════════════════════════════════
def build_result_excel(keyword: str, results: list, stype: str, user_id: int) -> str:
    """Build Excel file from search results."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header style
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["#", "URL/Domain", "Email", "Username", "Password", "Phone"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, r in enumerate(results, 1):
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=r.get("url") or r.get("domain", ""))
        ws.cell(row=i+1, column=3, value=r.get("email", ""))
        ws.cell(row=i+1, column=4, value=r.get("username", ""))
        ws.cell(row=i+1, column=5, value=r.get("password", ""))
        ws.cell(row=i+1, column=6, value=r.get("phone", ""))

    # Info sheet
    ws2 = wb.create_sheet("Info")
    ws2.append(["Field", "Value"])
    ws2.append(["Keyword", keyword])
    ws2.append(["Type", stype.upper()])
    ws2.append(["Total Results", len(results)])
    ws2.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    ws2.append(["Bot", WATERMARK_TEXT])
    ws2.append(["User ID", user_id])

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    path = os.path.join(FILES_DIR, f"tmp_excel_{user_id}_{int(time.time())}.xlsx")
    wb.save(path)
    return path

# ════════════════════════════════════════════
#     PHASE 1 — IP / IBAN / ADDRESS SEARCH
# ════════════════════════════════════════════
def is_ip_address(value: str) -> bool:
    return bool(re.fullmatch(r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}(/\d{1,2})?", value.strip()))

def is_iban(value: str) -> bool:
    cleaned = re.sub(r"\s", "", value).upper()
    return bool(re.fullmatch(r"[A-Z]{2}\d{2}[A-Z0-9]{4,30}", cleaned))

def detect_advanced_search_type(query: str) -> str:
    q = query.strip()
    if is_ip_address(q):
        return "ip"
    if is_iban(q):
        return "iban"
    # Check for address patterns (Arabic or contains city names)
    if re.search(r"شارع|حي|مدينة|محافظة|street|district|city|address", q, re.I):
        return "address"
    return None

def search_by_ip(ip: str, limit: int = 100) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT line FROM data_index WHERE line ILIKE %s LIMIT %s",
                    (f"%{ip}%", limit)
                )
                rows = cur.fetchall()
        return [{"line": r[0], "ip": ip} for r in rows]
    except Exception as e:
        log.error(f"search_by_ip error: {e}")
        return []

def search_by_iban(iban: str, limit: int = 100) -> list:
    cleaned = re.sub(r"\s", "", iban).upper()
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT line FROM data_index WHERE line ILIKE %s LIMIT %s",
                    (f"%{cleaned}%", limit)
                )
                rows = cur.fetchall()
        return [{"line": r[0], "iban": cleaned} for r in rows]
    except Exception as e:
        log.error(f"search_by_iban error: {e}")
        return []

def search_by_address(address: str, limit: int = 100) -> list:
    words = [w for w in address.split() if len(w) >= 3]
    if not words:
        return []
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                conditions = " AND ".join(["line ILIKE %s"] * len(words))
                params = [f"%{w}%" for w in words] + [limit]
                cur.execute(
                    f"SELECT line FROM data_index WHERE {conditions} LIMIT %s",
                    params
                )
                rows = cur.fetchall()
        return [{"line": r[0]} for r in rows]
    except Exception as e:
        log.error(f"search_by_address error: {e}")
        return []

def build_advanced_result_txt(query: str, results: list, qtype: str, user_id: int) -> str:
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    type_labels = {"ip": "🌐 IP Address", "iban": "🏦 IBAN", "address": "📍 Address/Location"}
    lines = [
        "═" * 55,
        f"  🔍 DATA SCANNER BOT v9.0 — ADVANCED SEARCH",
        "═" * 55,
        f"  📌 Query  : {query}",
        f"  📂 Type   : {type_labels.get(qtype, qtype)}",
        f"  📊 Total  : {len(results):,} records",
        f"  🕐 Time   : {now}",
        f"  👤 User   : {user_id}",
        "═" * 55, "",
    ]
    for r in results:
        lines.append(r.get("line", ""))
    lines += ["", "═"*55, f"  ✅ Total: {len(results):,} records", "═"*55]
    content = "\n".join(lines)
    return add_watermark(content, user_id)

# ════════════════════════════════════════════
#         PHASE 1 — BULK SEARCH
# ════════════════════════════════════════════
def do_bulk_search(keywords: list, stype: str, limit_per_kw: int) -> dict:
    results = {}
    for kw in keywords[:BULK_SEARCH_MAX]:
        kw = kw.strip()
        if len(kw) < MIN_KEYWORD_LEN:
            continue
        r = smart_search(kw, stype, limit_per_kw)
        results[kw] = r
    return results

def build_bulk_result_txt(results: dict, stype: str, user_id: int) -> str:
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    lines = [
        "═" * 55,
        "  🔍 DATA SCANNER BOT v9.0 — BULK SEARCH RESULTS",
        "═" * 55,
        f"  📂 Type     : {stype.upper()}",
        f"  🔢 Keywords : {len(results)}",
        f"  🕐 Time     : {now}",
        "═" * 55, "",
    ]
    total = 0
    for kw, res in results.items():
        lines.append(f"{'─'*55}")
        lines.append(f"  🎯 {kw} — {len(res):,} results")
        lines.append(f"{'─'*55}")
        for r in res:
            url = r.get("url") or r.get("domain", "")
            em  = r.get("email", "")
            pwd = r.get("password", "")
            if url and (em or r.get("username")) and pwd:
                lines.append(f"{url}|{em or r.get('username')}:{pwd}")
            elif em and pwd:
                lines.append(f"{em}:{pwd}")
            elif em:
                lines.append(em)
        lines.append("")
        total += len(res)
    lines += ["═"*55, f"  ✅ Total across all keywords: {total:,} records", "═"*55]
    content = "\n".join(lines)
    return add_watermark(content, user_id)

# ════════════════════════════════════════════
#    PHASE 1 — FAVORITES / SAVED SEARCHES
# ════════════════════════════════════════════
def save_favorite(user_id: int, keyword: str, stype: str, label: str = "") -> bool:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT COUNT(*) FROM user_favorites WHERE user_id=%s",
                    (user_id,)
                )
                if cur.fetchone()[0] >= 20:
                    return False  # max 20 favorites
                cur.execute(
                    "INSERT INTO user_favorites (user_id, keyword, stype, label, created_at) "
                    "VALUES (%s,%s,%s,%s,%s)",
                    (user_id, keyword, stype, label or keyword, datetime.utcnow().isoformat())
                )
            conn.commit()
        return True
    except Exception as e:
        log.error(f"save_favorite error: {e}")
        return False

def get_favorites(user_id: int) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, keyword, stype, label, created_at FROM user_favorites "
                    "WHERE user_id=%s ORDER BY created_at DESC",
                    (user_id,)
                )
                return cur.fetchall()
    except Exception:
        return []

def delete_favorite(fav_id: int, user_id: int) -> bool:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "DELETE FROM user_favorites WHERE id=%s AND user_id=%s",
                    (fav_id, user_id)
                )
            conn.commit()
        return True
    except Exception:
        return False

# ════════════════════════════════════════════
#    PHASE 1 — SUPPORT TICKETS
# ════════════════════════════════════════════
def create_ticket(user_id: int, username: str, full_name: str, subject: str, message: str) -> int:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """INSERT INTO support_tickets
                       (user_id, username, full_name, subject, message, status, created_at, updated_at)
                       VALUES (%s,%s,%s,%s,%s,'open',%s,%s) RETURNING id""",
                    (user_id, username, full_name, subject, message,
                     datetime.utcnow().isoformat(), datetime.utcnow().isoformat())
                )
                ticket_id = cur.fetchone()[0]
            conn.commit()
        return ticket_id
    except Exception as e:
        log.error(f"create_ticket error: {e}")
        return 0

def get_user_tickets(user_id: int) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, subject, status, created_at, admin_reply FROM support_tickets "
                    "WHERE user_id=%s ORDER BY created_at DESC LIMIT 10",
                    (user_id,)
                )
                return cur.fetchall()
    except Exception:
        return []

def get_open_tickets() -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, user_id, username, full_name, subject, message, created_at "
                    "FROM support_tickets WHERE status='open' ORDER BY created_at ASC LIMIT 20"
                )
                return cur.fetchall()
    except Exception:
        return []

def reply_ticket(ticket_id: int, reply: str) -> bool:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE support_tickets SET admin_reply=%s, status='closed', updated_at=%s WHERE id=%s",
                    (reply, datetime.utcnow().isoformat(), ticket_id)
                )
            conn.commit()
        return True
    except Exception:
        return False

# ════════════════════════════════════════════
#    PHASE 1 — AFFILIATE SYSTEM
# ════════════════════════════════════════════
def get_or_create_affiliate(user_id: int) -> str:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT code FROM affiliate_links WHERE user_id=%s", (user_id,))
                row = cur.fetchone()
                if row:
                    return row[0]
                code = hashlib.md5(f"{user_id}{time.time()}".encode()).hexdigest()[:8].upper()
                cur.execute(
                    "INSERT INTO affiliate_links (user_id, code, created_at) VALUES (%s,%s,%s)",
                    (user_id, code, datetime.utcnow().isoformat())
                )
            conn.commit()
        return code
    except Exception as e:
        log.error(f"get_or_create_affiliate error: {e}")
        return ""

def get_affiliate_stats(user_id: int) -> dict:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT code, clicks, conversions, earnings FROM affiliate_links WHERE user_id=%s",
                    (user_id,)
                )
                row = cur.fetchone()
        if not row:
            return {}
        return {"code": row[0], "clicks": row[1], "conversions": row[2], "earnings": row[3]}
    except Exception:
        return {}

def track_affiliate_click(code: str):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE affiliate_links SET clicks=clicks+1 WHERE code=%s",
                    (code,)
                )
            conn.commit()
    except Exception:
        pass

def track_affiliate_conversion(code: str, credits: int = REFERRAL_CREDITS):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE affiliate_links SET conversions=conversions+1, earnings=earnings+%s WHERE code=%s",
                    (credits, code)
                )
                cur.execute("SELECT user_id FROM affiliate_links WHERE code=%s", (code,))
                row = cur.fetchone()
            conn.commit()
        if row:
            add_points(row[0], POINTS_PER_REFERRAL, "affiliate_conversion")
    except Exception:
        pass

# ════════════════════════════════════════════
#    PHASE 1 — ADMIN INSTANT NOTIFICATIONS
# ════════════════════════════════════════════
async def notify_admins(app, message: str, parse_mode: str = "Markdown"):
    for admin_id in ADMIN_IDS:
        try:
            await app.bot.send_message(chat_id=admin_id, text=message, parse_mode=parse_mode)
        except Exception:
            pass

async def notify_admins_new_user(app, user):
    msg = (
        f"🆕 *New User!*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 ID: `{user.id}`\n"
        f"👤 Name: {user.first_name or 'N/A'}\n"
        f"🔖 Username: @{user.username or 'N/A'}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━"
    )
    await notify_admins(app, msg)

async def notify_admins_sub_request(app, user_id: int, username: str, tier: str, req_id: int):
    msg = (
        f"📋 *New Subscription Request!*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 User: @{username} (`{user_id}`)\n"
        f"📦 Plan: *{tier}*\n"
        f"🔢 Request ID: `#{req_id}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━"
    )
    kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton(f"✅ Approve #{req_id}", callback_data=f"adm_approve_req_{req_id}"),
            InlineKeyboardButton(f"❌ Reject #{req_id}",  callback_data=f"adm_reject_req_{req_id}"),
        ]
    ])
    for admin_id in ADMIN_IDS:
        try:
            await app.bot.send_message(chat_id=admin_id, text=msg, parse_mode="Markdown", reply_markup=kb)
        except Exception:
            pass

async def notify_admins_new_ticket(app, ticket_id: int, user_id: int, username: str, subject: str):
    msg = (
        f"🎫 *New Support Ticket!*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🔢 Ticket: `#{ticket_id}`\n"
        f"👤 User: @{username} (`{user_id}`)\n"
        f"📝 Subject: {subject}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━"
    )
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton(f"💬 Reply #{ticket_id}", callback_data=f"adm_reply_ticket_{ticket_id}")
    ]])
    for admin_id in ADMIN_IDS:
        try:
            await app.bot.send_message(chat_id=admin_id, text=msg, parse_mode="Markdown", reply_markup=kb)
        except Exception:
            pass

# ════════════════════════════════════════════
#    PHASE 1 — USER HANDLERS (NEW FEATURES)
# ════════════════════════════════════════════
async def show_points(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    pts  = get_points(uid)
    lb   = get_points_leaderboard(10)
    is_ar = get_lang(uid) == "ar"

    lb_text = ""
    medals = ["🥇", "🥈", "🥉"] + ["4️⃣","5️⃣","6️⃣","7️⃣","8️⃣","9️⃣","🔟"]
    for i, (lid, lname, lpts, _) in enumerate(lb):
        marker = " ◀️" if lid == uid else ""
        lb_text += f"{medals[i]} {lname or lid}: `{lpts:,}` pts{marker}\n"

    text = (
        f"🏆 *{'نقاطي' if is_ar else 'My Points'}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"⭐ {'نقاطك' if is_ar else 'Your Points'}: `{pts:,}`\n\n"
        f"🏅 *{'المتصدرون' if is_ar else 'Leaderboard'}*\n"
        f"{'─'*30}\n"
        f"{lb_text or 'No data yet'}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💡 {'اكسب نقاط بالبحث والإحالات' if is_ar else 'Earn points by searching & referring friends'}"
    )
    kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="user_home")]])
    if update.message:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.callback_query.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)

async def show_favorites(update: Update, context: ContextTypes.DEFAULT_TYPE, query=None):
    uid  = (query or update.callback_query).from_user.id if query or update.callback_query else update.effective_user.id
    favs = get_favorites(uid)
    is_ar = get_lang(uid) == "ar"

    if not favs:
        text = "⭐ *No favorites yet!*\n\nAfter a search, tap ⭐ Save to favorites."
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="user_home")]])
    else:
        text = f"⭐ *{'المفضلة' if is_ar else 'Saved Favorites'}* — {len(favs)}/20\n\n"
        rows = []
        for fid, kw, stype, label, _ in favs:
            rows.append([
                InlineKeyboardButton(f"🔍 {label[:25]} [{stype}]", callback_data=_cb_put(f"confirm_search:{stype}:{kw}")),
                InlineKeyboardButton("🗑️", callback_data=f"del_fav_{fid}"),
            ])
        rows.append([InlineKeyboardButton("🔙 Back", callback_data="user_home")])
        kb = InlineKeyboardMarkup(rows)

    q = query or (update.callback_query if update.callback_query else None)
    if q:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

async def show_support_menu(update: Update, context: ContextTypes.DEFAULT_TYPE, query=None):
    uid   = update.effective_user.id
    is_ar = get_lang(uid) == "ar"
    tickets = get_user_tickets(uid)

    text = (
        f"🎫 *{'الدعم الفني' if is_ar else 'Support Center'}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
    )
    if tickets:
        text += f"📋 {'تذاكرك' if is_ar else 'Your tickets'}:\n"
        for tid, subj, status, cat, reply in tickets:
            emoji = "✅" if status == "closed" else "⏳"
            text += f"{emoji} `#{tid}` — {subj[:30]}\n"
        text += "\n"
    else:
        text += f"{'لا توجد تذاكر بعد.' if is_ar else 'No tickets yet.'}\n\n"

    text += f"{'اضغط لفتح تذكرة جديدة.' if is_ar else 'Tap to open a new ticket.'}"

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ " + ("تذكرة جديدة" if is_ar else "New Ticket"), callback_data="support_new_ticket")],
        [InlineKeyboardButton("🔙 Back", callback_data="user_home")],
    ])
    q = query or (update.callback_query if update.callback_query else None)
    if q:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

async def show_affiliate_stats(update: Update, context: ContextTypes.DEFAULT_TYPE, query=None):
    uid   = update.effective_user.id
    is_ar = get_lang(uid) == "ar"
    code  = get_or_create_affiliate(uid)
    stats = get_affiliate_stats(uid)
    bot_username = context.bot.username if context.bot else "DataScannerBot"
    link  = f"https://t.me/{bot_username}?start=ref_{code}"

    text = (
        f"🔗 *{'رابط الإحالة المتقدم' if is_ar else 'Affiliate Dashboard'}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🔑 {'كودك' if is_ar else 'Your Code'}: `{code}`\n"
        f"🔗 {'رابطك' if is_ar else 'Your Link'}:\n`{link}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👆 {'نقرات' if is_ar else 'Clicks'}: `{stats.get('clicks', 0):,}`\n"
        f"✅ {'تحويلات' if is_ar else 'Conversions'}: `{stats.get('conversions', 0):,}`\n"
        f"💰 {'أرباح (رصيد)' if is_ar else 'Earnings (credits)'}: `{stats.get('earnings', 0):,}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💡 {'كل تحويل = {REFERRAL_CREDITS} رصيد + {POINTS_PER_REFERRAL} نقطة' if is_ar else f'Each conversion = {REFERRAL_CREDITS} credits + {POINTS_PER_REFERRAL} points'}"
    )
    kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="user_home")]])
    q = query or (update.callback_query if update.callback_query else None)
    if q:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

async def show_bulk_search_menu(update: Update, context: ContextTypes.DEFAULT_TYPE, query=None):
    uid   = update.effective_user.id
    is_ar = get_lang(uid) == "ar"
    u     = get_user_cached(uid)
    tier  = u[3] if u else "free"

    if tier not in ("premium", "vip") and not is_admin(uid):
        text = (
            "❌ *Bulk Search — Premium Feature*\n\n"
            "This feature requires Premium or VIP plan.\n\n"
            "💎 Upgrade to search multiple keywords at once!"
        )
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("💳 View Plans", callback_data="show_plans")],
            [InlineKeyboardButton("🔙 Back", callback_data="user_home")],
        ])
    else:
        text = (
            f"🔎 *{'البحث المتعدد' if is_ar else 'Bulk Search'}*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"{'ابعت كلمات البحث، كل كلمة في سطر (أقصى ' + str(BULK_SEARCH_MAX) + ' كلمة):' if is_ar else f'Send keywords, one per line (max {BULK_SEARCH_MAX}):'}\n\n"
            f"_مثال / Example:_\n"
            f"`google.com\nfacebook.com\nyahoo.com`"
        )
        context.user_data["bulk_search_waiting"] = True
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="user_home")]])

    q = query or (update.callback_query if update.callback_query else None)
    if q:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

# ════════════════════════════════════════════
#    PHASE 1 — ADMIN TICKETS HANDLER
# ════════════════════════════════════════════
async def show_admin_tickets(update, context, query=None):
    tickets = get_open_tickets()
    if not tickets:
        text = "✅ *No open tickets!*"
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="adm_home")]])
    else:
        text = f"🎫 *Open Tickets* — {len(tickets)}\n\n"
        rows = []
        for tid, uid, uname, fname, subj, msg, cat in tickets:
            text += f"`#{tid}` @{uname or uid} — {subj[:25]}\n"
            rows.append([InlineKeyboardButton(
                f"💬 Reply #{tid}", callback_data=f"adm_reply_ticket_{tid}"
            )])
        rows.append([InlineKeyboardButton("🔙 Back", callback_data="adm_home")])
        kb = InlineKeyboardMarkup(rows)

    q = query or (update.callback_query if update.callback_query else None)
    if q:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

# ════════════════════════════════════════════
#    PHASE 1 — SEARCH HISTORY (DETAILED)
# ════════════════════════════════════════════
async def show_search_history(update, context, query=None):
    uid = (query or update.callback_query).from_user.id if (query or update.callback_query) else update.effective_user.id
    is_ar = get_lang(uid) == "ar"

    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT keyword, category, results, timestamp FROM search_logs "
                    "WHERE user_id=%s ORDER BY timestamp DESC LIMIT 20",
                    (uid,)
                )
                rows = cur.fetchall()
    except Exception:
        rows = []

    if not rows:
        text = f"📜 {'لا يوجد تاريخ بحث بعد.' if is_ar else 'No search history yet.'}"
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="user_home")]])
    else:
        text = f"📜 *{'تاريخ البحث' if is_ar else 'Search History'}* — آخر 20\n\n"
        kb_rows = []
        for kw, cat, res, ts in rows:
            ts_str = str(ts)[:16] if ts else ""
            text += f"🔍 `{kw}` [{cat}] — {res:,} results — {ts_str}\n"
            kb_rows.append([InlineKeyboardButton(
                f"🔄 {kw[:20]} [{cat}]",
                callback_data=_cb_put(f"confirm_search:{cat}:{kw}")
            )])
        kb_rows.append([InlineKeyboardButton("🔙 Back", callback_data="user_home")])
        kb = InlineKeyboardMarkup(kb_rows)

    q = query or (update.callback_query if update.callback_query else None)
    if q:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

# ════════════════════════════════════════════
#    PHASE 1 — ADVANCED SEARCH HANDLER
# ════════════════════════════════════════════
async def handle_advanced_search(update: Update, context: ContextTypes.DEFAULT_TYPE, query: str):
    uid   = update.effective_user.id
    qtype = detect_advanced_search_type(query)
    if not qtype:
        return False  # not an advanced search

    if not can_search(uid):
        await update.message.reply_text(
            "❌ No searches remaining. Upgrade your plan.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("💳 Plans", callback_data="show_plans")]])
        )
        return True

    msg = await update.message.reply_text(
        f"⏳ *Advanced Search...*\n\n"
        f"🎯 Query: `{mesc(query)}`\n"
        f"📂 Type: `{qtype.upper()}`",
        parse_mode="Markdown"
    )

    loop = asyncio.get_running_loop()
    if qtype == "ip":
        results = await loop.run_in_executor(_executor, lambda: search_by_ip(query, 500))
    elif qtype == "iban":
        results = await loop.run_in_executor(_executor, lambda: search_by_iban(query, 500))
    else:
        results = await loop.run_in_executor(_executor, lambda: search_by_address(query, 500))

    if not results:
        await msg.edit_text(
            f"🔍 *No results found*\n\nQuery: `{mesc(query)}`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Menu", callback_data="user_home")]])
        )
        return True

    content  = build_advanced_result_txt(query, results, qtype, uid)
    safe_kw  = re.sub(r"[^\w\-]", "_", query)[:20]
    filename = f"{qtype}_{safe_kw}_{len(results)}_results.txt"
    tmppath  = os.path.join(FILES_DIR, f"tmp_adv_{uid}.txt")
    with open(tmppath, "w", encoding="utf-8") as f:
        f.write(content)

    caption = (
        f"✅ *Advanced Search Results*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🎯 Query: `{mesc(query)}`\n"
        f"📂 Type: `{qtype.upper()}`\n"
        f"📊 Total: `{len(results):,}` records"
    )

    await msg.delete()
    await safe_send_document(
        update.message.reply_document, tmppath, filename, caption,
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Menu", callback_data="user_home")]])
    )

    if not is_admin(uid):
        deduct(uid)
        add_points(uid, POINTS_PER_SEARCH, "advanced_search")
    log_search(uid, query, f"adv_{qtype}", len(results))
    return True


# ════════════════════════════════════════════
#         PAYMENT SYSTEM — DB HELPERS
# ════════════════════════════════════════════
def create_payment_order(user_id, plan, duration, method, amount_usd, amount_egp=None, currency="USD") -> int:
    try:
        now = datetime.utcnow().isoformat()
        exp = (datetime.utcnow() + timedelta(hours=PAYMENT_PENDING_TIMEOUT)).isoformat()
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO payment_orders
                    (user_id, plan, duration, method, amount_usd, amount_egp, currency, status, created_at, expires_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,'pending',%s,%s) RETURNING id
                """, (user_id, plan, duration, method, amount_usd, amount_egp, currency, now, exp))
                order_id = cur.fetchone()[0]
            conn.commit()
        return order_id
    except Exception as e:
        log.error(f"create_payment_order error: {e}")
        return 0

def get_order(order_id: int) -> dict:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT id, user_id, plan, duration, method, amount_usd, amount_egp,
                           currency, status, tx_id, screenshot_file_id, created_at, expires_at, confirmed_at, notes
                    FROM payment_orders WHERE id=%s
                """, (order_id,))
                r = cur.fetchone()
        if not r:
            return {}
        keys = ["id","user_id","plan","duration","method","amount_usd","amount_egp",
                "currency","status","tx_id","screenshot_file_id","created_at","expires_at","confirmed_at","notes"]
        return dict(zip(keys, r))
    except Exception:
        return {}

def get_user_orders(user_id: int, limit=10) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT id, plan, duration, method, amount_usd, status, created_at
                    FROM payment_orders WHERE user_id=%s ORDER BY created_at DESC LIMIT %s
                """, (user_id, limit))
                return cur.fetchall()
    except Exception:
        return []

def get_pending_orders(limit=50) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT po.id, po.user_id, u.username, u.full_name, po.plan, po.duration,
                           po.method, po.amount_usd, po.amount_egp, po.currency,
                           po.tx_id, po.screenshot_file_id, po.created_at
                    FROM payment_orders po
                    LEFT JOIN users u ON u.user_id = po.user_id
                    WHERE po.status='pending'
                    ORDER BY po.created_at ASC LIMIT %s
                """, (limit,))
                return cur.fetchall()
    except Exception:
        return []

def confirm_order(order_id: int, admin_id: int) -> dict:
    order = get_order(order_id)
    if not order or order["status"] != "pending":
        return {"ok": False, "msg": "Order not found or already processed"}
    try:
        plan      = order["plan"]
        duration  = order["duration"]
        user_id   = order["user_id"]
        days      = PLAN_DURATIONS.get(duration, PLAN_DURATIONS["1month"])["days"]
        t         = TIERS.get(plan, TIERS["free"])
        nt        = NAMEID_TIERS.get(plan, NAMEID_TIERS["free"])
        new_exp   = (datetime.utcnow() + timedelta(days=days)).isoformat()

        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE users SET tier=%s, daily_limit=%s, daily_nameid_limit=%s, expires_at=%s
                    WHERE user_id=%s
                """, (plan, t["daily"], nt["daily_nameid"], new_exp, user_id))
                cur.execute("""
                    UPDATE payment_orders SET status='confirmed', admin_id=%s, confirmed_at=%s WHERE id=%s
                """, (admin_id, datetime.utcnow().isoformat(), order_id))
                cur.execute("""
                    INSERT INTO invoices (order_id, user_id, plan, duration, amount, currency, method, status, issued_at, paid_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,'paid',%s,%s)
                """, (order_id, user_id, plan, duration, order["amount_usd"], order["currency"],
                      order["method"], datetime.utcnow().isoformat(), datetime.utcnow().isoformat()))
                cur.execute("""
                    INSERT INTO sub_history (user_id, tier, amount, admin_id, timestamp)
                    VALUES (%s,%s,%s,%s,%s)
                """, (user_id, plan, int(order["amount_usd"]), admin_id, datetime.utcnow().isoformat()))
            conn.commit()
        invalidate_user_cache(user_id)
        add_points(user_id, 50, "payment_confirmed")
        return {"ok": True, "user_id": user_id, "plan": plan, "days": days, "expires": new_exp}
    except Exception as e:
        log.error(f"confirm_order error: {e}")
        return {"ok": False, "msg": str(e)}

def reject_order(order_id: int, admin_id: int, reason: str = "") -> dict:
    order = get_order(order_id)
    if not order:
        return {"ok": False, "msg": "Order not found"}
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE payment_orders SET status='rejected', admin_id=%s, notes=%s WHERE id=%s
                """, (admin_id, reason, order_id))
            conn.commit()
        return {"ok": True, "user_id": order["user_id"]}
    except Exception as e:
        return {"ok": False, "msg": str(e)}

def update_order_screenshot(order_id: int, file_id: str):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE payment_orders SET screenshot_file_id=%s WHERE id=%s", (file_id, order_id))
            conn.commit()
    except Exception:
        pass

def update_order_txid(order_id: int, tx_id: str):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE payment_orders SET tx_id=%s WHERE id=%s", (tx_id, order_id))
            conn.commit()
    except Exception:
        pass

def get_invoice(order_id: int) -> dict:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT i.id, i.order_id, i.user_id, u.full_name, u.username,
                           i.plan, i.duration, i.amount, i.currency, i.method, i.status, i.issued_at, i.paid_at
                    FROM invoices i LEFT JOIN users u ON u.user_id=i.user_id
                    WHERE i.order_id=%s
                """, (order_id,))
                r = cur.fetchone()
        if not r:
            return {}
        keys = ["id","order_id","user_id","full_name","username","plan","duration",
                "amount","currency","method","status","issued_at","paid_at"]
        return dict(zip(keys, r))
    except Exception:
        return {}

def calc_price(plan: str, duration: str, currency: str = "USD") -> float:
    base = PLAN_PRICES.get(plan, PLAN_PRICES["basic"])
    dur  = PLAN_DURATIONS.get(duration, PLAN_DURATIONS["1month"])
    months_map = {"1month": 1, "3month": 3, "6month": 6}
    months = months_map.get(duration, 1)
    discount = dur["discount"] / 100
    if currency == "EGP":
        base_price = base["egp"] * months
    else:
        base_price = base["usd"] * months
    return round(base_price * (1 - discount), 2)

# ════════════════════════════════════════════
#         PAYMENT SYSTEM — UI BUILDERS
# ════════════════════════════════════════════
def build_invoice_text(order: dict, user_name: str = "") -> str:
    dur_info  = PLAN_DURATIONS.get(order.get("duration", "1month"), PLAN_DURATIONS["1month"])
    plan_info = PLAN_PRICES.get(order.get("plan", "basic"), PLAN_PRICES["basic"])
    days      = dur_info["days"]
    method    = order.get("method", "")

    method_labels = {
        "USDT_TRC20": "💵 USDT TRC-20", "USDT_ERC20": "💵 USDT ERC-20",
        "BTC": "🪙 Bitcoin", "vodafone": "📱 Vodafone Cash",
        "instapay": "💳 InstaPay", "fawry": "🏪 Fawry",
    }

    lines = [
        "🧾 *INVOICE — DATA SCANNER BOT*",
        "━━━━━━━━━━━━━━━━━━━━━━",
        f"📋 Order ID  : `#{order.get('id', '?')}`",
        f"👤 User      : {user_name or order.get('user_id', '?')}",
        f"📦 Plan      : {plan_info['label']}",
        f"⏱ Duration  : {dur_info['label']} ({days} days)",
        f"💳 Method    : {method_labels.get(method, method)}",
        f"💰 Amount    : `{order.get('amount_usd', '?')} USD`",
    ]
    if order.get("amount_egp"):
        lines.append(f"🇪🇬 Amount EGP: `{order.get('amount_egp')} EGP`")
    lines += [
        f"📌 Status    : {'✅ Confirmed' if order.get('status')=='confirmed' else '⏳ Pending' if order.get('status')=='pending' else '❌ Rejected'}",
        f"🕐 Created   : {str(order.get('created_at',''))[:16]}",
    ]
    if order.get("confirmed_at"):
        lines.append(f"✅ Confirmed : {str(order.get('confirmed_at',''))[:16]}")
    if order.get("tx_id"):
        lines.append(f"🔗 TX ID     : `{order.get('tx_id')}`")
    lines += ["━━━━━━━━━━━━━━━━━━━━━━", f"🤖 {WATERMARK_TEXT}"]
    return "\n".join(lines)

def build_crypto_payment_text(plan: str, duration: str, currency_key: str, order_id: int) -> str:
    wallet  = CRYPTO_WALLETS[currency_key]
    amount  = calc_price(plan, duration, "USD")
    dur_lbl = PLAN_DURATIONS[duration]["label"]
    pl_lbl  = PLAN_PRICES[plan]["label"]

    return (
        f"🪙 *Crypto Payment*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 Plan      : {pl_lbl}\n"
        f"⏱ Duration  : {dur_lbl}\n"
        f"💰 Amount    : `{amount} USD`\n"
        f"🌐 Network   : `{wallet['network']}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📋 *Send to this address:*\n"
        f"`{wallet['address']}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📌 Order ID  : `#{order_id}`\n\n"
        f"⚠️ *After sending:*\n"
        f"1️⃣ Copy the TX hash/ID\n"
        f"2️⃣ Send it here as a message\n"
        f"3️⃣ Or send a screenshot of the transaction\n\n"
        f"⏳ Pending orders expire in {PAYMENT_PENDING_TIMEOUT}h"
    )

def build_mobile_payment_text(plan: str, duration: str, method_key: str, order_id: int, currency: str = "EGP") -> str:
    mp     = MOBILE_PAYMENT[method_key]
    amount = calc_price(plan, duration, "EGP")
    dur_lbl = PLAN_DURATIONS[duration]["label"]
    pl_lbl  = PLAN_PRICES[plan]["label"]

    if method_key == "vodafone":
        dest_line = f"📱 *Vodafone Number:*\n`{mp['number']}`\n🏷 Name: `{mp['name']}`"
    elif method_key == "instapay":
        dest_line = f"💳 *InstaPay IPA:*\n`{mp['ipa']}`"
    else:
        dest_line = f"🏪 *Fawry Code:* `{mp['code']}`"

    return (
        f"{mp['label']}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 Plan      : {pl_lbl}\n"
        f"⏱ Duration  : {dur_lbl}\n"
        f"💰 Amount    : `{amount} EGP`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"{dest_line}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📌 Order ID  : `#{order_id}`\n\n"
        f"⚠️ *After sending:*\n"
        f"1️⃣ Take a screenshot of the receipt\n"
        f"2️⃣ Send the screenshot here\n"
        f"3️⃣ Wait for admin confirmation (usually < 1 hour)\n\n"
        f"⏳ Pending orders expire in {PAYMENT_PENDING_TIMEOUT}h"
    )

# ════════════════════════════════════════════
#         PAYMENT SYSTEM — HANDLERS
# ════════════════════════════════════════════
async def show_payment_plans(update, context, query=None):
    uid   = update.effective_user.id
    is_ar = get_lang(uid) == "ar"

    text = (
        f"💳 *{'اختر باقتك' if is_ar else 'Choose Your Plan'}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n\n"
    )

    for pk, pv in PLAN_PRICES.items():
        t   = TIERS.get(pk, TIERS["free"])
        text += (
            f"{pv['label']}\n"
            f"  📊 {'بحوث يومية' if is_ar else 'Daily searches'}: `{t['daily']}`\n"
            f"  📋 {'نتايج' if is_ar else 'Results'}: `{t['max_results']:,}`\n"
            f"  💵 {'من' if is_ar else 'From'} `{pv['usd']}$/mo` | `{pv['egp']} EGP/mo`\n\n"
        )

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("⭐ Basic",   callback_data="pay_plan_basic"),
         InlineKeyboardButton("💎 Premium", callback_data="pay_plan_premium")],
        [InlineKeyboardButton("👑 VIP",     callback_data="pay_plan_vip")],
        [InlineKeyboardButton("🔙 Back", callback_data="user_home")],
    ])

    q = query or (update.callback_query if update.callback_query else None)
    if q:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

async def show_duration_picker(q, plan: str, uid: int):
    is_ar = get_lang(uid) == "ar"
    pl    = PLAN_PRICES[plan]
    text  = (
        f"{pl['label']} — {'اختر المدة' if is_ar else 'Choose Duration'}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n\n"
    )
    rows = []
    for dk, dv in PLAN_DURATIONS.items():
        usd_price = calc_price(plan, dk, "USD")
        egp_price = calc_price(plan, dk, "EGP")
        discount_txt = f" 🔥 -{dv['discount']}%" if dv["discount"] > 0 else ""
        text += f"⏱ {dv['label']}{discount_txt}\n  💵 `{usd_price}$` | `{egp_price} EGP`\n\n"
        rows.append([InlineKeyboardButton(
            f"{dv['label']}{discount_txt}  ({usd_price}$)",
            callback_data=f"pay_dur_{plan}_{dk}"
        )])
    rows.append([InlineKeyboardButton("🔙 Back", callback_data="show_plans")])
    await q.edit_message_text(text, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(rows))

async def show_method_picker(q, plan: str, duration: str, uid: int):
    is_ar = get_lang(uid) == "ar"
    usd_p = calc_price(plan, duration, "USD")
    egp_p = calc_price(plan, duration, "EGP")
    text = (
        f"{'اختر طريقة الدفع' if is_ar else 'Choose Payment Method'}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 Plan: {PLAN_PRICES[plan]['label']}\n"
        f"⏱ Duration: {PLAN_DURATIONS[duration]['label']}\n"
        f"💰 Amount: `{usd_p}$` | `{egp_p} EGP`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"{'اختر طريقة:' if is_ar else 'Select a method:'}\n"
    )
    rows = []
    # Crypto
    for ck, cv in CRYPTO_WALLETS.items():
        rows.append([InlineKeyboardButton(cv["label"], callback_data=f"pay_method_{plan}_{duration}_crypto_{ck}")])
    # Mobile payments
    for mk, mv in MOBILE_PAYMENT.items():
        if mv.get("enabled"):
            rows.append([InlineKeyboardButton(mv["label"], callback_data=f"pay_method_{plan}_{duration}_mobile_{mk}")])
    rows.append([InlineKeyboardButton("🔙 Back", callback_data=f"pay_plan_{plan}")])
    await q.edit_message_text(text, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(rows))

async def handle_payment_method(q, context, plan: str, duration: str, method_type: str, method_key: str, uid: int):
    is_ar = get_lang(uid) == "ar"

    # Check for existing pending order
    existing = get_user_orders(uid, limit=5)
    for oid, oplan, odur, ometh, oamt, ostatus, ocat in existing:
        if ostatus == "pending":
            await q.edit_message_text(
                f"⏳ *{'لديك طلب قيد المراجعة بالفعل' if is_ar else 'You have a pending order'}*\n\n"
                f"📋 Order `#{oid}` — {PLAN_PRICES.get(oplan, {}).get('label', oplan)}\n"
                f"💳 {ometh}\n\n"
                f"{'انتظر تأكيد الأدمن أو تواصل معه.' if is_ar else 'Wait for admin confirmation or contact support.'}",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🧾 View Order", callback_data=f"view_order_{oid}")],
                    [InlineKeyboardButton("❌ Cancel Order", callback_data=f"cancel_order_{oid}")],
                    [InlineKeyboardButton("🔙 Back", callback_data="show_plans")],
                ])
            )
            return

    # Create order
    usd_p = calc_price(plan, duration, "USD")
    egp_p = calc_price(plan, duration, "EGP")

    # Apply coupon if active
    coupon_applied = None
    active_code = context.user_data.get("active_coupon")
    if active_code:
        cv = validate_coupon(active_code, uid, plan)
        if cv["ok"]:
            disc = apply_coupon(cv["coupon"], usd_p, egp_p, plan, duration)
            usd_p = disc["usd"]
            egp_p = disc["egp"]
            coupon_applied = cv["coupon"]
            context.user_data.pop("active_coupon", None)
            log.info(f"Coupon {active_code} applied for user {uid}")

    currency = "USD" if method_type == "crypto" else "EGP"
    order_id = create_payment_order(uid, plan, duration, method_key, usd_p, egp_p, currency)

    if not order_id:
        await q.edit_message_text("❌ Error creating order. Try again.", reply_markup=back_user_kb(uid))
        return

    # Record coupon use
    if coupon_applied:
        record_coupon_use(coupon_applied["id"], uid, order_id)

    # Store pending order in context
    context.user_data["pending_order_id"] = order_id

    if method_type == "crypto":
        text = build_crypto_payment_text(plan, duration, method_key, order_id)
    else:
        text = build_mobile_payment_text(plan, duration, method_key, order_id, "EGP")

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 View Invoice", callback_data=f"view_order_{order_id}")],
        [InlineKeyboardButton("❌ Cancel Order", callback_data=f"cancel_order_{order_id}")],
        [InlineKeyboardButton("🔙 Main Menu",   callback_data="user_home")],
    ])
    await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)

    # Notify admins
    u = get_user_cached(uid)
    uname = u[1] if u else str(uid)
    fname = u[2] if u else "User"
    method_label = CRYPTO_WALLETS.get(method_key, {}).get("label") or MOBILE_PAYMENT.get(method_key, {}).get("label", method_key)
    notif = (
        f"💳 *New Payment Order!*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🔢 Order: `#{order_id}`\n"
        f"👤 User: @{uname} (`{uid}`)\n"
        f"📦 Plan: {PLAN_PRICES[plan]['label']}\n"
        f"⏱ Duration: {PLAN_DURATIONS[duration]['label']}\n"
        f"💳 Method: {method_label}\n"
        f"💰 Amount: `{usd_p}$` | `{egp_p} EGP`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━"
    )
    adm_kb = InlineKeyboardMarkup([[
        InlineKeyboardButton(f"✅ Confirm #{order_id}", callback_data=f"adm_confirm_order_{order_id}"),
        InlineKeyboardButton(f"❌ Reject #{order_id}",  callback_data=f"adm_reject_order_{order_id}"),
    ]])
    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(chat_id=admin_id, text=notif, parse_mode="Markdown", reply_markup=adm_kb)
        except Exception:
            pass

async def show_order_view(q, order_id: int, uid: int):
    order = get_order(order_id)
    if not order or order["user_id"] != uid and not is_admin(uid):
        await q.answer("❌ Not found.", show_alert=True)
        return
    u   = get_user_cached(order["user_id"])
    txt = build_invoice_text(order, u[2] if u else str(order["user_id"]))
    if order["status"] == "pending":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("❌ Cancel Order", callback_data=f"cancel_order_{order_id}")],
            [InlineKeyboardButton("🔙 Back",         callback_data="user_home")],
        ])
    else:
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="user_home")]])
    await q.edit_message_text(txt, parse_mode="Markdown", reply_markup=kb)

async def show_my_orders(q, uid: int):
    orders = get_user_orders(uid, 10)
    is_ar  = get_lang(uid) == "ar"
    if not orders:
        text = f"📋 {'لا توجد طلبات بعد.' if is_ar else 'No orders yet.'}"
        kb   = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="user_home")]])
    else:
        text = f"📋 *{'طلباتي' if is_ar else 'My Orders'}*\n\n"
        rows = []
        for oid, oplan, odur, ometh, oamt, ostatus, ocat in orders:
            emoji = {"confirmed": "✅", "pending": "⏳", "rejected": "❌"}.get(ostatus, "❓")
            text += f"{emoji} `#{oid}` — {PLAN_PRICES.get(oplan,{}).get('label',oplan)} | {ometh} | ${oamt}\n"
            rows.append([InlineKeyboardButton(
                f"{emoji} Order #{oid} — {PLAN_PRICES.get(oplan,{}).get('label',oplan)}",
                callback_data=f"view_order_{oid}"
            )])
        rows.append([InlineKeyboardButton("🔙 Back", callback_data="user_home")])
        kb = InlineKeyboardMarkup(rows)
    await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)

# ════════════════════════════════════════════
#         PAYMENT — ADMIN PANEL
# ════════════════════════════════════════════
async def show_admin_payments(update, context, query=None):
    if not is_admin(update.effective_user.id):
        return
    orders = get_pending_orders(20)
    if not orders:
        text = "✅ *No pending payment orders!*"
        kb   = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="adm_home")]])
    else:
        text = f"💳 *Pending Orders* — {len(orders)}\n\n"
        rows = []
        for r in orders:
            oid, uid2, uname, fname, plan, dur, meth, ausd, aegp, curr, tx, shot, cat = r
            pl_lbl = PLAN_PRICES.get(plan, {}).get("label", plan)
            has_proof = "📎" if (tx or shot) else "❓"
            text += f"`#{oid}` {has_proof} @{uname or uid2} — {pl_lbl} ${ausd}\n"
            rows.append([
                InlineKeyboardButton(f"#{oid} {pl_lbl}", callback_data=f"adm_view_order_{oid}"),
            ])
        rows.append([InlineKeyboardButton("🔙 Back", callback_data="adm_home")])
        kb = InlineKeyboardMarkup(rows)

    q = query or (update.callback_query if update.callback_query else None)
    if q:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

async def show_admin_order_detail(q, order_id: int, context):
    order = get_order(order_id)
    if not order:
        await q.answer("❌ Order not found.", show_alert=True)
        return
    u    = get_user_cached(order["user_id"])
    text = build_invoice_text(order, u[2] if u else str(order["user_id"]))
    if order["status"] == "pending":
        kb = InlineKeyboardMarkup([
            [
                InlineKeyboardButton(f"✅ Confirm",  callback_data=f"adm_confirm_order_{order_id}"),
                InlineKeyboardButton(f"❌ Reject",   callback_data=f"adm_reject_order_{order_id}"),
            ],
            [InlineKeyboardButton("🔙 Back", callback_data="adm_payments")],
        ])
    else:
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="adm_payments")]])

    await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)

    # If has screenshot, send it separately
    if order.get("screenshot_file_id"):
        try:
            await q.message.reply_photo(
                order["screenshot_file_id"],
                caption=f"📎 Payment proof for Order #{order_id}"
            )
        except Exception:
            pass

# ════════════════════════════════════════════
#         PAYMENT — PHOTO HANDLER (PROOFS)
# ════════════════════════════════════════════
async def handle_payment_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid      = update.effective_user.id
    order_id = context.user_data.get("pending_order_id")
    if not order_id:
        # Try to find the latest pending order for this user
        orders = get_user_orders(uid, 3)
        for o in orders:
            if o[5] == "pending":
                order_id = o[0]
                break
    if not order_id:
        return  # Not a payment photo

    file_id = update.message.photo[-1].file_id
    update_order_screenshot(order_id, file_id)

    await update.message.reply_text(
        f"📎 *Payment Proof Received!*\n\n"
        f"📋 Order `#{order_id}`\n\n"
        f"✅ Admin will review and confirm shortly.",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📋 View Order", callback_data=f"view_order_{order_id}")]])
    )

    # Forward proof to admins
    order = get_order(order_id)
    u     = get_user_cached(uid)
    notif = (
        f"📎 *Payment Proof Submitted*\n"
        f"Order `#{order_id}` | @{u[1] if u else uid}\n"
        f"Plan: {PLAN_PRICES.get(order.get('plan',''),{}).get('label','?')}"
    )
    adm_kb = InlineKeyboardMarkup([[
        InlineKeyboardButton(f"✅ Confirm", callback_data=f"adm_confirm_order_{order_id}"),
        InlineKeyboardButton(f"❌ Reject",  callback_data=f"adm_reject_order_{order_id}"),
    ]])
    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_photo(
                chat_id=admin_id,
                photo=file_id,
                caption=notif,
                parse_mode="Markdown",
                reply_markup=adm_kb
            )
        except Exception:
            pass


# ════════════════════════════════════════════
#      GAMIFICATION — BADGES
# ════════════════════════════════════════════
def award_badge(user_id: int, badge_key: str) -> bool:
    """Award a badge if not already held. Returns True if newly awarded."""
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO user_badges (user_id, badge_key, awarded_at) VALUES (%s,%s,%s) "
                    "ON CONFLICT (user_id, badge_key) DO NOTHING",
                    (user_id, badge_key, datetime.utcnow().isoformat())
                )
                newly = cur.rowcount > 0
            conn.commit()
        return newly
    except Exception:
        return False

def get_user_badges(user_id: int) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT badge_key, awarded_at FROM user_badges WHERE user_id=%s ORDER BY awarded_at",
                    (user_id,)
                )
                return cur.fetchall()
    except Exception:
        return []

def check_and_award_badges(user_id: int, context=None) -> list:
    """Check all conditions and award applicable badges. Returns newly awarded list."""
    newly_awarded = []
    try:
        # Search count badge
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM search_logs WHERE user_id=%s", (user_id,))
                search_count = cur.fetchone()[0]
                cur.execute("SELECT referral_count FROM users WHERE user_id=%s", (user_id,))
                row = cur.fetchone()
                ref_count = row[0] if row else 0

        pts  = get_points(user_id)
        tier = get_tier(user_id)

        checks = [
            ("newcomer",   True),
            ("searcher",   search_count >= 10),
            ("hunter",     search_count >= 50),
            ("elite",      search_count >= 200),
            ("referrer",   ref_count >= 3),
            ("vip_member", tier == "vip"),
            ("data_pro",   pts >= 1000),
        ]
        for key, cond in checks:
            if cond and award_badge(user_id, key):
                newly_awarded.append(key)
    except Exception as e:
        log.error(f"check_and_award_badges error: {e}")
    return newly_awarded

# ════════════════════════════════════════════
#      GAMIFICATION — DAILY CHALLENGE
# ════════════════════════════════════════════
def get_daily_challenge_progress(user_id: int) -> dict:
    """Check today's challenge: do 5 searches → earn 10 bonus points."""
    today = datetime.utcnow().strftime("%Y-%m-%d")
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT COUNT(*) FROM search_logs WHERE user_id=%s AND timestamp LIKE %s",
                    (user_id, f"{today}%")
                )
                count = cur.fetchone()[0]
        done    = count >= 5
        return {"count": count, "target": 5, "done": done, "reward": 10}
    except Exception:
        return {"count": 0, "target": 5, "done": False, "reward": 10}

# ════════════════════════════════════════════
#      REAL-TIME STATS (ADMIN)
# ════════════════════════════════════════════
def get_realtime_stats() -> dict:
    try:
        today = datetime.utcnow().strftime("%Y-%m-%d")
        hour_ago = (datetime.utcnow() - timedelta(hours=1)).isoformat()
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM users WHERE joined_at LIKE %s",  (f"{today}%",))
                new_users_today = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM search_logs WHERE timestamp LIKE %s", (f"{today}%",))
                searches_today = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM search_logs WHERE timestamp > %s", (hour_ago,))
                searches_last_hour = cur.fetchone()[0]
                cur.execute(
                    "SELECT COALESCE(SUM(amount_usd),0) FROM payment_orders WHERE status='confirmed' AND created_at LIKE %s",
                    (f"{today}%",)
                )
                revenue_today = float(cur.fetchone()[0])
                cur.execute(
                    "SELECT COUNT(*) FROM payment_orders WHERE status='pending'"
                )
                pending_orders = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM support_tickets WHERE status='open'")
                open_tickets = cur.fetchone()[0]
                cur.execute("SELECT COUNT(DISTINCT user_id) FROM search_logs WHERE timestamp > %s", (hour_ago,))
                active_now = cur.fetchone()[0]
        return {
            "new_users_today":    new_users_today,
            "searches_today":     searches_today,
            "searches_last_hour": searches_last_hour,
            "revenue_today":      revenue_today,
            "pending_orders":     pending_orders,
            "open_tickets":       open_tickets,
            "active_now":         active_now,
        }
    except Exception as e:
        log.error(f"get_realtime_stats error: {e}")
        return {}

def get_weekly_stats() -> list:
    """Last 7 days stats."""
    rows = []
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                for i in range(6, -1, -1):
                    d = (datetime.utcnow() - timedelta(days=i)).strftime("%Y-%m-%d")
                    cur.execute("SELECT COUNT(*) FROM users WHERE joined_at LIKE %s",  (f"{d}%",))
                    nu = cur.fetchone()[0]
                    cur.execute("SELECT COUNT(*) FROM search_logs WHERE timestamp LIKE %s", (f"{d}%",))
                    ns = cur.fetchone()[0]
                    cur.execute(
                        "SELECT COALESCE(SUM(amount_usd),0) FROM payment_orders WHERE status='confirmed' AND created_at LIKE %s",
                        (f"{d}%",)
                    )
                    rev = float(cur.fetchone()[0])
                    rows.append({"date": d, "users": nu, "searches": ns, "revenue": rev})
    except Exception:
        pass
    return rows

def get_top_searches(limit=10) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT keyword, COUNT(*) as cnt FROM search_logs GROUP BY keyword ORDER BY cnt DESC LIMIT %s",
                    (limit,)
                )
                return cur.fetchall()
    except Exception:
        return []

def get_top_users_by_searches(limit=10) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT sl.user_id, u.full_name, COUNT(*) as cnt
                    FROM search_logs sl LEFT JOIN users u ON u.user_id=sl.user_id
                    GROUP BY sl.user_id, u.full_name ORDER BY cnt DESC LIMIT %s
                """, (limit,))
                return cur.fetchall()
    except Exception:
        return []

# ════════════════════════════════════════════
#      DEDUPLICATION
# ════════════════════════════════════════════
def run_deduplication() -> int:
    """Remove exact duplicate lines from data_index. Returns count removed."""
    removed = 0
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    DELETE FROM data_index
                    WHERE id NOT IN (
                        SELECT MIN(id) FROM data_index GROUP BY line
                    )
                """)
                removed = cur.rowcount
            conn.commit()
        log.info(f"Deduplication removed {removed:,} duplicate rows")
        # Invalidate count cache
        _count_cache.clear()
    except Exception as e:
        log.error(f"run_deduplication error: {e}")
    return removed

def run_nameid_deduplication() -> int:
    removed = 0
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    DELETE FROM name_id_index
                    WHERE id NOT IN (
                        SELECT MIN(id) FROM name_id_index
                        GROUP BY full_name, national_id
                    )
                """)
                removed = cur.rowcount
            conn.commit()
        _count_cache.clear()
    except Exception as e:
        log.error(f"run_nameid_dedup error: {e}")
    return removed

# ════════════════════════════════════════════
#      SCHEDULED IMPORTS
# ════════════════════════════════════════════
def get_scheduled_imports() -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, url, label, stype, frequency_hours, last_run, next_run, status, records_added "
                    "FROM scheduled_imports ORDER BY created_at DESC"
                )
                return cur.fetchall()
    except Exception:
        return []

def create_scheduled_import(url: str, label: str, stype: str, freq_hours: int, admin_id: int) -> int:
    try:
        now = datetime.utcnow().isoformat()
        nxt = (datetime.utcnow() + timedelta(hours=freq_hours)).isoformat()
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO scheduled_imports (url, label, stype, frequency_hours, next_run, status, created_by, created_at)
                    VALUES (%s,%s,%s,%s,%s,'active',%s,%s) RETURNING id
                """, (url, label, stype, freq_hours, nxt, admin_id, now))
                sid = cur.fetchone()[0]
            conn.commit()
        return sid
    except Exception as e:
        log.error(f"create_scheduled_import: {e}")
        return 0

def delete_scheduled_import(sid: int):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM scheduled_imports WHERE id=%s", (sid,))
            conn.commit()
    except Exception:
        pass

# ════════════════════════════════════════════
#      ADMIN — STATS DASHBOARD HANDLER
# ════════════════════════════════════════════
async def show_admin_stats(update, context, query=None):
    uid = update.effective_user.id
    if not is_admin(uid):
        return

    rt  = get_realtime_stats()
    wk  = get_weekly_stats()
    top = get_top_searches(5)

    # Build weekly bar chart (ASCII)
    chart = ""
    max_s = max((d["searches"] for d in wk), default=1) or 1
    for d in wk:
        bar_len = int(d["searches"] / max_s * 12)
        bar = "█" * bar_len + "░" * (12 - bar_len)
        day = d["date"][5:]  # MM-DD
        chart += f"`{day}` {bar} {d['searches']}\n"

    text = (
        f"📊 *Real-Time Statistics*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👥 New users today   : `{rt.get('new_users_today', 0)}`\n"
        f"🔍 Searches today    : `{rt.get('searches_today', 0):,}`\n"
        f"⚡ Searches/last hr  : `{rt.get('searches_last_hour', 0)}`\n"
        f"👤 Active now (1h)   : `{rt.get('active_now', 0)}`\n"
        f"💰 Revenue today     : `${rt.get('revenue_today', 0):.2f}`\n"
        f"💳 Pending orders    : `{rt.get('pending_orders', 0)}`\n"
        f"🎫 Open tickets      : `{rt.get('open_tickets', 0)}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📅 *Searches — Last 7 Days*\n"
        f"{chart}"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🔥 *Top Searches*\n"
    )
    for i, (kw, cnt) in enumerate(top, 1):
        text += f"`{i}.` `{kw[:25]}` — {cnt:,}x\n"

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("👥 Top Users",        callback_data="adm_top_users"),
         InlineKeyboardButton("📈 Weekly Report",    callback_data="adm_weekly_report")],
        [InlineKeyboardButton("🗑️ Run Dedup",        callback_data="adm_run_dedup"),
         InlineKeyboardButton("⏰ Scheduled Imports", callback_data="adm_scheduled")],
        [InlineKeyboardButton("🔙 Back", callback_data="adm_home")],
    ])

    q = query or (update.callback_query if update.callback_query else None)
    if q:
        try:
            await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
        except BadRequest:
            await q.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

async def show_weekly_report(q, uid: int):
    wk   = get_weekly_stats()
    text = "📈 *Weekly Report*\n━━━━━━━━━━━━━━━━━━━━━━\n"
    text += f"{'Date':<12} {'Users':>6} {'Searches':>9} {'Revenue':>9}\n"
    text += "─" * 40 + "\n"
    total_u = total_s = total_r = 0
    for d in wk:
        text += f"`{d['date'][5:]}` {d['users']:>6} {d['searches']:>9,} ${d['revenue']:>7.2f}\n"
        total_u += d["users"]
        total_s += d["searches"]
        total_r += d["revenue"]
    text += "─" * 40 + "\n"
    text += f"`Total  ` {total_u:>6} {total_s:>9,} ${total_r:>7.2f}\n"
    text += f"━━━━━━━━━━━━━━━━━━━━━━"
    kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="adm_stats")]])
    await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)

async def show_top_users(q, uid: int):
    top  = get_top_users_by_searches(15)
    text = "👥 *Top Users by Searches*\n━━━━━━━━━━━━━━━━━━━━━━\n"
    medals = ["🥇","🥈","🥉"] + [f"{i}." for i in range(4, 16)]
    for i, (lid, lname, cnt) in enumerate(top):
        text += f"{medals[i]} `{lid}` {lname or 'User'[:15]} — {cnt:,}\n"
    kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="adm_stats")]])
    await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)

async def show_scheduled_imports(q, uid: int):
    imports = get_scheduled_imports()
    if not imports:
        text = "⏰ *No scheduled imports yet.*\n\nUse the button to add one."
    else:
        text = f"⏰ *Scheduled Imports* — {len(imports)}\n\n"
        for sid, url, label, stype, freq, last, nxt, status, added in imports:
            emoji = "✅" if status == "active" else "⏸️"
            text += f"{emoji} `#{sid}` {label or url[:25]}\n  Every {freq}h | +{added:,} records\n"
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Add Import", callback_data="adm_add_scheduled")],
        [InlineKeyboardButton("🔙 Back", callback_data="adm_stats")],
    ])
    await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)

# ════════════════════════════════════════════
#      USER — GAMIFICATION SCREEN
# ════════════════════════════════════════════
async def show_gamification(update, context, query=None):
    uid    = update.effective_user.id
    is_ar  = get_lang(uid) == "ar"
    badges = get_user_badges(uid)
    pts    = get_points(uid)
    ch     = get_daily_challenge_progress(uid)
    newly  = check_and_award_badges(uid)

    # Badges display
    owned_keys = {b[0] for b in badges}
    badge_text = ""
    for key, info in BADGES.items():
        if key in owned_keys:
            badge_text += f"{info['label']} ✅\n"
        else:
            badge_text += f"{'🔒 ' + info['label']} — {info['desc']}\n"

    # Daily challenge bar
    prog    = min(ch["count"], ch["target"])
    bar_len = int(prog / ch["target"] * 10)
    bar     = "█" * bar_len + "░" * (10 - bar_len)

    todays_challenge = "Today's Challenge"
    text = (
        f"🎮 *{'الإنجازات والتحديات' if is_ar else 'Achievements & Challenges'}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"⭐ {'نقاطك' if is_ar else 'Your Points'}: `{pts:,}`\n\n"
        f"📅 *{'تحدي اليوم' if is_ar else todays_challenge}*\n"
        f"`{bar}` {prog}/{ch['target']} {'بحوث' if is_ar else 'searches'}\n"
        f"{'✅ مكتمل! +' + str(ch['reward']) + ' نقطة' if ch['done'] else '🎁 أكمل 5 بحوث → +' + str(ch['reward']) + ' نقطة'}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🏅 *{'شاراتي' if is_ar else 'My Badges'}* — {len(owned_keys)}/{len(BADGES)}\n\n"
        f"{badge_text}"
    )
    if newly:
        new_labels = " | ".join(BADGES[k]["label"] for k in newly if k in BADGES)
        text += f"\n🎉 *New badge{'s' if len(newly)>1 else ''}!* {new_labels}"

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🏆 Leaderboard", callback_data="my_points"),
         InlineKeyboardButton("🔙 Back",       callback_data="user_home")],
    ])
    q = query or (update.callback_query if update.callback_query else None)
    if q:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

# ════════════════════════════════════════════
#      ONBOARDING — NEW USER TUTORIAL
# ════════════════════════════════════════════
ONBOARDING_STEPS = [
    {
        "title":   "👋 Welcome to Data Scanner Bot!",
        "content": (
            "This bot lets you search through millions of leaked data records.\n\n"
            "🔍 *What you can search:*\n"
            "• Emails, passwords, usernames\n"
            "• Phone numbers, domains, URLs\n"
            "• Names & National IDs\n"
            "• IP addresses, IBANs, addresses\n\n"
            "Let's set you up in 3 quick steps!"
        ),
    },
    {
        "title":   "📋 Step 1 — Choose Your Language",
        "content": "Tap the button below to select your preferred language.\n\nYou can also change it anytime from the main menu.",
    },
    {
        "title":   "🔍 Step 2 — How to Search",
        "content": (
            "Tap *🔍 Search* from the main menu.\n\n"
            "Then choose a category:\n"
            "• `email` — search by email address\n"
            "• `domain` — search by website domain\n"
            "• `phone` — search by phone number\n"
            "• `all` — search everything\n\n"
            "Type your keyword and wait for results!"
        ),
    },
    {
        "title":   "💎 Step 3 — Upgrade Your Plan",
        "content": (
            "Free plan = limited searches.\n\n"
            "⭐ Basic | 💎 Premium | 👑 VIP\n\n"
            "Tap *💳 Plans* to see pricing and upgrade.\n\n"
            "You can pay with crypto or Vodafone Cash."
        ),
    },
]

async def send_onboarding(bot, user_id: int, step: int = 0):
    if step >= len(ONBOARDING_STEPS):
        return
    s  = ONBOARDING_STEPS[step]
    kb_rows = []
    if step < len(ONBOARDING_STEPS) - 1:
        kb_rows.append([InlineKeyboardButton(f"Next ➡️ ({step+2}/{len(ONBOARDING_STEPS)})", callback_data=f"onboard_{step+1}")])
    kb_rows.append([InlineKeyboardButton("⏩ Skip Tutorial", callback_data="onboard_done")])
    text = f"*{s['title']}*\n━━━━━━━━━━━━━━━━━━━━━━\n{s['content']}"
    try:
        await bot.send_message(chat_id=user_id, text=text, parse_mode="Markdown",
                               reply_markup=InlineKeyboardMarkup(kb_rows))
    except Exception:
        pass

# ════════════════════════════════════════════
#      AUTO DAILY STATS TRACKING
# ════════════════════════════════════════════
def record_daily_stats():
    today = datetime.utcnow().strftime("%Y-%m-%d")
    try:
        rt = get_realtime_stats()
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO daily_stats (stat_date, new_users, total_searches, revenue_usd, active_users)
                    VALUES (%s,%s,%s,%s,%s)
                    ON CONFLICT (stat_date) DO UPDATE SET
                        new_users=EXCLUDED.new_users,
                        total_searches=EXCLUDED.total_searches,
                        revenue_usd=EXCLUDED.revenue_usd,
                        active_users=EXCLUDED.active_users
                """, (today, rt.get("new_users_today",0), rt.get("searches_today",0),
                      rt.get("revenue_today",0), rt.get("active_now",0)))
            conn.commit()
    except Exception as e:
        log.error(f"record_daily_stats: {e}")

# ════════════════════════════════════════════
#      AUTO DAILY REPORT TO ADMIN
# ════════════════════════════════════════════
async def send_daily_report(app):
    rt  = get_realtime_stats()
    wk  = get_weekly_stats()
    counts = get_cached_counts()
    today = datetime.utcnow().strftime("%Y-%m-%d")
    text = (
        f"📊 *Daily Report — {today}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👥 New users    : `{rt.get('new_users_today', 0)}`\n"
        f"🔍 Searches     : `{rt.get('searches_today', 0):,}`\n"
        f"💰 Revenue      : `${rt.get('revenue_today', 0):.2f}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 Total DB     : `{counts.get('data', 0):,}`\n"
        f"👤 Total Users  : `{counts.get('users', 0):,}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"7-day searches  : `{sum(d['searches'] for d in wk):,}`\n"
        f"7-day revenue   : `${sum(d['revenue'] for d in wk):.2f}`\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🤖 {WATERMARK_TEXT}"
    )
    await notify_admins(app, text)


# ════════════════════════════════════════════
#   REGEX SEARCH
# ════════════════════════════════════════════
def search_by_regex(pattern: str, limit: int = 200) -> list:
    """Search data_index using a Python regex pattern."""
    results = []
    try:
        compiled = re.compile(pattern, re.IGNORECASE)
    except re.error as e:
        return {"error": str(e)}
    try:
        # Fetch in batches and filter client-side (PG regex too)
        with pool_conn() as conn:
            with conn.cursor() as cur:
                # Use PostgreSQL regex for efficiency, then validate with Python
                cur.execute(
                    "SELECT line FROM data_index WHERE line ~ %s LIMIT %s",
                    (pattern, limit * 2)
                )
                rows = cur.fetchall()
        for (line,) in rows:
            if compiled.search(line):
                parsed = parse_line(line)
                results.append(parsed)
                if len(results) >= limit:
                    break
    except Exception as e:
        log.error(f"search_by_regex error: {e}")
        return []
    return results

# ════════════════════════════════════════════
#   DATA CATEGORIES & TAGS
# ════════════════════════════════════════════
def classify_line(line: str) -> str:
    """Auto-classify a data line into a category."""
    for cat_key, cat in DATA_CATEGORIES.items():
        if cat["pattern"] and re.search(cat["pattern"], line):
            return cat_key
    return "other"

def filter_results_by_category(results: list, category: str) -> list:
    """Filter a result list to only those matching a category."""
    if not category or category == "all":
        return results
    cat = DATA_CATEGORIES.get(category)
    if not cat or not cat["pattern"]:
        return results
    compiled = re.compile(cat["pattern"], re.IGNORECASE)
    return [r for r in results if compiled.search(r.get("line", "") or str(r))]

def get_category_stats(results: list) -> dict:
    """Count how many results fall into each category."""
    counts = {k: 0 for k in DATA_CATEGORIES}
    for r in results:
        line = r.get("line", "") or ""
        counts[classify_line(line)] = counts.get(classify_line(line), 0) + 1
    return counts

def add_file_tags(file_id: int, tags: list):
    """Add tag labels to an uploaded file."""
    try:
        tag_str = ",".join(tags)
        with pool_conn() as conn:
            with conn.cursor() as cur:
                # Add tags column if not exists
                cur.execute("ALTER TABLE uploaded_files ADD COLUMN IF NOT EXISTS tags TEXT DEFAULT ''")
                cur.execute("UPDATE uploaded_files SET tags=%s WHERE id=%s", (tag_str, file_id))
            conn.commit()
    except Exception:
        pass

def search_files_by_tag(tag: str) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, original_name, records, uploaded_at, tags FROM uploaded_files WHERE tags ILIKE %s",
                    (f"%{tag}%",)
                )
                return cur.fetchall()
    except Exception:
        return []

# ════════════════════════════════════════════
#   2FA — TWO FACTOR AUTHENTICATION
# ════════════════════════════════════════════
_2fa_codes: dict = {}   # {user_id: {"code": "123456", "ts": timestamp, "purpose": "login"}}

def generate_2fa_code(user_id: int, purpose: str = "login") -> str:
    import random
    code = "".join([str(random.randint(0, 9)) for _ in range(TWO_FA_CODE_LENGTH)])
    _2fa_codes[user_id] = {"code": code, "ts": time.time(), "purpose": purpose}
    return code

def verify_2fa_code(user_id: int, code: str) -> bool:
    entry = _2fa_codes.get(user_id)
    if not entry:
        return False
    if time.time() - entry["ts"] > TWO_FA_CODE_EXPIRY:
        _2fa_codes.pop(user_id, None)
        return False
    if entry["code"] == code.strip():
        _2fa_codes.pop(user_id, None)
        return True
    return False

def is_2fa_enabled(user_id: int) -> bool:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM users WHERE user_id=%s AND updated_at='2fa_on'", (user_id,))
                # Using updated_at as a flag — in production add a real column
                return False  # Default off; toggle via settings
    except Exception:
        return False

def set_2fa_enabled(user_id: int, enabled: bool):
    # Stores flag — in full deploy, add a dedicated column
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "ALTER TABLE users ADD COLUMN IF NOT EXISTS twofa_enabled INTEGER DEFAULT 0"
                )
                cur.execute(
                    "UPDATE users SET twofa_enabled=%s WHERE user_id=%s",
                    (1 if enabled else 0, user_id)
                )
            conn.commit()
        invalidate_user_cache(user_id)
    except Exception:
        pass

async def send_2fa_code(bot, user_id: int, purpose: str = "sensitive action") -> str:
    code = generate_2fa_code(user_id, purpose)
    try:
        await bot.send_message(
            chat_id=user_id,
            text=(
                f"🔐 *2FA Verification Code*\n\n"
                f"Your one-time code for *{purpose}*:\n\n"
                f"```{code}```\n\n"
                f"⏳ Expires in {TWO_FA_CODE_EXPIRY // 60} minutes.\n"
                f"Do NOT share this code with anyone."
            ),
            parse_mode="Markdown"
        )
    except Exception:
        pass
    return code

# ════════════════════════════════════════════
#   URL IMPORT (Direct Link Import)
# ════════════════════════════════════════════
async def import_from_url(url: str, label: str, admin_id: int, bot) -> dict:
    """Download a file from URL and index it."""
    import urllib.request
    import tempfile

    result = {"ok": False, "records": 0, "msg": ""}
    try:
        # Check extension
        url_lower = url.lower().split("?")[0]
        if not any(url_lower.endswith(ext) for ext in URL_IMPORT_ALLOWED):
            result["msg"] = f"Unsupported file type. Allowed: {', '.join(URL_IMPORT_ALLOWED)}"
            return result

        await bot.send_message(chat_id=admin_id, text=f"⏳ Downloading `{url}`...", parse_mode="Markdown")

        # Download to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(url_lower)[-1]) as tmp:
            tmp_path = tmp.name

        def _download():
            urllib.request.urlretrieve(url, tmp_path)
        await asyncio.get_running_loop().run_in_executor(_executor, _download)

        size_mb = os.path.getsize(tmp_path) / 1024 / 1024
        if size_mb > URL_IMPORT_MAX_MB:
            os.unlink(tmp_path)
            result["msg"] = f"File too large: {size_mb:.1f}MB (max {URL_IMPORT_MAX_MB}MB)"
            return result

        # Index the file using existing process_file
        dest = os.path.join(FILES_DIR, f"url_import_{int(time.time())}_{label[:20]}{os.path.splitext(tmp_path)[-1]}")
        shutil.move(tmp_path, dest)

        records = await asyncio.get_running_loop().run_in_executor(
            _executor, lambda: process_file(dest, label)
        )
        result["ok"]      = True
        result["records"] = records
        result["msg"]     = f"✅ Imported {records:,} records from URL"

        # Update scheduled import record if exists
        try:
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "UPDATE scheduled_imports SET last_run=%s, records_added=records_added+%s "
                        "WHERE url=%s",
                        (datetime.utcnow().isoformat(), records, url)
                    )
                conn.commit()
        except Exception:
            pass

    except Exception as e:
        result["msg"] = f"❌ Error: {e}"
        log.error(f"import_from_url error: {e}")
    return result

# ════════════════════════════════════════════
#   SCHEDULED IMPORTS — EXECUTION ENGINE
# ════════════════════════════════════════════
async def run_scheduled_imports(app):
    """Run all due scheduled imports."""
    imports = get_scheduled_imports()
    now_iso = datetime.utcnow().isoformat()
    for row in imports:
        sid, url, label, stype, freq, last_run, next_run, status, records_added = row
        if status != "active":
            continue
        if next_run and now_iso < next_run:
            continue  # Not due yet
        try:
            log.info(f"⏰ Running scheduled import #{sid}: {url}")
            admin_id = ADMIN_IDS[0] if ADMIN_IDS else None
            result   = await import_from_url(url, label or f"sched_{sid}", admin_id, app.bot)
            # Update next_run
            nxt = (datetime.utcnow() + timedelta(hours=freq)).isoformat()
            with pool_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "UPDATE scheduled_imports SET last_run=%s, next_run=%s, records_added=records_added+%s WHERE id=%s",
                        (now_iso, nxt, result.get("records", 0), sid)
                    )
                conn.commit()
            if admin_id:
                await app.bot.send_message(
                    chat_id=admin_id,
                    text=f"⏰ *Scheduled Import #{sid}*\n{result['msg']}\n🔗 {url}",
                    parse_mode="Markdown"
                )
        except Exception as e:
            log.error(f"Scheduled import #{sid} failed: {e}")

# ════════════════════════════════════════════
#   ADVANCED RESULT FILTER (UI)
# ════════════════════════════════════════════
async def show_filter_menu(q, keyword: str, stype: str, uid: int, context):
    """Show category filter options after search."""
    is_ar = get_lang(uid) == "ar"
    text = (
        f"🔽 *{'فلترة النتايج' if is_ar else 'Filter Results'}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🎯 Keyword: `{mesc(keyword)}`\n\n"
        f"{'اختر فئة لتصفية النتايج:' if is_ar else 'Choose a category to filter by:'}"
    )
    rows = [[InlineKeyboardButton("📋 All (no filter)", callback_data=_cb_put(f"filter_search:all:{keyword}:{stype}"))]]
    for cat_key, cat in DATA_CATEGORIES.items():
        rows.append([InlineKeyboardButton(
            cat["label"],
            callback_data=_cb_put(f"filter_search:{cat_key}:{keyword}:{stype}")
        )])
    rows.append([InlineKeyboardButton("🔙 Back", callback_data="user_home")])
    await q.edit_message_text(text, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(rows))

# ════════════════════════════════════════════
#   2FA SETTINGS SCREEN (User)
# ════════════════════════════════════════════
async def show_security_settings(update, context, query=None):
    uid   = update.effective_user.id
    is_ar = get_lang(uid) == "ar"
    text = (
        f"🔐 *{'الأمان' if is_ar else 'Security Settings'}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"🔑 *2FA (Two-Factor Authentication)*\n"
        f"{'أضف طبقة حماية إضافية على حسابك.' if is_ar else 'Add an extra layer of protection to your account.'}\n\n"
        f"⚠️ {'عند تفعيله، ستحتاج كود تأكيد لكل عملية حساسة.' if is_ar else 'When enabled, sensitive actions require a verification code.'}"
    )
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Enable 2FA",  callback_data="2fa_enable"),
         InlineKeyboardButton("❌ Disable 2FA", callback_data="2fa_disable")],
        [InlineKeyboardButton("🧪 Test 2FA",    callback_data="2fa_test")],
        [InlineKeyboardButton("🔙 Back",        callback_data="user_home")],
    ])
    q = query or (update.callback_query if update.callback_query else None)
    if q:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

# ════════════════════════════════════════════
#   ADMIN: URL IMPORT HANDLER
# ════════════════════════════════════════════
async def show_url_import_menu(q, uid: int, context):
    await q.edit_message_text(
        "🔗 *URL Import*\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Send a direct URL to a data file:\n"
        f"Supported: `{', '.join(URL_IMPORT_ALLOWED)}`\n"
        f"Max size: `{URL_IMPORT_MAX_MB}MB`\n\n"
        "Format: `URL Optional_Label`\n"
        "Example:\n`https://example.com/data.txt MyData`",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Cancel", callback_data="adm_home")]])
    )
    context.user_data["admin_action"] = "url_import"


# ════════════════════════════════════════════
#   USER ACTIVITY LOG
# ════════════════════════════════════════════
def log_activity(user_id: int, action: str, details: str = ""):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO user_activity_log (user_id, action, details, timestamp) VALUES (%s,%s,%s,%s)",
                    (user_id, action, details, datetime.utcnow().isoformat())
                )
            conn.commit()
    except Exception:
        pass

def get_user_activity(user_id: int, limit: int = 30) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT action, details, timestamp FROM user_activity_log "
                    "WHERE user_id=%s ORDER BY timestamp DESC LIMIT %s",
                    (user_id, limit)
                )
                return cur.fetchall()
    except Exception:
        return []

# ════════════════════════════════════════════
#   CALLBACKS — WIRE ALL REMAINING FEATURES
# ════════════════════════════════════════════

# Patch log_search to also write activity log
_orig_log_search = None

def log_search_with_activity(uid: int, keyword: str, stype: str, results: int):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO search_logs (user_id, keyword, category, results, timestamp) VALUES (%s,%s,%s,%s,%s)",
                    (uid, keyword, stype, results, datetime.utcnow().isoformat())
                )
            conn.commit()
    except Exception:
        pass
    log_activity(uid, "search", f"kw={keyword} type={stype} results={results}")

# ════════════════════════════════════════════
#   ADMIN: USER ACTIVITY LOG VIEWER
# ════════════════════════════════════════════
async def show_user_activity(q, target_uid: int, admin_uid: int):
    rows = get_user_activity(target_uid, 20)
    if not rows:
        text = f"📋 No activity log for user `{target_uid}`."
    else:
        text = f"📋 *Activity Log — `{target_uid}`*\n\n"
        for action, details, ts in rows:
            ts_str = str(ts)[:16]
            text += f"`{ts_str}` — {action}"
            if details:
                text += f": {details[:40]}"
            text += "\n"
    kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="adm_home")]])
    try:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    except Exception:
        await q.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

# ════════════════════════════════════════════
#   REGEX SEARCH HANDLER
# ════════════════════════════════════════════
async def handle_regex_search(update, context, pattern: str):
    uid = update.effective_user.id
    if not can_search(uid):
        await update.message.reply_text(
            "❌ No searches remaining.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("💳 Plans", callback_data="show_plans")]])
        )
        return
    msg = await update.message.reply_text(
        f"🔍 *Regex Search...*\n`{mesc(pattern)}`", parse_mode="Markdown"
    )
    results = await asyncio.get_running_loop().run_in_executor(
        _executor, lambda: search_by_regex(pattern, 500)
    )
    if isinstance(results, dict) and "error" in results:
        await msg.edit_text(f"❌ Invalid regex: `{mesc(results['error'])}`", parse_mode="Markdown",
                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Menu", callback_data="user_home")]]))
        return
    if not results:
        await msg.edit_text("🔍 *No results found.*", parse_mode="Markdown",
                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Menu", callback_data="user_home")]]))
        return

    content  = build_result_txt(f"regex:{pattern}", results, "regex")
    content  = add_watermark(content, uid)
    safe_pat = re.sub(r"[^\w\-]", "_", pattern)[:20]
    filename = f"regex_{safe_pat}_{len(results)}_results.txt"
    tmppath  = os.path.join(FILES_DIR, f"tmp_regex_{uid}.txt")
    with open(tmppath, "w", encoding="utf-8") as f:
        f.write(content)
    caption = (
        f"✅ *Regex Search Results*\n"
        f"🔍 Pattern: `{mesc(pattern)}`\n"
        f"📊 Total: `{len(results):,}` records"
    )
    await msg.delete()
    await safe_send_document(
        update.message.reply_document, tmppath, filename, caption,
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Menu", callback_data="user_home")]])
    )
    if not is_admin(uid):
        deduct(uid)
        add_points(uid, POINTS_PER_SEARCH, "regex_search")
    log_search(uid, f"regex:{pattern}", "regex", len(results))
    log_activity(uid, "regex_search", pattern[:60])

# ════════════════════════════════════════════
#   SCHEDULED IMPORTS — HOURLY CHECKER
# ════════════════════════════════════════════
async def scheduled_import_loop(app):
    """Background loop that checks and runs scheduled imports every hour."""
    while True:
        await asyncio.sleep(SCHEDULED_IMPORT_INTERVAL)
        try:
            await run_scheduled_imports(app)
        except Exception as e:
            log.error(f"scheduled_import_loop error: {e}")


# ════════════════════════════════════════════
#   COUPON SYSTEM
# ════════════════════════════════════════════
def create_coupon(code: str, ctype: str, value: float, max_uses: int,
                  valid_days: int, plan_restriction: str, admin_id: int) -> int:
    try:
        now      = datetime.utcnow().isoformat()
        valid_until = (datetime.utcnow() + timedelta(days=valid_days)).isoformat() if valid_days else None
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO coupons (code, type, value, max_uses, valid_from, valid_until,
                    plan_restriction, created_by, created_at, is_active)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,1) RETURNING id
                """, (code.upper(), ctype, value, max_uses, now, valid_until,
                      plan_restriction, admin_id, now))
                cid = cur.fetchone()[0]
            conn.commit()
        return cid
    except Exception as e:
        log.error(f"create_coupon: {e}")
        return 0

def validate_coupon(code: str, user_id: int, plan: str = None) -> dict:
    """Validate a coupon. Returns {ok, coupon, msg}."""
    try:
        now = datetime.utcnow().isoformat()
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM coupons WHERE code=%s AND is_active=1", (code.upper(),))
                row = cur.fetchone()
        if not row:
            return {"ok": False, "msg": "❌ Invalid coupon code."}
        cols = ["id","code","type","value","max_uses","used_count","valid_from",
                "valid_until","plan_restriction","created_by","created_at","is_active"]
        cp = dict(zip(cols, row))

        if cp["used_count"] >= cp["max_uses"]:
            return {"ok": False, "msg": "❌ This coupon has been fully used."}
        if cp["valid_until"] and now > cp["valid_until"]:
            return {"ok": False, "msg": "❌ This coupon has expired."}
        if cp["valid_from"] and now < cp["valid_from"]:
            return {"ok": False, "msg": "❌ This coupon is not active yet."}
        if cp["plan_restriction"] and plan and cp["plan_restriction"] != plan:
            return {"ok": False, "msg": f"❌ This coupon is only for *{cp['plan_restriction']}* plan."}

        # Check if user already used it
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM coupon_uses WHERE coupon_id=%s AND user_id=%s",
                            (cp["id"], user_id))
                already = cur.fetchone()
        if already:
            return {"ok": False, "msg": "❌ You already used this coupon."}

        # Calculate discount text
        if cp["type"] == "percent":
            desc = f"🏷 {cp['value']}% off"
        elif cp["type"] == "fixed":
            desc = f"🏷 ${cp['value']} off"
        elif cp["type"] == "days":
            desc = f"🏷 +{int(cp['value'])} free days"
        else:
            desc = f"🏷 Free {cp['plan_restriction'] or ''} plan"

        return {"ok": True, "coupon": cp, "msg": f"✅ Coupon valid! {desc}"}
    except Exception as e:
        log.error(f"validate_coupon: {e}")
        return {"ok": False, "msg": "❌ Error validating coupon."}

def apply_coupon(coupon: dict, base_price_usd: float, base_price_egp: float,
                 plan: str, duration: str) -> dict:
    """Apply coupon discount. Returns adjusted prices and extra_days."""
    cp = coupon
    extra_days = 0
    new_usd    = base_price_usd
    new_egp    = base_price_egp
    new_plan   = plan

    if cp["type"] == "percent":
        disc = cp["value"] / 100
        new_usd = round(base_price_usd * (1 - disc), 2)
        new_egp = round(base_price_egp * (1 - disc), 2)
    elif cp["type"] == "fixed":
        new_usd = max(0, round(base_price_usd - cp["value"], 2))
        new_egp = max(0, round(base_price_egp - cp["value"] * 50, 2))  # approx EGP
    elif cp["type"] == "days":
        extra_days = int(cp["value"])
    elif cp["type"] == "plan":
        new_plan = cp["plan_restriction"] or plan

    return {
        "usd": new_usd, "egp": new_egp,
        "extra_days": extra_days, "plan": new_plan,
        "original_usd": base_price_usd,
    }

def record_coupon_use(coupon_id: int, user_id: int, order_id: int = 0):
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE coupons SET used_count=used_count+1 WHERE id=%s", (coupon_id,))
                cur.execute(
                    "INSERT INTO coupon_uses (coupon_id, user_id, order_id, used_at) VALUES (%s,%s,%s,%s)",
                    (coupon_id, user_id, order_id, datetime.utcnow().isoformat())
                )
            conn.commit()
    except Exception as e:
        log.error(f"record_coupon_use: {e}")

def list_coupons(active_only: bool = True) -> list:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                q = "SELECT id, code, type, value, max_uses, used_count, valid_until, plan_restriction, is_active FROM coupons"
                if active_only:
                    q += " WHERE is_active=1"
                q += " ORDER BY created_at DESC"
                cur.execute(q)
                return cur.fetchall()
    except Exception:
        return []

def delete_coupon(code: str) -> bool:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE coupons SET is_active=0 WHERE code=%s", (code.upper(),))
            conn.commit()
        return True
    except Exception:
        return False

# ════════════════════════════════════════════
#   AUTO-RENEWAL SYSTEM
# ════════════════════════════════════════════
def set_auto_renewal(user_id: int, enabled: bool, method: str = None,
                     plan: str = None, duration: str = None):
    try:
        now = datetime.utcnow().isoformat()
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO auto_renewal (user_id, enabled, preferred_method, preferred_plan,
                    preferred_duration, created_at)
                    VALUES (%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (user_id) DO UPDATE SET
                    enabled=%s, preferred_method=%s, preferred_plan=%s, preferred_duration=%s
                """, (user_id, int(enabled), method, plan, duration, now,
                      int(enabled), method, plan, duration))
            conn.commit()
        invalidate_user_cache(user_id)
    except Exception as e:
        log.error(f"set_auto_renewal: {e}")

def get_auto_renewal(user_id: int) -> dict:
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT enabled, preferred_method, preferred_plan, preferred_duration "
                    "FROM auto_renewal WHERE user_id=%s", (user_id,)
                )
                row = cur.fetchone()
        if not row:
            return {"enabled": False}
        return {"enabled": bool(row[0]), "method": row[1], "plan": row[2], "duration": row[3]}
    except Exception:
        return {"enabled": False}

async def check_auto_renewals(app):
    """Check users nearing expiry and send renewal reminders."""
    alert_date = (datetime.utcnow() + timedelta(days=AUTO_RENEW_DAYS_BEFORE)).strftime("%Y-%m-%d")
    try:
        with pool_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT u.user_id, u.full_name, u.tier, u.expires_at, ar.enabled,
                           ar.preferred_plan, ar.preferred_duration, ar.preferred_method
                    FROM users u
                    LEFT JOIN auto_renewal ar ON ar.user_id = u.user_id
                    WHERE u.expires_at IS NOT NULL
                    AND u.expires_at LIKE %s
                    AND u.tier != 'free'
                    AND u.is_banned = 0
                """, (f"{alert_date}%",))
                rows = cur.fetchall()

        for uid2, name, tier, exp, ar_en, ar_plan, ar_dur, ar_meth in rows:
            is_ar = get_lang(uid2) == "ar"
            is_rtl = get_lang(uid2) in RTL_LANGS
            try:
                text = (
                    f"⏰ *{'اشتراكك على وشك الانتهاء!' if is_ar else 'Your subscription is expiring!'}*\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━\n"
                    f"📦 {'الباقة' if is_ar else 'Plan'}: *{tier}*\n"
                    f"📅 {'ينتهي' if is_ar else 'Expires'}: `{str(exp)[:10]}`\n\n"
                    f"{'جدد الآن للاستمرار!' if is_ar else 'Renew now to keep access!'}"
                )
                kb = InlineKeyboardMarkup([
                    [InlineKeyboardButton(
                        "🔄 " + ("تجديد الآن" if is_ar else "Renew Now"),
                        callback_data="user_subscribe"
                    )],
                    [InlineKeyboardButton(
                        "⚙️ " + ("إعدادات التجديد" if is_ar else "Renewal Settings"),
                        callback_data="my_autorenewal"
                    )],
                ])
                await app.bot.send_message(chat_id=uid2, text=text,
                                           parse_mode="Markdown", reply_markup=kb)
            except Exception:
                pass
    except Exception as e:
        log.error(f"check_auto_renewals: {e}")

# ════════════════════════════════════════════
#   RTL SUPPORT
# ════════════════════════════════════════════
def rtl_wrap(text: str, lang: str) -> str:
    """Wrap text with RTL marker if language is RTL."""
    if lang in RTL_LANGS:
        return "\u200F" + text  # Right-to-left mark
    return text

def get_direction_emoji(lang: str) -> str:
    return "◀️" if lang in RTL_LANGS else "▶️"

# ════════════════════════════════════════════
#   QUICK SEARCH SHORTCUTS
# ════════════════════════════════════════════
# Shortcuts: user types prefix to trigger quick search
SEARCH_SHORTCUTS = {
    "e:":  "email",    # e:gmail.com
    "p:":  "phone",    # p:+201234
    "d:":  "domain",   # d:google.com
    "u:":  "username", # u:john_doe
    "pw:": "password", # pw:123456
    "ip:": "ip",       # ip:192.168.1.1
}

def detect_shortcut(text: str):
    """Returns (stype, keyword) if text starts with a shortcut, else None."""
    for prefix, stype in SEARCH_SHORTCUTS.items():
        if text.lower().startswith(prefix):
            keyword = text[len(prefix):].strip()
            if keyword:
                return stype, keyword
    return None, None

# ════════════════════════════════════════════
#   UI: AUTO-RENEWAL SETTINGS SCREEN
# ════════════════════════════════════════════
async def show_autorenewal_settings(update, context, query=None):
    uid   = update.effective_user.id
    is_ar = get_lang(uid) == "ar"
    ar    = get_auto_renewal(uid)
    u     = get_user_cached(uid)
    tier  = u[3] if u else "free"

    status_emoji = "✅" if ar.get("enabled") else "❌"
    text = (
        f"🔄 *{'التجديد التلقائي' if is_ar else 'Auto-Renewal'}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"{'الحالة' if is_ar else 'Status'}: {status_emoji} "
        f"{'مفعّل' if ar.get('enabled') else 'معطّل' if is_ar else 'Enabled' if ar.get('enabled') else 'Disabled'}\n"
    )
    if ar.get("plan"):
        text += f"📦 {'الباقة' if is_ar else 'Plan'}: `{ar['plan']}`\n"
    if ar.get("duration"):
        text += f"⏱ {'المدة' if is_ar else 'Duration'}: `{ar['duration']}`\n"
    if ar.get("method"):
        text += f"💳 {'الطريقة' if is_ar else 'Method'}: `{ar['method']}`\n"
    text += (
        f"\n{'سيتم إخطارك قبل ' + str(AUTO_RENEW_DAYS_BEFORE) + ' أيام من انتهاء الاشتراك.' if is_ar else f'You will be notified {AUTO_RENEW_DAYS_BEFORE} days before expiry.'}"
    )

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(
            "✅ " + ("تفعيل" if is_ar else "Enable"),  callback_data="autorenewal_on"),
         InlineKeyboardButton(
            "❌ " + ("تعطيل" if is_ar else "Disable"), callback_data="autorenewal_off")],
        [InlineKeyboardButton("⚙️ " + ("إعداد الباقة والطريقة" if is_ar else "Set Plan & Method"),
                              callback_data="autorenewal_setup")],
        [InlineKeyboardButton("🔙 Back", callback_data="user_home")],
    ])
    q = query or (update.callback_query if update.callback_query else None)
    if q:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

# ════════════════════════════════════════════
#   UI: ADMIN COUPON MANAGEMENT SCREEN
# ════════════════════════════════════════════
async def show_admin_coupons(update, context, query=None):
    uid = update.effective_user.id
    if not is_admin(uid):
        return
    coupons = list_coupons(active_only=False)
    if not coupons:
        text = "🏷 *No coupons created yet.*"
    else:
        text = f"🏷 *Coupons* — {len(coupons)}\n\n"
        for cid, code, ctype, val, mx, used, vuntil, plan_r, active in coupons:
            status = "✅" if active else "❌"
            exp = str(vuntil)[:10] if vuntil else "∞"
            text += f"{status} `{code}` — {ctype} {val} | {used}/{mx} | exp:{exp}\n"
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Create Coupon", callback_data="adm_create_coupon")],
        [InlineKeyboardButton("🗑️ Delete Coupon", callback_data="adm_delete_coupon")],
        [InlineKeyboardButton("🔙 Back",          callback_data="adm_home")],
    ])
    q = query or (update.callback_query if update.callback_query else None)
    if q:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)


if __name__ == "__main__":
    main()
